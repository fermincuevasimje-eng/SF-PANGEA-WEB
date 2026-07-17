import streamlit as st
import pandas as pd
import numpy as np
from scipy.spatial.distance import cdist
import re
import unicodedata
import simplekml
import io
import requests
import time
import os
import json
import base64
from datetime import datetime, timedelta, timezone
import gspread
from google.oauth2.service_account import Credentials
from streamlit_gsheets import GSheetsConnection
from openpyxl.styles import PatternFill
import config
import data_manager

# --- 1. CONFIGURACIÓN E INTERFAZ (MARCA DE AGUA SF) ---
st.set_page_config(page_title="SF PANGEA VPLUS ULTRA", layout="wide")

st.markdown(
    """
    <style>
    .main::before {
        content: "SF";
        position: fixed;
        top: 50%;
        left: 50%;
        transform: translate(-50%, -50%) rotate(-45deg);
        font-size: 25vw;
        color: rgba(0, 0, 0, 0.07);
        z-index: -1;
        pointer-events: none;
        font-weight: bold;
    }
    [data-testid="stMetricValue"] {
        font-size: 28px;
        color: #1f4e78;
    }
    </style>
    """,
    unsafe_allow_html=True
)

BASE_COORDS = (19.291395219739588, -99.63555838631413)
URL_DB = "https://docs.google.com/spreadsheets/d/14_fewol5DiFXoiO102wviiWR08Lw3PKHzEjSbMwxUm8/edit?gid=0#gid=0"
HOJA_PRINCIPAL = "Sheet1"
HOJA_PAPELERA = "Trash"

# --- 1.5 CATÁLOGO MAESTRO ACTUALIZADO ---
CATALOGO_MAESTRO = {
    "ADOLFO LOPEZ MATEOS": ['PARQUES NACIONALES I', 'MIGUEL HIDALGO  (CORRALITOS)', 'PARQUES NACIONALES  II'],
    "ARBOL DE LAS MANITAS": ['ZOPILOCALCO SUR', 'ZOPILOCALCO NORTE', 'LOMAS ALTAS', 'HUITZILA Y DOCTORES', 'NIÑOS HEROES (PENSIONES)'],
    "BARRIO TRADICIONALES": ['SANTA BARBARA', 'EL COPORO', 'LA RETAMA', 'SAN MIGUEL APINAHUISCO', 'UNION', 'SAN LUIS OBISPO'],
    "CACALOMACAN": ['CENTRO', 'RANCHO SAN MIGUEL ZACANGO', 'SAGRADO CORAZON', 'EL ARENAL'],
    "CALIXTLAHUACA": ['SAN FRANCISCO DE ASIS', 'ZONA ARQUEOLOGICA', 'EL CALVARIO', 'PALMILLAS'],
    "CAPULTITLAN": ['SAN ISIDRO LABRADOR', 'PASEOS DEL  VALLE', 'SAN JUDAS TADEO', 'LA SOLEDAD', 'LOS PINOS', 'GUADALUPE'],
    "CENTRO HISTORICO": ['CENTRO', 'SANTA CLARA', '5 DE MAYO', 'FRANCISCO MURGUIA (EL RANCHITO)', 'LA MERCED ( ALAMEDA)'],
    "CERRILLO VISTA HERMOSA": ['EL CERRILLO', 'EL EMBARCADERO'],
    "CIUDAD UNIVERSITARIA": ['PLAZAS DE SAN BUENAVENTURA', 'SAN BERNARDINO', 'VICENTE GUERRERO'],
    "COLON": ['COLON Y CIPRES I', 'COLON Y CIPRES II', 'ISIDRO FABELA PRIMERA SECCION', 'ISIDRO FABELA SEGUNDA SECCION', 'RANCHO DOLORES'],
    "DEL PARQUE": ['DEL PARQUE   I', 'DEL PARQUE  II', 'LAZARO CARDENAS', 'AMPLIACION LAZARO CARDENAS', 'AZTECA'],
    "INDEPENDENCIA": ['REFORMA Y FERROCARRILES NACIONALES', 'METEORO', 'INDEPENDENCIA', 'LAS TORRES (CIENTIFICOS)', 'SAN JUAN BUENAVISTA'],
    "LA MAQUINITA": ['RANCHO LA MORA', 'LOS ANGELES', 'CARLOS HANK Y LOS FRAILES', 'GUADALUPE, CLUB JARDIN Y LA MAGDALENA', 'TLACOPA'],
    "METROPOLITANA": ['LAS PALOMAS', 'LAS MARGARITAS', 'RANCHO MAYA'],
    "MODERNA DE LA CRUZ": ['MODERNA DE LA CRUZ  I', 'MODERNA DE LA CRUZ  II', 'BOSQUES DE COLON'],
    "MORELOS": ['MORELOS 1A SECCION', 'MORELOS 2A SECCION', 'FEDERAL (ADOLFO LOPEZ MATEOS)'],
    "NUEVA OXTOTITLAN": ['NUEVA OXTOTITLAN  I', 'NUEVA OXTOTITLAN II'],
    "OCHO CEDROS": ['OCHO CEDROS  I', 'VILLA HOGAR', 'OCHO CEDROS  II', '8 CEDROS SEGUNDA SECCION'],
    "SAN ANDRES CUEXCONTITLAN": ['SAN ANDRES', 'LA CONCEPCION', 'SANTA ROSA', 'LA NATIVIDAD', 'EJIDO SAN DIEGO DE LOS PADRES CUEXCONTITLAN', 'SAN DIEGO DE LOS PADRES I', 'SAN DIEGO DE LOS PADRES II', 'JICALTEPEC  CUEXCONTITLAN', 'LOMA LA PROVIDENCIA', 'EJIDO DE LA Y', 'LA LOMA CUEXCONTITLAN'],
    "SAN ANTONIO BUENAVISTA": ['CAMINO REAL', 'JOSE MARIA HEREDIA', 'LOS ROSALES'],
    "SAN BUENAVENTURA": ['INSURGENTES', 'PENSADOR MEXICANO', 'ALAMEDA 2000', 'CULTURAL', 'DEL DEPORTE', 'GUADALUPE'],
    "SAN CAYETANO DE MORELOS": ['SAN CAYETANO', 'CERRILLO PIEDRAS BLANCAS'],
    "SAN CRISTOBAL HUICHOCHITLAN": ['SAN GABRIEL', 'SAN JOSE GUADALUPE HUICHOCHITLAN', 'LA CONCEPCION', 'LA TRINIDAD I', 'LA TRINIDAD II', 'SAN SALVADOR II', 'SAN SALVADOR I'],
    "SAN FELIPE TLALMIMILOLPAN": ['CENTRO', 'EL CALVARIO', 'JARDINES DE SAN PEDRO', 'LA CURVA', 'LOS ALAMOS', 'LA VENTA', 'EL FRONTON', 'DEL PANTEON'],
    "SAN JUAN TILAPA": ['CENTRO', 'LAZARO CARDENAS', 'EL DURAZNO', 'GUADALUPE'],
    "SAN LORENZO TEPALTITLAN": ['CENTRO', 'LAS FLORES', 'EL CHARCO', 'SAN ANGELIN', 'LA CRUZ COMALCO', 'SAN ISIDRO', 'DEL PANTEON', 'RINCON DE SAN LORENZO', 'LA LOMA', 'CELANESE', 'EL MOGOTE'],
    "SAN MARCOS YACHIHUACALTEPEC": ['NORTE', 'SUR'],
    "SAN MARTIN TOLTEPEC": ['SAN MARTIN', 'PASEOS DE SAN MARTIN', 'SAN ISIDRO', 'LA PALMA TOLTEPEC', 'SEBASTIAN LERDO DE TEJADA', 'EJIDO DE SAN MARCOS YACHIHUACALTEPEC'],
    "SAN MATEO OTZACATIPAN": ['PONIENTE   I', 'PONIENTE  I I', 'RANCHO SAN JOSE', 'CANALEJA', 'ORIENTE  I', 'ORIENTE  II', 'LA MAGDALENA OTZACATIPAN', 'SANTA CRUZ OTZACATIPAN', 'SAN JOSE GUADALUPE OTZACATIPAN', 'SAN DIEGO DE LOS PADRES OTZACATIPAN', 'SAN BLAS OTZACATIPAN', 'SAN NICOLAS TOLENTINO I', 'SAN NICOLAS TOLENTINO II', 'LA CRESPA', 'JARDINES DE LA CRESPA', 'GEOVILLAS ARBOLEDA', 'LA FLORESTA', 'GEOVILLAS DE LA INDEPENDENCIA', 'VICENTE LOMBARDO', 'ARBOLEDAS'],
    "SAN MATEO OXTOTITLAN": ['CENTRO', 'TLALNEPANTLA', 'ATOTONILCO', 'RINCON DEL PARQUE', 'NIÑOS HEROES I', 'NIÑOS HEROES  II', 'TIERRA Y LIBERTAD', 'PROTIMBOS', '20 DE NOVIEMBRE', '14 DE DICIEMBRE', 'EL TRIGO', 'SAN JORGE'],
    "SAN PABLO AUTOPAN": ['DE JESUS 1A  SECCION', 'STA MARIA TLACHALOYITA', 'PUEBLO NUEVO  I', 'PUEBLO NUEVO  II', 'SANTA CRUZ  I', 'SANTA CRUZ  II', 'DE JESUS 3A SECCION', 'DE JESUS 2A SECCION', 'OJO DE AGUA', 'AVIACION AUTOPAN', 'SAN CARLOS AUTOPAN', 'SAN DIEGO LINARES', 'SAN DIEGO', 'REAL DE SAN PABLO', 'XICALTEPEC', 'GALAXIA TOLUCA', 'JICALTEPEC AUTOPAN'],
    "SAN PEDRO TOTOLTEPEC": ['DEL CENTRO', 'MANZANA SUR', 'DEL PANTEON', 'GEOVILLAS', 'FRANCISCO I. MADERO', 'LA GALIA', 'NUEVA SAN FRANCISCO', 'SAN MIGUEL TOTOLTEPEC', 'BORDO DE LAS CANASTAS', 'SAN FRANCISCO TOTOLTEPEC', 'GUADALUPE TOTOLTEPEC', 'SAN BLAS TOTOLTEPEC', 'LA CONSTITUCION TOTOLTEPEC', 'ARROYO VISTA HERMOSA'],
    "SAN SEBASTIAN": ['VALLE VERDE Y TERMINAL', 'PROGRESO', 'IZCALLI IPIEM', 'SAN SEBASTIAN Y VERTICE', 'IZCALLI TOLUCA', 'SALVADOR SANCHEZ COLIN', 'COMISION FEDERAL DE ELECTRICIDAD', 'VALLE DON CAMILO'],
    "SANCHEZ": ['SOR JUANA INES DE LA CRUZ', 'ELECTRICISTAS LOCALES', 'LA TERESONA I', 'LA TERESONA  II', 'LA TERESONA   III', 'SECTOR POPULAR'],
    "SANTA ANA TLAPALTITLAN": ['16 DE SEPTIEMBRE', 'PINO SUAREZ', 'DEL PANTEON', 'INDEPENDENCIA', 'SANTA MARIA SUR', 'SANTA MARIA NORTE', 'BUENAVISTA'],
    "SANTA CRUZ ATZCAPOTZALTONGO": ['SANTA CRUZ SUR', 'SANTA CRUZ NORTE', 'EX HACIENDA LA MAGDALENA'],
    "SANTA MARIA DE LAS ROSAS": ['SANTA MARIA DE LAS ROSAS', 'NUEVA SANTA MARIA DE LAS ROSAS', 'UNIDAD VICTORIA', 'LA MAGDALENA', 'NUEVA SANTA MARIA', 'BENITO JUAREZ', 'EVA SAMANO DE LOPEZ MATEOS', 'EMILIANO ZAPATA'],
    "SANTA MARIA TOTOLTEPEC": ['CENTRO', 'EL COECILLO', 'HEROES', 'PASEO TOTOLTEPEC', 'EL OLIMPO', 'EL CARMEN TOTOLTEPEC'],
    "SANTIAGO MILTEPEC": ['MILTEPEC CENTRO', 'MILTEPEC SUR', 'MILTEPEC NORTE'],
    "SANTIAGO TLACOTEPEC": ['DEL CENTRO', 'SANTA MARIA', 'SHINGADE', 'CRISTO REY', 'EL CALVARIO', 'SANTA JUANITA', 'EL REFUGIO'],
    "SANTIAGO TLAXOMULCO": ['EL CALVARIO', 'LA PEÑA', 'JUNTA LOCAL DE CAMINOS'],
    "SAUCES": ['SAUCES I', 'SAUCES III', 'SAUCES IV', 'SAUCES VI', 'SAUCES V', 'VILLAS SANTIN I', 'VILLAS SANTIN II', 'FRANCISCO VILLA', 'SAUCES II'],
    "SEMINARIO 2 DE MARZO": ['SEMINARIO 4A SECCION I', 'SEMINARIO 4A SECCION II', 'HEROES 5 DE MAYO I', 'HEROES 5 DE MAYO II'],
    "SEMINARIO CONCILIAR": ['SEMINARIO EL PARQUE', 'SEMINARIO 3A SECCION', 'SEMINARIO 1A SECCION', 'SEMINARIO EL MODULO'],
    "SEMINARIO LAS TORRES": ['SEMINARIO SAN FELIPE DE JESUS', 'SEMINARIO 2A SECCION', 'SEMINARIO 5A SECCION'],
    "TECAXIC": ['TECAXIC ORIENTE', 'TECAXIC PONIENTE'],
    "TLACHALOYA": ['TLACHALOYA', 'BALBUENA', 'SAN CARLOS', 'SAN JOSE BUENAVISTA', 'DEL CENTRO', 'EL TEJOCOTE', 'SAN JOSE LA COSTA'],
    "UNIVERSIDAD": ['UNIVERSIDAD', 'CUAUHTEMOC', 'AMERICAS', 'ALTAMIRANO'],
}

MAPA_UTB_DEL = {utb: dl for dl, lista in CATALOGO_MAESTRO.items() for utb in lista}

# --- 1.6 INVENTARIO MAESTRO () ---
# --- 1.6 INVENTARIO MAESTRO () ---
STOCK_INICIAL = [
    {"ID": "MAT-01", "Material": "FOTOCELDA 220V", "Stock": 100, "Min": 10, "Unidad": "Piezas"},
    {"ID": "MAT-02", "Material": "CABLE 2+1 # 6 BOBINA DE 300", "Stock": 100, "Min": 10, "Unidad": "Metros"},
    {"ID": "MAT-03", "Material": "ALAMBRE RECOCIDO", "Stock": 100, "Min": 10, "Unidad": "Kilos"},
    {"ID": "MAT-04", "Material": "CINTA AISLAR", "Stock": 100, "Min": 10, "Unidad": "Piezas"},
    {"ID": "MAT-05", "Material": "FOTOCELDA 110 - 220 V", "Stock": 100, "Min": 10, "Unidad": "Piezas"},
    {"ID": "MAT-06", "Material": "CABLE TWH # 14 IUSA (CAJA DE 100)", "Stock": 100, "Min": 10, "Unidad": "Cajas"},
    {"ID": "MAT-07", "Material": "FLEJE 5/8” ROLLO DE 30", "Stock": 100, "Min": 10, "Unidad": "Metros"},
    {"ID": "MAT-08", "Material": "HEBILLA 5/8", "Stock": 100, "Min": 10, "Unidad": "Piezas"},
    {"ID": "MAT-09", "Material": "CINTA AISLAR IUSA", "Stock": 100, "Min": 10, "Unidad": "Piezas"},
    {"ID": "MAT-10", "Material": "MENSULA 1.20", "Stock": 100, "Min": 10, "Unidad": "Piezas"},
    {"ID": "MAT-11", "Material": "MENSULA 1.50", "Stock": 100, "Min": 10, "Unidad": "Piezas"},
    {"ID": "MAT-12", "Material": "CABLE 1+1 # 6 BOBINA DE 500", "Stock": 100, "Min": 10, "Unidad": "Metros"},
    {"ID": "MAT-13", "Material": "MANGUERA LED BLANCO FRIO (ROLLO 100 M.)", "Stock": 100, "Min": 10, "Unidad": "Metros"},
    {"ID": "MAT-14", "Material": "MANGUERA LED BLANCO CALIDO (ROLLO 100 M.)", "Stock": 100, "Min": 10, "Unidad": "Metros"},
    {"ID": "MAT-15", "Material": "CONECTOR CON DRIVER PARA MANGUERA LED", "Stock": 100, "Min": 10, "Unidad": "Piezas"},
    {"ID": "MAT-16", "Material": "MANGUERA LED VERDE (ROLLO 100 M.)", "Stock": 100, "Min": 10, "Unidad": "Metros"},
    {"ID": "MAT-17", "Material": "MANGUERA LED ROJO (ROLLO 100 M.)", "Stock": 100, "Min": 10, "Unidad": "Metros"},
    {"ID": "MAT-18", "Material": "CINCHOS PLASTICO USO RUDO DE 100 M.M", "Stock": 100, "Min": 10, "Unidad": "Piezas"},
    {"ID": "MAT-19", "Material": "CINCHOS PLASTICO USO RUDO DE 150 M.M", "Stock": 100, "Min": 10, "Unidad": "Piezas"},
    {"ID": "MAT-20", "Material": "FESTON GUIRNALDA VERDE SEGUNDA CALIDAD", "Stock": 100, "Min": 10, "Unidad": "Piezas"},
    {"ID": "MAT-21", "Material": "FESTON GUIRNALDA ROJO SEGUNDA CALIDAD", "Stock": 100, "Min": 10, "Unidad": "Piezas"},
    {"ID": "MAT-22", "Material": "FESTON GUIRNALDA BLANCO PRIMERA CALIDAD", "Stock": 100, "Min": 10, "Unidad": "Piezas"},
    {"ID": "MAT-23", "Material": "FESTON GUIRNALDA VERDE PRIMERA CALIDAD", "Stock": 100, "Min": 10, "Unidad": "Piezas"},
    {"ID": "MAT-24", "Material": "CABLE DUPLEX POT # 16", "Stock": 100, "Min": 10, "Unidad": "Metros"},
    {"ID": "MAT-25", "Material": "ALAMBRE GALVANIZADO # 14.5", "Stock": 100, "Min": 10, "Unidad": "Kilos"},
    {"ID": "MAT-26", "Material": "ALAMBRE GALVANIZADO # 16", "Stock": 100, "Min": 10, "Unidad": "Kilos"},
    {"ID": "MAT-27", "Material": "ACEITE DIRECCION HIDRAHULICA", "Stock": 100, "Min": 10, "Unidad": "Litros"},
    {"ID": "MAT-28", "Material": "ACEITE TRANSMICION STANDART", "Stock": 100, "Min": 10, "Unidad": "Litros"},
    {"ID": "MAT-29", "Material": "ACEITE MOTOR DIESEL", "Stock": 100, "Min": 10, "Unidad": "Litros"},
    {"ID": "MAT-30", "Material": "ACEITE HIDRAHULICO", "Stock": 100, "Min": 10, "Unidad": "Litros"},
    {"ID": "MAT-31", "Material": "ANTICONGELANTE ROSA", "Stock": 100, "Min": 10, "Unidad": "Litros"},
    {"ID": "MAT-32", "Material": "BOBINA PARA CONTACTOR", "Stock": 100, "Min": 10, "Unidad": "Piezas"},
    {"ID": "MAT-33", "Material": "CAMPANA DEL REFLECTOR ESTADIO", "Stock": 100, "Min": 10, "Unidad": "Piezas"},
    {"ID": "MAT-34", "Material": "FOTOCONTACTOR RECUPERADO", "Stock": 100, "Min": 10, "Unidad": "Piezas"},
    {"ID": "MAT-35", "Material": "LAMPARA LED VARIOS WATS", "Stock": 100, "Min": 10, "Unidad": "Piezas"},
    {"ID": "MAT-36", "Material": "LAMPARA FLUORECENTE VARIOS WATS", "Stock": 100, "Min": 10, "Unidad": "Piezas"},
    {"ID": "MAT-37", "Material": "FUSIBLE 100 A", "Stock": 100, "Min": 10, "Unidad": "Piezas"},
    {"ID": "MAT-38", "Material": "INTERRUPTOR DE SEGURIDAD REHABILITADO", "Stock": 100, "Min": 10, "Unidad": "Piezas"},
    {"ID": "MAT-39", "Material": "INTERRUPTOR TRIFASICO 200 A", "Stock": 100, "Min": 10, "Unidad": "Piezas"},
    {"ID": "MAT-40", "Material": "LAMPARA ADITIVO METALICO CERAMICO 140 W RECUPERADO", "Stock": 100, "Min": 10, "Unidad": "Piezas"},
    {"ID": "MAT-41", "Material": "LAMPARA CPO-TW (TORCION) 140 W", "Stock": 100, "Min": 10, "Unidad": "Piezas"},
    {"ID": "MAT-42", "Material": "LUMINARIO BELLOTA", "Stock": 100, "Min": 10, "Unidad": "Piezas"},
    {"ID": "MAT-43", "Material": "LUMINARIO DELTA", "Stock": 100, "Min": 10, "Unidad": "Piezas"},
    {"ID": "MAT-44", "Material": "LUMINARIO LED PUNTA DE POSTE", "Stock": 100, "Min": 10, "Unidad": "Piezas"},
    {"ID": "MAT-45", "Material": "LUMINARIO AMC VARIOS W COMPLETO", "Stock": 100, "Min": 10, "Unidad": "Piezas"},
    {"ID": "MAT-46", "Material": "LUMINARIO POSEIDON 2.75 M.", "Stock": 100, "Min": 10, "Unidad": "Piezas"},
    {"ID": "MAT-47", "Material": "POSTE POSEIDON 2.75 M.", "Stock": 100, "Min": 10, "Unidad": "Piezas"},
    {"ID": "MAT-48", "Material": "MÉNSULA 1.20", "Stock": 100, "Min": 10, "Unidad": "Piezas"},
    {"ID": "MAT-49", "Material": "MÉNSULA 1.80 M. RECUPERADA", "Stock": 100, "Min": 10, "Unidad": "Piezas"},
    {"ID": "MAT-50", "Material": "PANEL SOLAR N/F", "Stock": 100, "Min": 10, "Unidad": "Piezas"},
    {"ID": "MAT-51", "Material": "INTERRUPTOR TERMOMAGNETICO VARIOS A", "Stock": 100, "Min": 10, "Unidad": "Piezas"},
    {"ID": "MAT-52", "Material": "POSTE CON BASE VARIAS MEDIDAS", "Stock": 100, "Min": 10, "Unidad": "Piezas"},
    {"ID": "MAT-53", "Material": "POSTE SIN BASE VARIAS MEDIDAS", "Stock": 100, "Min": 10, "Unidad": "Piezas"},
    {"ID": "MAT-54", "Material": "POSTE RIZO 8", "Stock": 100, "Min": 10, "Unidad": "Piezas"},
    {"ID": "MAT-55", "Material": "REFLECTOR LED 200 W", "Stock": 100, "Min": 10, "Unidad": "Piezas"},
    {"ID": "MAT-56", "Material": "REFLECTOR ESTADIO", "Stock": 100, "Min": 10, "Unidad": "Piezas"},
    {"ID": "MAT-57", "Material": "REFLECTOR LED RGB 30 W", "Stock": 100, "Min": 10, "Unidad": "Piezas"},
    {"ID": "MAT-58", "Material": "REFLECTOR LED RGB 50 W", "Stock": 100, "Min": 10, "Unidad": "Piezas"},
    {"ID": "MAT-59", "Material": "REFLECTOR LED 500W DE USO", "Stock": 100, "Min": 10, "Unidad": "Piezas"},
    {"ID": "MAT-60", "Material": "REFLECTOR LED 300 Y 400 W RECUPERADO", "Stock": 100, "Min": 10, "Unidad": "Piezas"},
    {"ID": "MAT-61", "Material": "CEMENTO GRIS CPC 30", "Stock": 100, "Min": 15, "Unidad": "Bultos"},
    {"ID": "MAT-62", "Material": "GRAVA TRITURADA", "Stock": 100, "Min": 5, "Unidad": "Metro Cúbico"},
    {"ID": "MAT-63", "Material": "ARENA DE MINA", "Stock": 100, "Min": 5, "Unidad": "Metro Cúbico"},
    {"ID": "MAT-64", "Material": "ABRAZADERA DE U UNICANAL 2\"", "Stock": 100, "Min": 10, "Unidad": "Piezas"},
    {"ID": "MAT-65", "Material": "ABRAZADERA DE U UNICANAL 3\"", "Stock": 100, "Min": 10, "Unidad": "Piezas"},
    {"ID": "MAT-66", "Material": "ABRAZADERA PARA POSTE DE 5\"", "Stock": 100, "Min": 5, "Unidad": "Piezas"},
    {"ID": "MAT-67", "Material": "ABRAZADERA PARA POSTE DE 6\"", "Stock": 100, "Min": 5, "Unidad": "Piezas"},
    {"ID": "MAT-68", "Material": "ABRAZADERA SIN FIN VARIAS MEDIDAS", "Stock": 100, "Min": 15, "Unidad": "Piezas"},
    {"ID": "MAT-69", "Material": "ABRAZADERA UNICANAL 1/2\"", "Stock": 100, "Min": 15, "Unidad": "Piezas"},
    {"ID": "MAT-70", "Material": "AISLADOR DE PORCELANA CORNETA", "Stock": 100, "Min": 10, "Unidad": "Piezas"},
    {"ID": "MAT-71", "Material": "AISLADOR TIPO CARRETE 2\"", "Stock": 100, "Min": 10, "Unidad": "Piezas"},
    {"ID": "MAT-72", "Material": "ANCLA PARA POSTE", "Stock": 100, "Min": 5, "Unidad": "Piezas"},
    {"ID": "MAT-73", "Material": "APAGADOR SENCILLO QUESITO", "Stock": 100, "Min": 10, "Unidad": "Piezas"},
    {"ID": "MAT-74", "Material": "ARANDELA PRESION 1/2\"", "Stock": 100, "Min": 20, "Unidad": "Piezas"},
    {"ID": "MAT-75", "Material": "ARANDELA PLANA 1/2\"", "Stock": 100, "Min": 20, "Unidad": "Piezas"},
    {"ID": "MAT-76", "Material": "ARMADURA PARA LUMINARIA LED", "Stock": 100, "Min": 5, "Unidad": "Piezas"},
    {"ID": "MAT-77", "Material": "ARRANCADOR PARA LAMPARA DE 400W", "Stock": 100, "Min": 10, "Unidad": "Piezas"},
    {"ID": "MAT-78", "Material": "BASE PARA FOTOCELDA", "Stock": 100, "Min": 15, "Unidad": "Piezas"},
    {"ID": "MAT-79", "Material": "BASE PARA POSTE POSEIDON", "Stock": 100, "Min": 5, "Unidad": "Piezas"},
    {"ID": "MAT-80", "Material": "BALASTRO ADITIVO METALICO 100W", "Stock": 100, "Min": 5, "Unidad": "Piezas"},
    {"ID": "MAT-81", "Material": "BALASTRO ADITIVO METALICO 150W", "Stock": 100, "Min": 5, "Unidad": "Piezas"},
    {"ID": "MAT-82", "Material": "BALASTRO ADITIVO METALICO 250W", "Stock": 100, "Min": 5, "Unidad": "Piezas"},
    {"ID": "MAT-83", "Material": "BALASTRO ADITIVO METALICO 400W", "Stock": 100, "Min": 5, "Unidad": "Piezas"},
    {"ID": "MAT-84", "Material": "BALASTRO VAPOR DE SODIO 70W", "Stock": 100, "Min": 5, "Unidad": "Piezas"},
    {"ID": "MAT-85", "Material": "BALASTRO VAPOR DE SODIO 100W", "Stock": 100, "Min": 5, "Unidad": "Piezas"},
    {"ID": "MAT-86", "Material": "BALASTRO VAPOR DE SODIO 150W", "Stock": 100, "Min": 5, "Unidad": "Piezas"},
    {"ID": "MAT-87", "Material": "BALASTRO VAPOR DE SODIO 250W", "Stock": 100, "Min": 5, "Unidad": "Piezas"},
    {"ID": "MAT-88", "Material": "BALASTRO VAPOR DE SODIO 400W", "Stock": 100, "Min": 5, "Unidad": "Piezas"},
    {"ID": "MAT-89", "Material": "BALASTRO DOBLE POTENCIA 250-400W", "Stock": 100, "Min": 5, "Unidad": "Piezas"},
    {"ID": "MAT-90", "Material": "BLOKE DE CONEXION DE PORCELANA", "Stock": 100, "Min": 20, "Unidad": "Piezas"},
    {"ID": "MAT-91", "Material": "BOLA DE CRISTAL PARA LUMINARIO AMC", "Stock": 100, "Min": 5, "Unidad": "Piezas"},
    {"ID": "MAT-92", "Material": "BOVEDA POLIURETANO INSPECCION", "Stock": 100, "Min": 5, "Unidad": "Piezas"},
    {"ID": "MAT-93", "Material": "BROCA PARA CONCRETO 1/2\"", "Stock": 100, "Min": 5, "Unidad": "Piezas"},
    {"ID": "MAT-94", "Material": "BROCA PARA METAL 1/4\"", "Stock": 100, "Min": 5, "Unidad": "Piezas"},
    {"ID": "MAT-95", "Material": "BROCHE PARA LUMINARIA AMC", "Stock": 100, "Min": 15, "Unidad": "Piezas"},
    {"ID": "MAT-96", "Material": "CABLE POTENCIA 3X10 AWG", "Stock": 100, "Min": 50, "Unidad": "Metros"},
    {"ID": "MAT-97", "Material": "CABLE POTENCIA 3X12 AWG", "Stock": 100, "Min": 50, "Unidad": "Metros"},
    {"ID": "MAT-98", "Material": "CABLE DESNUDO COBRE #8", "Stock": 100, "Min": 30, "Unidad": "Metros"},
    {"ID": "MAT-99", "Material": "CABLE ALUMINIO MULTICONDUCTOR 2+1 #4", "Stock": 100, "Min": 50, "Unidad": "Metros"},
    {"ID": "MAT-100", "Material": "CAJA DE REGISTRO DE ALUMINIO 3/4\"", "Stock": 100, "Min": 10, "Unidad": "Piezas"},
    {"ID": "MAT-101", "Material": "CAJA TERMOPLASTICA PARA INT. TERMOMAGNETICO", "Stock": 100, "Min": 10, "Unidad": "Piezas"},
    {"ID": "MAT-102", "Material": "CANALETA DE PVC 20X10MM", "Stock": 100, "Min": 15, "Unidad": "Tramos"},
    {"ID": "MAT-103", "Material": "CAPUCHON PARA CABLE VARIAS MEDIDAS", "Stock": 100, "Min": 50, "Unidad": "Piezas"},
    {"ID": "MAT-104", "Material": "CARRETE DE HILO NYLON PARA AMARRAR", "Stock": 100, "Min": 5, "Unidad": "Piezas"},
    {"ID": "MAT-105", "Material": "CARTUCHO FUSIBLE 30 A", "Stock": 100, "Min": 20, "Unidad": "Piezas"},
    {"ID": "MAT-106", "Material": "CARTUCHO FUSIBLE 60 A", "Stock": 100, "Min": 20, "Unidad": "Piezas"},
    {"ID": "MAT-107", "Material": "CHALUPA METALICA ESTANDAR", "Stock": 100, "Min": 15, "Unidad": "Piezas"},
    {"ID": "MAT-108", "Material": "CINTA ADHESIVA DOBLE CAPA ESPUMA", "Stock": 100, "Min": 10, "Unidad": "Piezas"},
    {"ID": "MAT-109", "Material": "CINTA METÁLICA PASACABLES 15M", "Stock": 100, "Min": 2, "Unidad": "Piezas"},
    {"ID": "MAT-110", "Material": "CLEMA DE PORCELANA 2 POLOS", "Stock": 100, "Min": 25, "Unidad": "Piezas"},
    {"ID": "MAT-111", "Material": "CLEMA DE PORCELANA 3 POLOS", "Stock": 100, "Min": 25, "Unidad": "Piezas"},
    {"ID": "MAT-112", "Material": "COLETA DE COBRE PARA TIERRA", "Stock": 100, "Min": 15, "Unidad": "Piezas"},
    {"ID": "MAT-113", "Material": "CONDUIT ACERO GALVANIZADO 1/2\"", "Stock": 100, "Min": 10, "Unidad": "Tramos"},
    {"ID": "MAT-114", "Material": "CONDUIT ACERO GALVANIZADO 3/4\"", "Stock": 100, "Min": 10, "Unidad": "Tramos"},
    {"ID": "MAT-115", "Material": "CONDUIT ACERO GALVANIZADO 1\"", "Stock": 100, "Min": 5, "Unidad": "Tramos"},
    {"ID": "MAT-116", "Material": "CONDUIT PVC PESADO 1/2\"", "Stock": 100, "Min": 15, "Unidad": "Tramos"},
    {"ID": "MAT-117", "Material": "CONDUIT PVC PESADO 3/4\"", "Stock": 100, "Min": 15, "Unidad": "Tramos"},
    {"ID": "MAT-118", "Material": "CONDUIT PVC PESADO 2\"", "Stock": 100, "Min": 10, "Unidad": "Tramos"},
    {"ID": "MAT-119", "Material": "CONECTOR LIQUIDTIGHT 1/2\"", "Stock": 100, "Min": 15, "Unidad": "Piezas"},
    {"ID": "MAT-120", "Material": "CONECTOR LIQUIDTIGHT 3/4\"", "Stock": 100, "Min": 15, "Unidad": "Piezas"},
    {"ID": "MAT-121", "Material": "CONECTOR RECTO CONDUIT 1/2\"", "Stock": 100, "Min": 25, "Unidad": "Piezas"},
    {"ID": "MAT-122", "Material": "CONECTOR RECTO CONDUIT 3/4\"", "Stock": 100, "Min": 25, "Unidad": "Piezas"},
    {"ID": "MAT-123", "Material": "CONTACTO RECEPTACULO DUPLEX", "Stock": 100, "Min": 15, "Unidad": "Piezas"},
    {"ID": "MAT-124", "Material": "CONTACTOR ELECTROMAGNETICO 30 A", "Stock": 100, "Min": 5, "Unidad": "Piezas"},
    {"ID": "MAT-125", "Material": "CONTACTOR ELECTROMAGNETICO 60 A", "Stock": 100, "Min": 5, "Unidad": "Piezas"},
    {"ID": "MAT-126", "Material": "COPLE ACERO GALVANIZADO 1/2\"", "Stock": 100, "Min": 20, "Unidad": "Piezas"},
    {"ID": "MAT-127", "Material": "COPLE ACERO GALVANIZADO 3/4\"", "Stock": 100, "Min": 20, "Unidad": "Piezas"},
    {"ID": "MAT-128", "Material": "COPLE PVC CONDUIT 1/2\"", "Stock": 100, "Min": 30, "Unidad": "Piezas"},
    {"ID": "MAT-129", "Material": "COPLE PVC CONDUIT 3/4\"", "Stock": 100, "Min": 30, "Unidad": "Piezas"},
    {"ID": "MAT-130", "Material": "CONTRA Y MONITOR 1/2\"", "Stock": 100, "Min": 40, "Unidad": "Juegos"},
    {"ID": "MAT-131", "Material": "CONTRA Y MONITOR 3/4\"", "Stock": 100, "Min": 40, "Unidad": "Juegos"},
    {"ID": "MAT-132", "Material": "CRUZ METALICA REFORZADA CR-4", "Stock": 100, "Min": 5, "Unidad": "Piezas"},
    {"ID": "MAT-133", "Material": "DISCO LIJA DE DESBASTE 7\"", "Stock": 100, "Min": 10, "Unidad": "Piezas"},
    {"ID": "MAT-134", "Material": "DISCO CORTE METAL 4 1/2\"", "Stock": 100, "Min": 15, "Unidad": "Piezas"},
    {"ID": "MAT-135", "Material": "DISCO CORTE METAL 7\"", "Stock": 100, "Min": 15, "Unidad": "Piezas"},
    {"ID": "MAT-136", "Material": "ELECTRODO PARA SOLDAR ARCO 1/8\"", "Stock": 100, "Min": 10, "Unidad": "Kilos"},
    {"ID": "MAT-137", "Material": "EMPAQUE NEOPRENO LUMINARIA AMC", "Stock": 100, "Min": 15, "Unidad": "Piezas"},
    {"ID": "MAT-138", "Material": "FUSIBLE AUTOMOTRIZ VARIOS AMPERES", "Stock": 100, "Min": 30, "Unidad": "Piezas"},
    {"ID": "MAT-139", "Material": "FUSIBLE DE EXPULSION TIPO APERTURA", "Stock": 100, "Min": 10, "Unidad": "Piezas"},
    {"ID": "MAT-140", "Material": "GANCHO C PARA SOPORTE DE AISLADOR", "Stock": 100, "Min": 15, "Unidad": "Piezas"},
    {"ID": "MAT-141", "Material": "GRANITO GRIS FINO DECORATIVO", "Stock": 100, "Min": 5, "Unidad": "Bultos"},
    {"ID": "MAT-142", "Material": "GUÍA DE ACERO PARA ALAMBRADO 30M", "Stock": 100, "Min": 2, "Unidad": "Piezas"},
    {"ID": "MAT-143", "Material": "IGNITOR UNIVERSAL 70-400W", "Stock": 100, "Min": 15, "Unidad": "Piezas"},
    {"ID": "MAT-144", "Material": "INTERRUPTOR TERMOMAGNETICO 1P 15A", "Stock": 100, "Min": 10, "Unidad": "Piezas"},
    {"ID": "MAT-145", "Material": "INTERRUPTOR TERMOMAGNETICO 1P 20A", "Stock": 100, "Min": 10, "Unidad": "Piezas"},
    {"ID": "MAT-146", "Material": "INTERRUPTOR TERMOMAGNETICO 1P 30A", "Stock": 100, "Min": 10, "Unidad": "Piezas"},
    {"ID": "MAT-147", "Material": "INTERRUPTOR TERMOMAGNETICO 2P 30A", "Stock": 100, "Min": 10, "Unidad": "Piezas"},
    {"ID": "MAT-148", "Material": "INTERRUPTOR TERMOMAGNETICO 2P 40A", "Stock": 100, "Min": 10, "Unidad": "Piezas"},
    {"ID": "MAT-149", "Material": "INTERRUPTOR TERMOMAGNETICO 2P 50A", "Stock": 100, "Min": 10, "Unidad": "Piezas"},
    {"ID": "MAT-150", "Material": "INTERRUPTOR TERMOMAGNETICO 2P 60A", "Stock": 100, "Min": 10, "Unidad": "Piezas"},
    {"ID": "MAT-151", "Material": "INTERRUPTOR TERMOMAGNETICO 3P 100A", "Stock": 100, "Min": 5, "Unidad": "Piezas"},
    {"ID": "MAT-152", "Material": "KIT ANCLAJE PARA POSTE DE ALUMBRADO", "Stock": 100, "Min": 4, "Unidad": "Juegos"},
    {"ID": "MAT-153", "Material": "LAMPARA VAPOR SODIO OVALADA 70W", "Stock": 100, "Min": 15, "Unidad": "Piezas"},
    {"ID": "MAT-154", "Material": "LAMPARA VAPOR SODIO TUBULAR 100W", "Stock": 100, "Min": 15, "Unidad": "Piezas"},
    {"ID": "MAT-155", "Material": "LAMPARA VAPOR SODIO TUBULAR 150W", "Stock": 100, "Min": 15, "Unidad": "Piezas"},
    {"ID": "MAT-156", "Material": "LAMPARA VAPOR SODIO TUBULAR 250W", "Stock": 100, "Min": 15, "Unidad": "Piezas"},
    {"ID": "MAT-157", "Material": "LAMPARA VAPOR SODIO TUBULAR 400W", "Stock": 100, "Min": 10, "Unidad": "Piezas"},
    {"ID": "MAT-158", "Material": "LÍQUIDO AFLOJATODO WD-40 GRANDE", "Stock": 100, "Min": 8, "Unidad": "Piezas"},
    {"ID": "MAT-159", "Material": "MANGUERA LICUATITE FLEXIBLE 1/2\"", "Stock": 100, "Min": 25, "Unidad": "Metros"},
    {"ID": "MAT-160", "Material": "MANGUERA LICUATITE FLEXIBLE 3/4\"", "Stock": 100, "Min": 25, "Unidad": "Metros"},
    {"ID": "MAT-161", "Material": "OPRESORES ALLEN PARA LUMINARIA", "Stock": 100, "Min": 50, "Unidad": "Piezas"},
    {"ID": "MAT-162", "Material": "PINTURA AEROSOL NEGRO BRILLANTE", "Stock": 100, "Min": 12, "Unidad": "Piezas"},
    {"ID": "MAT-163", "Material": "PINTURA AEROSOL GRIS PRIMARIO", "Stock": 100, "Min": 12, "Unidad": "Piezas"},
    {"ID": "MAT-164", "Material": "PINTURA ESMALTE ALKIDALIK GRIS 4L", "Stock": 100, "Min": 4, "Unidad": "Botes"},
    {"ID": "MAT-165", "Material": "PLACA ACERO 1/4\" PARA BASE POSTE", "Stock": 100, "Min": 4, "Unidad": "Piezas"},
    {"ID": "MAT-166", "Material": "PORTAFUSIBLE AEREO TIPO EN LINEA", "Stock": 100, "Min": 20, "Unidad": "Piezas"},
    {"ID": "MAT-167", "Material": "PUNTA PARA POSTE TUBULAR SINGLE", "Stock": 100, "Min": 5, "Unidad": "Piezas"},
    {"ID": "MAT-168", "Material": "PUNTA PARA POSTE TUBULAR DOBLE", "Stock": 100, "Min": 5, "Unidad": "Piezas"},
    {"ID": "MAT-169", "Material": "RELEVADOR DE CONTROL 12 A 220V", "Stock": 100, "Min": 10, "Unidad": "Piezas"},
    {"ID": "MAT-170", "Material": "RODILLO PARA PINTAR FELPA EXTRADURA", "Stock": 100, "Min": 6, "Unidad": "Piezas"},
    {"ID": "MAT-171", "Material": "SEGUETA PARA ARCO DE MANO 18T", "Stock": 100, "Min": 15, "Unidad": "Piezas"},
    {"ID": "MAT-172", "Material": "SEGUETA PARA ARCO DE MANO 24T", "Stock": 100, "Min": 15, "Unidad": "Piezas"},
    {"ID": "MAT-173", "Material": "SELLO CONDUIT DE EPOXICO DE 1/2\"", "Stock": 100, "Min": 10, "Unidad": "Piezas"},
    {"ID": "MAT-174", "Material": "SOLERA DE FIERRO 1\" X 1/8\" TRAMO", "Stock": 100, "Min": 5, "Unidad": "Tramos"},
    {"ID": "MAT-175", "Material": "SOPORTE UNICANAL PERFORADO 4X2 CM", "Stock": 100, "Min": 10, "Unidad": "Tramos"},
    {"ID": "MAT-176", "Material": "TAQUETE EXPANSIVO DE ACERO 1/2\"", "Stock": 100, "Min": 50, "Unidad": "Piezas"},
    {"ID": "MAT-177", "Material": "TAQUETE EXPANSIVO DE ACERO 3/8\"", "Stock": 100, "Min": 50, "Unidad": "Piezas"},
    {"ID": "MAT-178", "Material": "TAQUETE PLÁSTICO CON TORNILLO #10", "Stock": 100, "Min": 100, "Unidad": "Piezas"},
    {"ID": "MAT-179", "Material": "TERMINAL DE OJO PARA CABLE #6", "Stock": 100, "Min": 40, "Unidad": "Piezas"},
    {"ID": "MAT-180", "Material": "TERMINAL DE OJO PARA CABLE #8", "Stock": 100, "Min": 40, "Unidad": "Piezas"},
    {"ID": "MAT-181", "Material": "THINNER ESTÁNDAR LIMPIEZA 4L", "Stock": 100, "Min": 4, "Unidad": "Botes"},
    {"ID": "MAT-182", "Material": "TORNILLO GRADO 5 1/2\" X 1 1/2\"", "Stock": 100, "Min": 50, "Unidad": "Piezas"},
    {"ID": "MAT-183", "Material": "TORNILLO GRADO 5 1/2\" X 2\"", "Stock": 100, "Min": 50, "Unidad": "Piezas"},
    {"ID": "MAT-184", "Material": "TORNILLO MAQUINA G5 3/8\" X 1\"", "Stock": 100, "Min": 50, "Unidad": "Piezas"},
    {"ID": "MAT-185", "Material": "TUBO GALVANIZADO LICUATITE 1/2\"", "Stock": 100, "Min": 10, "Unidad": "Tramos"},
    {"ID": "MAT-186", "Material": "TUERCA HEXAGONAL ST 1/2\"", "Stock": 100, "Min": 100, "Unidad": "Piezas"},
    {"ID": "MAT-187", "Material": "TUERCA HEXAGONAL ST 3/8\"", "Stock": 100, "Min": 100, "Unidad": "Piezas"},
    {"ID": "MAT-188", "Material": "VALVULA DE SEGURIDAD PARA COMPRESOR", "Stock": 100, "Min": 2, "Unidad": "Piezas"},
    {"ID": "MAT-189", "Material": "VARILLA DE TIERRA COPPERWELD 1.5M", "Stock": 100, "Min": 10, "Unidad": "Piezas"},
    {"ID": "MAT-190", "Material": "ZANCO DE FIERRO REFORZADO INTERNO", "Stock": 100, "Min": 5, "Unidad": "Piezas"},
    {"ID": "MAT-191", "Material": "ZAPATA TERMINAL MECANICA AL/CU #2", "Stock": 100, "Min": 15, "Unidad": "Piezas"},
    {"ID": "MAT-192", "Material": "ZAPATA TERMINAL MECANICA AL/CU #4", "Stock": 100, "Min": 15, "Unidad": "Piezas"}
]

# --- 2. MOTOR LÓGICO MEJORADO ---
def get_real_route(coords_list):
    locs = ";".join([f"{lon},{lat}" for lat, lon in coords_list])
    url = f"http://router.project-osrm.org/route/v1/driving/{locs}?overview=full&geometries=geojson"
    try:
        r = requests.get(url, timeout=5) 
        if r.status_code == 200:
            data = r.json()
            if data.get('code') == 'Ok':
                return data['routes'][0]['geometry']['coordinates'], data['routes'][0]['distance'] / 1000
        return None, None
    except Exception: 
        return None, None

def normalizar_texto(texto):
    if not isinstance(texto, str): texto = str(texto)
    texto = "".join(c for c in unicodedata.normalize('NFD', texto) if unicodedata.category(c) != 'Mn')
    return texto.lower().strip()

def extraer_carga_robusta(punto_dict, tipo):
    d_letras = {'un ':'1 ','uno ':'1 ','una ':'1 ','dos ':'2 ','tres ':'3 ','cuatro ':'4 ','cinco ':'5 '}
    posibles_cols = ['ASUNTO', 'Observaciones', 'asunto', 'observaciones', 'Asunto', 'OBSERVACIONES']
    texto_fuente = ""
    for col in posibles_cols:
        if col in punto_dict and str(punto_dict[col]).strip() != "":
            texto_fuente = str(punto_dict[col]); break
    t_norm = normalizar_texto(texto_fuente)
    for p, n in d_letras.items(): t_norm = t_norm.replace(p, n)
    patrones = {
        'lum': r'(\d+)\s*(?:lampara|foco|reflector|arbotante|luminari[oa]|unidad|brazo|farol[a]?|punto de luz)s?',
        'poste': r'(\d+)\s*(?:poste|estructura|columna)s?',
        'cable': r'(\d+)\s*(?:metro|m)\.?\s*(?:de\s*)?(?:cable|conductor|linea|red|alambre|potencia)s?'
    }
    if tipo == 'cable':
        m = re.search(patrones['cable'], t_norm)
        if m: return int(m.group(1))
        if any(w in t_norm for w in ['cable', 'conductor', 'linea', 'red']):
            m_flex = re.search(r'(\d+)\s*(?:metro|m)s?', t_norm)
            return int(m_flex.group(1)) if m_flex else 0
        return 0
    m = re.search(patrones[tipo], t_norm)
    return int(m.group(1)) if m else 0

@st.cache_data
def load_massive_data(file, extension):
    df = pd.read_excel(file, engine='openpyxl') if extension == 'xlsx' else pd.read_csv(file)
    df = df.dropna(how='all').reset_index(drop=True)
    df = df[df.iloc[:, 0].astype(str).str.strip() != "nan"]
    df = df[df.iloc[:, 0].astype(str).str.strip() != ""]
    df = df[df.iloc[:, 0].notna()]
    df['del_norm'] = df.iloc[:, 22].astype(str).apply(normalizar_texto)
    df['utb_norm'] = df.iloc[:, 23].astype(str).apply(normalizar_texto)
    return df

# --- 3. AUTENTICACIÓN Y ESTADO ---
if "autenticado" not in st.session_state:
    st.session_state.autenticado, st.session_state.perfil, st.session_state.usuario_nombre = False, None, ""
if "menu" not in st.session_state:
    st.session_state.menu = "Inicio"
if "lista_bajas" not in st.session_state:
    st.session_state.lista_bajas = {}
if "input_key" not in st.session_state:
    st.session_state.input_key = 0
if "pasos_sf4" not in st.session_state:
    st.session_state.pasos_sf4 = [] 
if "edit_index" not in st.session_state:
    st.session_state.edit_index = -1

if "boveda_mmd" not in st.session_state:
    if os.path.exists("boveda_pangea.json"):
        with open("boveda_pangea.json", "r", encoding="utf-8") as f:
            st.session_state.boveda_mmd = json.load(f)
    else:
        st.session_state.boveda_mmd = {}

if not st.session_state.autenticado:
    st.title("🔐 Acceso SF PANGEA")
    col_u, col_p = st.columns(2)
    with col_u: u = st.text_input("Usuario")
    with col_p: p = st.text_input("Contraseña", type="password")
    if st.button("🚀 Ingresar", use_container_width=True):
        if u == "SF" and p == "1827":
            st.session_state.autenticado, st.session_state.perfil, st.session_state.usuario_nombre = True, "ADMIN", "SF_ADMIN"
            st.rerun()
        elif u == "GuaDAP" and p == "1111":
            st.session_state.autenticado, st.session_state.perfil, st.session_state.usuario_nombre = True, "CONSULTA", "GuaDAP"
            st.rerun()
        else:
            st.error("Acceso denegado")
else:
    # --- 4. SIDEBAR ---
    with st.sidebar:
        st.title("⚙️ Panel Operativo")
        st.write(f"**Usuario:** {st.session_state.usuario_nombre}")
        st.write("---")
        if st.button("🏠 Inicio", use_container_width=True): st.session_state.menu = "Inicio"
        if st.button("🚀 SF1-Generador de Rutas", use_container_width=True): st.session_state.menu = "SF1"
        if st.button("📁 SF2-Bajas", use_container_width=True): st.session_state.menu = "SF2"
        if st.button("📊 SF3-Captura y Métricas", use_container_width=True): st.session_state.menu = "SF3"
        if st.button("🏗️ SF4-Diseño de Procesos", use_container_width=True): st.session_state.menu = "SF4"
        if st.button("🛡️ SF5-Anti-Duplicados", use_container_width=True): st.session_state.menu = "SF5"
        if st.button("📦 SF6-Almacén e Inventario", use_container_width=True): st.session_state.menu = "SF6"
        st.write("---")
        if st.session_state.menu == "SF1":
            st.subheader("📊 Ajustes GdR Multi-Ruta")
            t_por_punto = st.slider("Minutos por Atención", 5, 60, 20)
            v_promedio = st.slider("Velocidad km/h", 10, 80, 25)
            max_puntos_ruta = st.slider("Puntos Máximos por Ruta (Segmentación):", 5, 50, 15) # <--- NUEVO CONTROL DE LA V24
            st.write("---")
        if st.button("🚪 Cerrar Sesión", use_container_width=True):
            st.session_state.autenticado = False
            st.rerun()
        st.info("SF PANGEA VPLUS ULTRA")

# --- 5. CUERPO LÓGICO ---
    if st.session_state.menu == "Inicio":
        st.title("👋 Bienvenido a SF PANGEA")
        st.info("Sistema de Gestión Operativa - Dirección de Alumbrado Público")
        st.write("Seleccione un módulo en el menú lateral para comenzar.")
        st.image("https://img.icons8.com/clouds/500/000000/map-marker.png", width=150)

    elif st.session_state.menu == "SF3":
        st.title(f"🛠️ Módulo SF3 - Gestión y Métricas")

        # --- CONEXIÓN DE SEGURIDAD CON GOOGLE SHEETS ---
        scope = ["https://www.googleapis.com/auth/spreadsheets"]
        creds_dict = {
            "type": st.secrets["connections"]["gsheets"]["type"],
            "project_id": st.secrets["connections"]["gsheets"]["project_id"],
            "private_key_id": st.secrets["connections"]["gsheets"]["private_key_id"],
            "private_key": st.secrets["connections"]["gsheets"]["private_key"].replace('\\n', '\n'),
            "client_email": st.secrets["connections"]["gsheets"]["client_email"],
            "client_id": st.secrets["connections"]["gsheets"]["client_id"],
            "auth_uri": st.secrets["connections"]["gsheets"]["auth_uri"],
            "token_uri": st.secrets["connections"]["gsheets"]["token_uri"],
            "auth_provider_x509_cert_url": st.secrets["connections"]["gsheets"]["auth_provider_x509_cert_url"],
            "client_x509_cert_url": st.secrets["connections"]["gsheets"]["client_x509_cert_url"]
        }
        creds = Credentials.from_service_account_info(creds_dict, scopes=scope)
        client = gspread.authorize(creds)
        SHEET_ID = "14_fewol5DiFXoiO102wviiWR08Lw3PKHzEjSbMwxUm8"
        
        try:
            sh = client.open_by_key(SHEET_ID)
            ws = sh.worksheet("Boveda_Bajas")
        except:
            ws = sh.get_worksheet(0)

        # --- SINCRONIZACIÓN AUTOMÁTICA ADAPTADA ---
        if "manual_db" not in st.session_state:
            st.session_state.manual_db = []
            try:
                valores_sheet = ws.get_all_values()
                if valores_sheet:
                    headers = [str(h).strip() for h in valores_sheet[0]]
                    if "ID Registro" in headers and "Datos Captura" in headers:
                        idx_id = headers.index("ID Registro")
                        idx_datos = headers.index("Datos Captura")
                        for row in valores_sheet[1:]:
                            if len(row) > idx_id and str(row[idx_id]).startswith("SF3-MET-"):
                                try:
                                    datos_raw = row[idx_datos]
                                    vale_parsed = json.loads(datos_raw) if isinstance(datos_raw, str) else datos_raw
                                    st.session_state.manual_db.append(vale_parsed)
                                except: pass
            except Exception as e:
                st.sidebar.warning(f"Sincronizando métricas: {e}")

        if "reset_key" not in st.session_state:
            st.session_state.reset_key = 0
        rk = st.session_state.reset_key

        # --- FORMULARIO ---
        with st.expander("📝 REGISTRAR NUEVA ATENCIÓN (FORMULARIO)", expanded=False):
            st.write("📍 **Paso 1: Ubicación**")
            col_geo1, col_geo2 = st.columns(2)
            with col_geo1:
                f_del = st.selectbox("Delegación", sorted(list(CATALOGO_MAESTRO.keys())), key=f"del_manual_{rk}")
            with col_geo2:
                opciones_utb_f = sorted(CATALOGO_MAESTRO.get(f_del, []))
                f_utb = st.selectbox("UTB", opciones_utb_f, key=f"utb_manual_{rk}")

            with st.form(key=f"form_sf3_core_{rk}", clear_on_submit=True):
                st.write("📝 **Paso 2: Detalles de la Atención**")
                c1, c2, c3 = st.columns([1, 1, 2])
                with c1: f_fecha = st.date_input("Fecha")
                with c2: f_ot = st.text_input("O.T.")
                with c3: f_folio = st.text_input("Folio / Ticket / IMEI")
                f_calle = st.text_input("Calle")

                m1, m2, m3, m4 = st.columns(4)
                with m1: f_rehab = st.number_input("7. Rehabilitación", min_value=0, step=1)
                with m2: f_manto = st.number_input("8. Mantenimiento", min_value=0, step=1)
                with m3: f_sust = st.number_input("9. Sustitución", min_value=0, step=1)
                with m4: f_ampli = st.number_input("10. Ampliación", min_value=0, step=1)

                f_obs = st.text_area("11. Observaciones")
                btn_guardar = st.form_submit_button("🚀 GUARDAR REGISTRO EN LISTA PERMANENTE", use_container_width=True)

                if btn_guardar:
                    payload = {"FECHA":f_fecha.strftime("%d/%m/%Y"), "OT":f_ot.upper(), "CALLE":f_calle.upper(), "DELEGACIÓN":f_del, "UTB":f_utb, "FOLIO":f_folio.upper(), "REHAB":int(f_rehab), "MANTO":int(f_manto), "SUST":int(f_sust), "AMPLI":int(f_ampli), "OBS":f_obs}
                    st.session_state.manual_db.append(payload)
                    ws.append_row([f"SF3-MET-{f_ot.upper()}", datetime.now().strftime("%d/%m/%Y %H:%M"), f_del, "", "", json.dumps(payload), ""])
                    st.session_state.reset_key += 1
                    st.rerun()

        # --- CONSOLA DE GESTIÓN (EDIT/ELIMINAR) ---
        with st.expander("🛠️ GESTIÓN DE REGISTROS (EDICIÓN/BORRADO)", expanded=False):
            if st.session_state.manual_db:
                sel = st.selectbox("Seleccionar:", [f"{r['OT']} | {r['FECHA']}" for r in st.session_state.manual_db])
                idx = [f"{r['OT']} | {r['FECHA']}" for r in st.session_state.manual_db].index(sel)
                col_e, col_d, col_b = st.columns(3)
                if col_e.button("✏️ Editar"): st.session_state.edit = idx
                if col_d.button("🗑️ Eliminar"): st.session_state.borra = idx
                if col_b.button("💥 Borrar Último"): st.session_state.borra_ult = True
                
                if "edit" in st.session_state:
                    reg = st.session_state.manual_db[st.session_state.edit]
                    nr = st.number_input("Rehab", value=reg['REHAB'])
                    if st.button("💾 Guardar Cambios"):
                        st.session_state.manual_db[st.session_state.edit].update({"REHAB": nr})
                        cell = ws.find(f"SF3-MET-{reg['OT']}", in_column=1)
                        if cell: ws.update_cell(cell.row, 6, json.dumps(st.session_state.manual_db[st.session_state.edit]))
                        del st.session_state.edit; st.rerun()
                if "borra" in st.session_state and st.checkbox("Confirmar"):
                    if st.button("Ejecutar"):
                        reg = st.session_state.manual_db.pop(st.session_state.borra)
                        cell = ws.find(f"SF3-MET-{reg['OT']}", in_column=1)
                        if cell: ws.delete_rows(cell.row)
                        del st.session_state.borra; st.rerun()
                if "borra_ult" in st.session_state and st.checkbox("Confirmar"):
                    if st.button("Ejecutar"):
                        last = st.session_state.manual_db.pop()
                        cell = ws.find(f"SF3-MET-{last['OT']}", in_column=1)
                        if cell: ws.delete_rows(cell.row)
                        del st.session_state.borra_ult; st.rerun()

        # --- FILTROS Y RESUMEN (TU LÓGICA ORIGINAL) ---
        st.markdown("---")
        up_cap = st.file_uploader("📂 Opcional: Cargar Archivo de Captura Masiva", type=["csv", "xlsx"], key="up_cap_sf3")
        if up_cap:
            try:
                ext = 'xlsx' if up_cap.name.endswith('.xlsx') else 'csv'
                st.session_state.masivo_pangea = load_massive_data(up_cap, ext)
            except Exception as e: st.error(f"Error: {e}")
        
        if "masivo_pangea" not in st.session_state: st.session_state.masivo_pangea = None
        
        col_f1, col_f2 = st.columns(2)
        if 'sel_del_val' not in st.session_state: st.session_state.sel_del_val = "TODAS"
        if 'sel_utb_val' not in st.session_state: st.session_state.sel_utb_val = "TODAS"
        
        sel_del = col_f1.selectbox("📍 Filtrar TODO por Delegación:", ["TODAS"] + sorted(list(CATALOGO_MAESTRO.keys())), key="sel_del_val")
        lista_utbs = ["TODAS"] + (sorted(CATALOGO_MAESTRO.get(sel_del, [])) if sel_del != "TODAS" else sorted(list(MAPA_UTB_DEL.keys())))
        sel_utb = col_f2.selectbox("🔍 Filtrar TODO por UTB:", lista_utbs, key="sel_utb_val")

        pieces_reporte = []
        if st.session_state.manual_db:
            df_m = pd.DataFrame(st.session_state.manual_db)
            if sel_del != "TODAS": df_m = df_m[df_m['DELEGACIÓN'] == sel_del]
            if sel_utb != "TODAS": df_m = df_m[df_m['UTB'] == sel_utb]
            if not df_m.empty: pieces_reporte.append(df_m)

        if st.session_state.masivo_pangea is not None:
            df_filt = st.session_state.masivo_pangea.copy()
            if sel_del != "TODAS": df_filt = df_filt[df_filt['del_norm'] == normalizar_texto(sel_del)]
            if sel_utb != "TODAS": df_filt = df_filt[df_filt['utb_norm'] == normalizar_texto(sel_utb)]
            if not df_filt.empty:
                df_av = df_filt.iloc[:, [4, 6, 15, 19, 22, 23, 29, 30, 31, 39]].copy()
                df_av.columns = ["FECHA", "OT", "FOLIO", "CALLE", "DELEGACIÓN", "UTB", "REHAB", "MANTO", "SUST", "AMPLI"]
                df_av["OBS"] = ""
                pieces_reporte.append(df_av)

        df_final = pd.concat(pieces_reporte, ignore_index=True) if pieces_reporte else pd.DataFrame()
        if not df_final.empty:
            for c in ["REHAB", "MANTO", "SUST", "AMPLI"]: df_final[c] = pd.to_numeric(df_final[c], errors='coerce').fillna(0).astype(int)
            st.metric("🔧 Totales", int(df_final["REHAB"].sum()), delta=f"Manto: {int(df_final['MANTO'].sum())}")
            st.dataframe(df_final.astype(str), use_container_width=True)

        # --- DESCARGAS ---
        if not df_final.empty:
            def generar_reporte(df, hoja):
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer: df.to_excel(writer, index=False, sheet_name=hoja)
                return output.getvalue()
            d1, d2, d3 = st.columns(3)
            d1.download_button("📂 Reporte MASIVO", generar_reporte(df_final, "MASIVO"), "REPORTE_MASIVO.xlsx")
            d2.download_button("📝 Reporte MANUAL", generar_reporte(df_final, "MANUAL"), "REPORTE_MANUAL.xlsx")
            d3.download_button("🚀 Reporte UNIFICADO", generar_reporte(df_final, "UNIFICADO"), "REPORTE_UNIFICADO.xlsx")
    elif st.session_state.menu == "SF2":
        st.title("📁 SF2 - Módulo de Baja de Folios")

        # --- 1. CONEXIÓN ---
        scope = ["https://www.googleapis.com/auth/spreadsheets"]
        creds_dict = {
            "type": st.secrets["connections"]["gsheets"]["type"],
            "project_id": st.secrets["connections"]["gsheets"]["project_id"],
            "private_key_id": st.secrets["connections"]["gsheets"]["private_key_id"],
            "private_key": st.secrets["connections"]["gsheets"]["private_key"].replace('\\n', '\n'),
            "client_email": st.secrets["connections"]["gsheets"]["client_email"],
            "client_id": st.secrets["connections"]["gsheets"]["client_id"],
            "auth_uri": st.secrets["connections"]["gsheets"]["auth_uri"],
            "token_uri": st.secrets["connections"]["gsheets"]["token_uri"],
            "auth_provider_x509_cert_url": st.secrets["connections"]["gsheets"]["auth_provider_x509_cert_url"],
            "client_x509_cert_url": st.secrets["connections"]["gsheets"]["client_x509_cert_url"]
        }
        creds = Credentials.from_service_account_info(creds_dict, scopes=scope)
        client = gspread.authorize(creds)
        SHEET_ID = "14_fewol5DiFXoiO102wviiWR08Lw3PKHzEjSbMwxUm8"
        
        try:
            sh = client.open_by_key(SHEET_ID)
            ws = sh.worksheet("Boveda_Bajas")
        except:
            ws = sh.get_worksheet(0)

        # --- 2. CARGA DE BÓVEDA FILTRADA ---
        if "db_bajas_historico" not in st.session_state:
            registros = ws.get_all_records()
            st.session_state.db_bajas_historico = {
                str(r.get("ID Registro")): r 
                for r in registros 
                if r.get("ID Registro") and str(r.get("ID Registro")).startswith("BAJA-")
            }

        st.write("Cargue el archivo original y digite los folios para generar el documento de cierre.")
        up_sf2 = st.file_uploader("Subir Archivo de Referencia (Excel/CSV)", type=["csv", "xlsx"], key="sf2_up")
        
        c_input, c_lista = st.columns([1, 1])
        
        with c_lista:
            tab_actual, tab_boveda = st.tabs(["📋 Captura Actual", "📂 Bóveda de Historial"])
            
            with tab_actual:
                st.subheader("Folios en proceso de baja")
                if "lista_bajas" in st.session_state and st.session_state.lista_bajas:
                    df_res = pd.DataFrame([{"Folio": rk, "Respuesta 127": v} for rk, v in st.session_state.lista_bajas.items()])
                    st.dataframe(df_res, use_container_width=True, hide_index=True)
                    
                    # --- EDICIÓN DE LISTA ---
                    with st.expander("✂️ Editar Lista Actual"):
                        folios_a_eliminar = st.multiselect("Seleccionar folios a quitar:", list(st.session_state.lista_bajas.keys()))
                        confirma_del = st.checkbox("🔐 Confirmar eliminación de seleccionados")
                        if st.button("🗑️ Eliminar seleccionados") and confirma_del:
                            for f in folios_a_eliminar:
                                del st.session_state.lista_bajas[f]
                            st.rerun()

                    if up_sf2 and st.button("📥 Generar Documento de Bajas", use_container_width=True, type="primary"):
                        try:
                            df_ref = pd.read_excel(up_sf2, dtype=str).fillna("") if up_sf2.name.endswith('.xlsx') else pd.read_csv(up_sf2, encoding='latin-1', dtype=str).fillna("")
                            id_col = next((c for c in df_ref.columns if any(p in str(c).upper() for p in ['FOLIO','TICKET','ID','IMEI'])), df_ref.columns[0])
                            
                            st.balloons()
                            mapa_limpio = {str(k).strip(): str(v) for k, v in st.session_state.lista_bajas.items()}
                            df_final = df_ref[df_ref[id_col].astype(str).isin(mapa_limpio.keys())].copy()
                            df_final['RESPUESTA 127'] = df_final[id_col].astype(str).map(mapa_limpio)
                            
                            output = io.BytesIO()
                            with pd.ExcelWriter(output, engine='openpyxl') as writer: df_final.to_excel(writer, index=False)
                            excel_data = output.getvalue()
                            
                            tz_mx = timezone(timedelta(hours=-6))
                            ahora = datetime.now(tz_mx)
                            id_reg = f"BAJA-{ahora.strftime('%Y%m%d-%H%M%S')}"
                            fecha_mx = ahora.strftime("%d/%m/%Y %H:%M:%S")
                            b64_str = base64.b64encode(excel_data).decode('utf-8').replace('\n', '').replace('\r', '')
                            ws.append_row([id_reg, fecha_mx, up_sf2.name, len(mapa_limpio), json.dumps(mapa_limpio), b64_str])
                            st.session_state.db_bajas_historico[id_reg] = {
                                "ID Registro": id_reg, 
                                "Fecha": fecha_mx, 
                                "Origen": up_sf2.name, 
                                "Folios": len(mapa_limpio), 
                                "Datos Captura": json.dumps(mapa_limpio), 
                                "Excel Base64": b64_str
                            }
                            
                            st.success(f"✅ ¡Guardado! ID: {id_reg}")
                            st.download_button("📗 Descargar Excel", data=excel_data, file_name=f"BAJAS_{up_sf2.name}", use_container_width=True)
                        except Exception as e: st.error(f"Error procesando: {e}")
                    
                    st.write("---")
                    seguro_limpiar = st.checkbox("🔐 Confirmar vaciado total", key="limpiar_seguro")
                    if st.button("🗑️ Limpiar TODO", disabled=not seguro_limpiar):
                        st.session_state.lista_bajas = {}
                        st.rerun()
                else: st.info("Esperando captura...")

            with tab_boveda:
                st.subheader("🗄️ Historial Permanente")
                if st.session_state.db_bajas_historico:
                    lista_tabla = [{"ID Registro": k, "Fecha": v.get("Fecha", "N/A"), "Origen": v.get("Origen", "N/A"), "Folios": v.get("Folios", 0)} for k, v in st.session_state.db_bajas_historico.items()]
                    df_h = pd.DataFrame(lista_tabla).sort_values(by="ID Registro", ascending=False)
                    id_rec = st.selectbox("Seleccione ID:", list(st.session_state.db_bajas_historico.keys())[::-1])
                    
                    def resaltar_fila(row):
                        color = '#d1e7dd' if row['ID Registro'] == id_rec else ''
                        return [f'background-color: {color}'] * len(row)

                    st.dataframe(df_h.style.apply(resaltar_fila, axis=1), use_container_width=True, hide_index=True)
                    
                    if id_rec:
                        data = st.session_state.db_bajas_historico[id_rec]
                        with st.expander(f"🔍 Detalle de folios ({id_rec})"):
                            # Auto-detección por contenido: identifica el JSON y el Excel sin importar el orden de la columna
                            v1 = data.get("Datos Captura") or data.get("Datos") or ""
                            v2 = data.get("Excel Base64") or data.get("Excel") or ""
                            
                            str_v1, str_v2 = str(v1).strip(), str(v2).strip()
                            raw_datos = str_v2 if str_v2.startswith(("{", "[")) else (str_v1 if str_v1.startswith(("{", "[")) else "{}")
                            raw_b64 = str_v1 if str_v1.startswith("UEsDB") else (str_v2 if str_v2.startswith("UEsDB") else "")
                            
                            try:
                                df_mostrar = None
                                
                                # 1. Intento por JSON de captura (Texto de respaldo rápido)
                                if raw_datos and raw_datos != "{}":
                                    try:
                                        datos_dict = json.loads(raw_datos)
                                        if isinstance(datos_dict, str):
                                            datos_dict = json.loads(datos_dict)
                                        
                                        if isinstance(datos_dict, dict) and datos_dict:
                                            df_mostrar = pd.DataFrame([{"Folio": k, "Detalle": v} for k, v in datos_dict.items()])
                                        elif isinstance(datos_dict, list) and datos_dict:
                                            df_mostrar = pd.DataFrame(datos_dict)
                                    except Exception:
                                        pass
                                
                                # 2. Intento por Reconstrucción Binaria (Decodifica el Excel Base64 guardado)
                                if df_mostrar is None and raw_b64:
                                    try:
                                        excel_bytes = base64.b64decode(raw_b64)
                                        df_mostrar = pd.read_excel(io.BytesIO(excel_bytes))
                                    except Exception as e_excel:
                                        st.warning(f"⚠️ Archivo Excel detectado, pero la vista previa falló: {e_excel}")
                                
                                # 3. Renderizado final en la interfaz
                                if df_mostrar is not None:
                                    st.dataframe(df_mostrar, use_container_width=True, hide_index=True)
                                elif raw_b64:
                                    st.info("Este registro contiene un archivo Excel. Usa el botón de descarga de abajo.")
                                else:
                                    st.info("No hay folios individuales capturados para este registro.")
                                    
                            except Exception as e:
                                st.error(f"Error al visualizar folios: {e}")
                                st.code(raw_datos)

                        # --- BOTÓN DE DESCARGA Y ADMINISTRACIÓN ---
                        if raw_b64:
                            try:
                                raw_b64 = str(raw_b64).replace('\n', '').replace('\r', '')
                                st.download_button("🔄 Descargar Excel", data=base64.b64decode(raw_b64), file_name=f"{id_rec}.xlsx", use_container_width=True)
                            except Exception as e: st.error(f"Error en descarga: {e}")
                        
                        st.markdown("---")
                        st.markdown("#### 🛠️ Administración del Registro")
                        col1, col2 = st.columns([1, 1])
                        with col1: 
                            if st.button("🔄 Retornar a Captura", key=f"retornar_{id_rec}"):
                                try:
                                    st.session_state.lista_bajas = json.loads(raw_datos) if isinstance(raw_datos, str) else raw_datos
                                    st.success("¡Datos cargados en Captura Actual!")
                                    st.rerun()
                                except Exception as e_ret:
                                    st.error("⚠️ El archivo original rebasó el límite de Google Sheets y quedó incompleto en la nube. No se puede retornar automáticamente.")
                        with col2:
                            confirmar_del = st.checkbox("🔐 Habilitar borrado permanente", key=f"chk_del_{id_rec}")
                            if confirmar_del:
                                if st.button("🗑️ BORRAR DE BÓVEDA", type="primary", key=f"borrar_{id_rec}"):
                                    try:
                                        cell = ws.find(id_rec, in_column=1)
                                        if cell: ws.delete_rows(cell.row)
                                        del st.session_state.db_bajas_historico[id_rec]
                                        st.success("¡Eliminado!")
                                        time.sleep(1)
                                        st.rerun()
                                    except Exception as e: st.error(f"Error al borrar: {e}")
                else: st.info("Bóveda vacía.")

        with c_input:
            st.subheader("⌨️ Captura de Folios")
            if "input_key" not in st.session_state: st.session_state.input_key = 0
            with st.form(key=f"form_bajas_{st.session_state.input_key}", clear_on_submit=True):
                col_f, col_ot = st.columns([1.2, 1.0])
                in_f_val = col_f.text_input("Digite Folio/Ticket/IMEi:", key=f"f_{st.session_state.input_key}")
                in_ot_val = col_ot.text_input("Orden de Trabajo (O.T.):", key=f"ot_{st.session_state.input_key}")
                c_cal, c_man = st.columns([1.1, 1.1])
                d_p = c_cal.date_input("Fecha (Calendario):", value=pd.Timestamp.now().date(), key=f"dt_p_{st.session_state.input_key}")
                d_m = c_man.text_input("Fecha (Copiar/Pegar):", placeholder="DD/MM/AAAA", key=f"dt_m_{st.session_state.input_key}")
                in_obs = st.text_input("Respuesta Libre / Observaciones:", max_chars=30, key=f"lb_{st.session_state.input_key}")
                if st.form_submit_button("➕ Agregar"):
                    if not up_sf2: 
                        st.error("⚠️ Sube un archivo de referencia.")
                    else:
                        df_ref = pd.read_excel(up_sf2, dtype=str).fillna("") if up_sf2.name.endswith('.xlsx') else pd.read_csv(up_sf2, encoding='latin-1', dtype=str).fillna("")
                        id_col = next((c for c in df_ref.columns if any(p in str(c).upper() for p in ['FOLIO','TICKET','ID','IMEI'])), df_ref.columns[0])
                        
                        # --- NUEVA LÓGICA DE VALIDACIÓN ---
                        folio_buscado = in_f_val.strip()
                        if folio_buscado in df_ref[id_col].astype(str).values:
                            fec = d_m.strip() if d_m.strip() else d_p.strftime("%d/%m/%Y")
                            ot_part = f"O.T. {in_ot_val.strip()}" if in_ot_val.strip() else ""
                            obs_val = in_obs.strip() if in_obs.strip() else "ATENDIDO"
                            componentes = [c for c in [ot_part, fec, obs_val] if c]
                            st.session_state.lista_bajas[folio_buscado] = " | ".join(componentes)
                            st.session_state.input_key += 1
                            st.rerun()
                        else:
                            st.error(f"⚠️ El folio '{folio_buscado}' no existe en el archivo '{up_sf2.name}'. Verifica el ID.")
    
    elif st.session_state.menu == "SF1":
        st.title("🚀 GdR V24 - Generador de Rutas Inteligente")
        tab1, tab1_multi, tab2, tab3 = st.tabs([
            "📍 Generador de Ruta Clásico (V23 Pro)", 
            "🚚 Nuevo Motor Multi-Ruta (Pro)", 
            "📂 Bitácora", 
            "🗑️ Papelera"
        ])

        # ==========================================
        # PESTAÑA 1: GENERADOR DE RUTA CLÁSICO (V23 PRO)
        # ==========================================
        with tab1:
            if st.session_state.perfil == "CONSULTA":
                st.warning("⚠️ Modo Consulta activo.")
            else:
                datos_vienen_de_sf5 = "df_transferido" in st.session_state and st.session_state.df_transferido is not None
                if datos_vienen_de_sf5:
                    st.info(f"📦 Usando datos procesados de: {st.session_state.nombre_archivo_transferido}")
                    if st.button("❌ Cancelar y subir otro archivo", key="cancel_c"):
                        st.session_state.df_transferido = None
                        st.rerun()
                    up_c = True
                else:
                    up_c = st.file_uploader("Subir Archivo (Excel/CSV) - Modo Clásico", type=["csv", "xlsx"], key="up_clasico")

                if up_c:
                    try:
                        if datos_vienen_de_sf5:
                            df_raw = st.session_state.df_transferido.copy()
                            up_name = st.session_state.nombre_archivo_transferido
                        else:
                            df_raw = pd.read_excel(up_c, dtype=str).fillna("") if up_c.name.endswith('.xlsx') else pd.read_csv(up_c, encoding='latin-1', dtype=str).fillna("")
                            up_name = up_c.name

                        if 'lat_aux' in df_raw.columns and 'lon_aux' in df_raw.columns and df_raw['lat_aux'].notna().any():
                            df_raw['lat_aux'] = pd.to_numeric(df_raw['lat_aux'], errors='coerce')
                            df_raw['lon_aux'] = pd.to_numeric(df_raw['lon_aux'], errors='coerce')
                        else:
                            col_coor = next((c for c in df_raw.columns if any(p in str(c).lower() for p in ['coordenadas', 'gps', 'ubicacion', 'coord'])), df_raw.columns[0])
                            def limpiar_y_extraer_coordenadas(valor):
                                texto = str(valor).lower().replace("latitude:", "").replace("longitude:", "")
                                numeros = re.findall(r'(-?\d+\.\d+)', texto)
                                if len(numeros) >= 2: return float(numeros[0]), float(numeros[1])
                                return None, None
                            res_coor = df_raw[col_coor].apply(limpiar_y_extraer_coordenadas)
                            df_raw['lat_aux'] = [r[0] for r in res_coor]
                            df_raw['lon_aux'] = [r[1] for r in res_coor]

                        id_col = next((c for c in df_raw.columns if any(p in str(c).upper() for p in ['FOLIO','TICKET','ID'])), df_raw.columns[0])
                        df_v = df_raw.dropna(subset=['lat_aux', 'lon_aux']).reset_index(drop=True)

                        if not df_v.empty:
                            pts = df_v.to_dict('records')
                            coords_base = np.array([BASE_COORDS])
                            coords_puntos = np.array([[p['lat_aux'], p['lon_aux']] for p in pts])
                            distancias_a_base = cdist(coords_base, coords_puntos)[0]
                            
                            idx_mas_lejano = np.argmax(distancias_a_base)
                            punto_inicial = pts.pop(idx_mas_lejano)
                            ruta_ordenada = [punto_inicial]
                            last_coord = (punto_inicial['lat_aux'], punto_inicial['lon_aux'])

                            while pts:
                                rest_coords = np.array([[p['lat_aux'], p['lon_aux']] for p in pts])
                                dist_al_ultimo = cdist([last_coord], rest_coords)[0]
                                dist_a_base = cdist(coords_base, rest_coords)[0]
                                puntuacion_ruta = dist_al_ultimo + (dist_a_base * 0.2)
                                
                                idx_proximo = np.argmin(puntuacion_ruta)
                                proximo_punto = pts.pop(idx_proximo)
                                ruta_ordenada.append(proximo_punto)
                                last_coord = (proximo_punto['lat_aux'], proximo_punto['lon_aux'])

                            route_coords = [BASE_COORDS] + [(p['lat_aux'], p['lon_aux']) for p in ruta_ordenada] + [BASE_COORDS]
                            geo_trazo, dist_real_km = get_real_route(route_coords)
                            if not dist_real_km: 
                                dist_real_km = (len(ruta_ordenada) + 1) * 1.3
                                st.warning("🛰️ Servidor de rutas fuera de línea. El KML usará trazo directo.")

                            tot_lums, tot_postes, tot_cable = 0, 0, 0
                            cols_orig = [c for c in df_raw.columns if c not in ['lat_aux', 'lon_aux']]
                            
                            for idx_r, p in enumerate(ruta_ordenada, 1):
                                p['Ruta_Asignada'] = "Ruta_Unica"
                                p['No_Ruta'] = idx_r
                                p['ID_Pangea_Nombre'] = p[id_col]
                                p['Cant_Luminarias'] = extraer_carga_robusta(p, 'lum') or (1 if extraer_carga_robusta(p, 'poste')==0 and extraer_carga_robusta(p, 'cable')==0 else 0)
                                p['Cant_Postes'] = extraer_carga_robusta(p, 'poste')
                                p['Cant_Cable_m'] = extraer_carga_robusta(p, 'cable')
                                p['Maps'] = f"https://www.google.com/maps?q={p['lat_aux']},{p['lon_aux']}"
                                
                                tot_lums += p['Cant_Luminarias']
                                tot_postes += p['Cant_Postes']
                                tot_cable += p['Cant_Cable_m']

                            min_totales = ((tot_lums + tot_postes) * t_por_punto) + (dist_real_km / v_promedio * 60)
                            t_estimado = f"{int(min_totales // 60)} h {int(min_totales % 60)} m"

                            st.subheader("📊 Resumen de Ruta Única (Clásica)")
                            mc1, mc2, mc3, mc4, mc5, mc6 = st.columns(6)
                            mc1.metric("📍 Puntos", len(ruta_ordenada))
                            mc2.metric("💡 Luminarias", tot_lums)
                            mc3.metric("🏗️ Postes", tot_postes)
                            mc4.metric("🧶 Cable", f"{tot_cable} m")
                            mc5.metric("🛣️ Distancia", f"{round(dist_real_km, 2)} km")
                            mc6.metric("⏱️ Tiempo Est.", t_estimado)

                            df_export_c = pd.DataFrame(ruta_ordenada)
                            cols_vits = ['No_Ruta', 'ID_Pangea_Nombre', 'Cant_Luminarias', 'Cant_Postes', 'Cant_Cable_m', 'Maps']
                            columnas_finales = cols_vits + [c for c in df_raw.columns if c != id_col and c not in ['lat_aux', 'lon_aux', 'ï»¿No_Ruta', 'Maps', 'Ruta_Asignada']]
                            df_export_c = df_export_c[columnas_finales]

                            st.dataframe(df_export_c, use_container_width=True, hide_index=True)

                            st.write("---")
                            cc1, cc2, cc3, cc4 = st.columns(4)
                            
                           # Convertir columnas críticas a numéricas para permitir sumatorias en Excel
                            df_export_excel = df_export_c.copy()
                            for col_num in ['Cant_Luminarias', 'Cant_Postes', 'Cant_Cable_m']:
                                df_export_excel[col_num] = pd.to_numeric(df_export_excel[col_num], errors='coerce').fillna(0).astype(int)

                            buf_xlsx_c = io.BytesIO()
                            with pd.ExcelWriter(buf_xlsx_c, engine='openpyxl') as writer:
                                df_export_excel.to_excel(writer, index=False, sheet_name='Ruta_Clasica_SF')
                                ws = writer.sheets['Ruta_Clasica_SF']
                                last_row = len(ruta_ordenada) + 1
                                res_row = last_row + 2
                                
                                ws.cell(row=res_row, column=2, value="--- RESUMEN OPERATIVO DINÁMICO ---")
                                ws.cell(row=res_row+1, column=1, value="Total Puntos:"); ws.cell(row=res_row+1, column=2, value=len(ruta_ordenada))
                                ws.cell(row=res_row+2, column=1, value="Total Luminarias:"); ws.cell(row=res_row+2, column=2, value=f"=SUM(C2:C{last_row})")
                                ws.cell(row=res_row+3, column=1, value="Total Postes:"); ws.cell(row=res_row+3, column=2, value=f"=SUM(D2:D{last_row})")
                                ws.cell(row=res_row+4, column=1, value="Total Cable:"); ws.cell(row=res_row+4, column=2, value=f"=SUM(E2:E{last_row})")
                                ws.cell(row=res_row+5, column=1, value="Distancia:"); ws.cell(row=res_row+5, column=2, value=f"{round(dist_real_km,2)} km")
                                
                                f_calc = f"ROUND(((B{res_row+2}+B{res_row+3})*{t_por_punto})+({round(dist_real_km,2)}/{v_promedio}*60),0)"
                                ws.cell(row=res_row+6, column=1, value="Tiempo Estimado:")
                                ws.cell(row=res_row+6, column=2, value=f'=INT({f_calc}/60) & " h " & MOD({f_calc},60) & " m"')
                                
                                fg, fa = PatternFill(start_color="E2E2E2", end_color="E2E2E2", fill_type="solid"), PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
                                for r in range(2, last_row + 1):
                                    if int(df_export_c.iloc[r-2]['Cant_Postes']) > 0:
                                        for cell in ws[r]: cell.fill = fg
                                    elif int(df_export_c.iloc[r-2]['Cant_Cable_m']) > 0:
                                        for cell in ws[r]: cell.fill = fa

                            cc1.download_button("📗 Excel Pro Dinámico", buf_xlsx_c.getvalue(), file_name=f"SF_CLASICA_{up_name}.xlsx", use_container_width=True)
                            
                            csv_buffer = io.StringIO()
                            df_export_c.to_csv(csv_buffer, index=False)
                            csv_buffer.write(f"\n--- RESUMEN OPERATIVO DINÁMICO ---\n")
                            csv_buffer.write(f"Total Puntos:,{len(ruta_ordenada)}\n")
                            csv_buffer.write(f"Total Luminarias:,{tot_lums}\n")
                            csv_buffer.write(f"Total Postes:,{tot_postes}\n")
                            csv_buffer.write(f"Total Cable:,{tot_cable} m\n")
                            csv_buffer.write(f"Distancia Total:,{round(dist_real_km,2)} km\n")
                            csv_buffer.write(f"Tiempo Estimado:,{t_estimado}\n")
                            cc2.download_button("📊 CSV Estático", csv_buffer.getvalue().encode('utf-8-sig'), file_name=f"SF_CLASICA_{up_name}.csv", use_container_width=True)

                            kml_c = simplekml.Kml()
                            folder_c = kml_c.newfolder(name=f"🚚 Ruta Única Clásica ({len(ruta_ordenada)} Pts)")
                            
                            for p in ruta_ordenada:
                                pnt = folder_c.newpoint(name=f"{p['ID_Pangea_Nombre']}", coords=[(p['lon_aux'], p['lat_aux'])])
                                h = "<![CDATA[<table border='1' style='width:300px; border-collapse:collapse; font-family:Arial; font-size:12px;'>"
                                h += "<tr><td bgcolor='#767171' colspan='2' align='center'><b style='color:white;'>DATOS DEL REPORTE</b></td></tr>"
                                for col in cols_orig:
                                    val = str(p.get(col, '')).strip()
                                    if val: h += f"<tr><td bgcolor='#F2F2F2'><b>{col}:</b></td><td>{val}</td></tr>"
                                h += "<tr><td bgcolor='#1F4E78' colspan='2' align='center'><b style='color:white;'>DESGLOSE OPERATIVO</b></td></tr>"
                                h += f"<tr><td bgcolor='#D9EAD3'><b>Punto de Ruta:</b></td><td>{p['No_Ruta']}</td></tr>"
                                h += f"<tr><td bgcolor='#D9EAD3'><b>Luminarias:</b></td><td>{p['Cant_Luminarias']}</td></tr>"
                                h += f"<tr><td bgcolor='#D9EAD3'><b>Postes:</b></td><td>{p['Cant_Postes']}</td></tr>"
                                h += f"<tr><td bgcolor='#D9EAD3'><b>Cable:</b></td><td>{p['Cant_Cable_m']} m</td></tr>"
                                h += "<tr><td bgcolor='#C00000' colspan='2' align='center'><b style='color:white;'>--- RESUMEN OPERATIVO DINÁMICO ---</b></td></tr>"
                                h += f"<tr><td><b>Total Puntos:</b></td><td>{len(ruta_ordenada)}</td></tr>"
                                h += f"<tr><td><b>Total Luminarias:</b></td><td>{tot_lums}</td></tr>"
                                h += f"<tr><td><b>Total Postes:</b></td><td>{tot_postes}</td></tr>"
                                h += f"<tr><td><b>Total Cable:</b></td><td>{tot_cable} m</td></tr>"
                                h += f"<tr><td><b>Distancia:</b></td><td>{round(dist_real_km,2)} km</td></tr>"
                                h += f"<tr><td><b>Tiempo Est.:</b></td><td>{t_estimado}</td></tr>"
                                h += "</table>]]>"
                                pnt.description = h

                            if geo_trazo:
                                ls = folder_c.newlinestring(name="TRAYECTO VIAL COMPLETO (BASE-RUTA-BASE)")
                                ls.coords = [(float(c[0]), float(c[1])) for c in geo_trazo]
                                ls.style.linestyle.width = 6
                                ls.style.linestyle.color = 'ff0000ff'
                            else:
                                ls = folder_c.newlinestring(name="TRAYECTO DIRECTO (SIN CALLES)")
                                ls.coords = [(float(c[1]), float(c[0])) for c in route_coords]
                                ls.style.linestyle.width = 4
                                ls.style.linestyle.color = 'ff00ffff'

                            cc3.download_button("🗺️ KML Maestro Clásico", kml_c.kml(), file_name=f"SF_CLASICA_{up_name}.kml", use_container_width=True)
                            cc4.link_button("🚀 My Maps", "https://www.google.com/maps/d/", use_container_width=True)

                            if st.button("💾 REGISTRAR RUTA CLÁSICA EN BITÁCORA", use_container_width=True, key="reg_c"):
                                try:
                                    conn = st.connection("gsheets", type=GSheetsConnection)
                                    hist = conn.read(spreadsheet=URL_DB, worksheet=HOJA_PRINCIPAL, ttl=0).dropna(how='all')
                                    info_j = f"Modo: Clásico, Pts: {len(ruta_ordenada)}, Lums: {tot_lums}, Cab: {tot_cable}m, Dist: {round(dist_real_km,1)}km"
                                    n_f = pd.DataFrame([{"Fecha": pd.Timestamp.now().strftime("%d/%m/%Y %H:%M"), "Nombre_Ruta": f"CLASICA_{up_name}", "Usuario_Generador": st.session_state.usuario_nombre, "Datos_JSON": info_j}])
                                    conn.update(spreadsheet=URL_DB, worksheet=HOJA_PRINCIPAL, data=pd.concat([hist, n_f], ignore_index=True))
                                    st.balloons(); st.success("¡Bitácora actualizada!")
                                except Exception as e: st.error(f"Error GSheets: {e}")
                        else:
                            st.error("No se pudieron extraer coordenadas válidas en Modo Clásico.")
                    except Exception as e: st.error(f"Error en Motor Clásico: {e}")



        # ==========================================
        # PESTAÑA 2: NUEVO MOTOR MULTI-RUTA (PRO)
        # ==========================================
        with tab1_multi:
            if st.session_state.perfil == "CONSULTA":
                st.warning("⚠️ Modo Consulta activo.")
            else:
                datos_vienen_de_sf5 = "df_transferido" in st.session_state and st.session_state.df_transferido is not None
                if datos_vienen_de_sf5:
                    st.info(f"📦 Usando datos procesados de: {st.session_state.nombre_archivo_transferido}")
                    if st.button("❌ Cancelar y subir otro archivo", key="cancel_m"):
                        st.session_state.df_transferido = None
                        st.rerun()
                    up_m = True
                else:
                    up_m = st.file_uploader("Subir Archivo (Excel/CSV) - Modo Multi-Ruta", type=["csv", "xlsx"], key="up_multiruta")

                if up_m:
                    try:
                        if datos_vienen_de_sf5:
                            df_raw = st.session_state.df_transferido.copy()
                            up_name = st.session_state.nombre_archivo_transferido
                        else:
                            df_raw = pd.read_excel(up_m, dtype=str).fillna("") if up_m.name.endswith('.xlsx') else pd.read_csv(up_m, encoding='latin-1', dtype=str).fillna("")
                            up_name = up_m.name

                        if 'lat_aux' in df_raw.columns and 'lon_aux' in df_raw.columns and df_raw['lat_aux'].notna().any():
                            df_raw['lat_aux'] = pd.to_numeric(df_raw['lat_aux'], errors='coerce')
                            df_raw['lon_aux'] = pd.to_numeric(df_raw['lon_aux'], errors='coerce')
                        else:
                            col_coor = next((c for c in df_raw.columns if any(p in str(c).lower() for p in ['coordenadas', 'gps', 'ubicacion', 'coord'])), df_raw.columns[0])
                            def limpiar_y_extraer_coordenadas(valor):
                                texto = str(valor).lower().replace("latitude:", "").replace("longitude:", "")
                                numeros = re.findall(r'(-?\d+\.\d+)', texto)
                                if len(numeros) >= 2: return float(numeros[0]), float(numeros[1])
                                return None, None
                            res_coor = df_raw[col_coor].apply(limpiar_y_extraer_coordenadas)
                            df_raw['lat_aux'] = [r[0] for r in res_coor]
                            df_raw['lon_aux'] = [r[1] for r in res_coor]

                        id_col = next((c for c in df_raw.columns if any(p in str(c).upper() for p in ['FOLIO','TICKET','ID'])), df_raw.columns[0])
                        df_v = df_raw.dropna(subset=['lat_aux', 'lon_aux']).reset_index(drop=True)

                        if not df_v.empty:
                            pts_restantes = df_v.to_dict('records')
                            lista_rutas_finales = []
                            contador_rutas = 1
                            coords_base = np.array([BASE_COORDS])

                            while len(pts_restantes) > 0:
                                rest_coords = np.array([[p['lat_aux'], p['lon_aux']] for p in pts_restantes])
                                distancias_a_base = cdist(coords_base, rest_coords)[0]
                                idx_mas_lejano = np.argmax(distancias_a_base)
                                
                                punto_inicial = pts_restantes.pop(idx_mas_lejano)
                                ruta_actual_puntos = [punto_inicial]
                                last_coord = (punto_inicial['lat_aux'], punto_inicial['lon_aux'])

                                while len(pts_restantes) > 0 and len(ruta_actual_puntos) < max_puntos_ruta:
                                    rest_coords_loop = np.array([[p['lat_aux'], p['lon_aux']] for p in pts_restantes])
                                    dist_al_ultimo = cdist([last_coord], rest_coords_loop)[0]
                                    dist_a_base_loop = cdist(coords_base, rest_coords_loop)[0]
                                    puntuacion_ruta = dist_al_ultimo + (dist_a_base_loop * 0.2)
                                    
                                    idx_proximo = np.argmin(puntuacion_ruta)
                                    proximo_punto = pts_restantes.pop(idx_proximo)
                                    ruta_actual_puntos.append(proximo_punto)
                                    last_coord = (proximo_punto['lat_aux'], proximo_punto['lon_aux'])

                                lista_rutas_finales.append({
                                    "id_ruta": f"Ruta_{contador_rutas}",
                                    "puntos": ruta_actual_puntos
                                })
                                contador_rutas += 1

                            st.success(f"📦 ¡Segmentación Multi-Ruta Exitosa! Se generaron **{len(lista_rutas_finales)} rutas independientes**.")
                            
                            kml = simplekml.Kml()
                            cols_orig = [c for c in df_raw.columns if c not in ['lat_aux', 'lon_aux']]
                            
                            excel_rutas_desglose = []
                            resumen_global_texto = ""
                            csv_multi_buffer = io.StringIO()
                            
                            tot_puntos_global, tot_lums_global, tot_postes_global, tot_cable_global, tot_dist_global = 0, 0, 0, 0, 0.0
                            metricas_por_ruta = {}

                            for r_info in lista_rutas_finales:
                                r_id = r_info["id_ruta"]
                                r_pts = r_info["puntos"]
                                
                                folder = kml.newfolder(name=f"🚚 {r_id} ({len(r_pts)} Pts)")
                                route_coords = [BASE_COORDS] + [(p['lat_aux'], p['lon_aux']) for p in r_pts] + [BASE_COORDS]
                                
                                geo_trazo, dist_real_km = get_real_route(route_coords)
                                time.sleep(0.3)
                                if not dist_real_km: dist_real_km = (len(r_pts) + 1) * 1.3
                                
                                r_lums, r_postes, r_cable = 0, 0, 0
                                for idx_r, p in enumerate(r_pts, 1):
                                    p['Ruta_Asignada'] = r_id
                                    p['No_Ruta'] = idx_r
                                    p['ID_Pangea_Nombre'] = p[id_col]
                                    p['Cant_Luminarias'] = extraer_carga_robusta(p, 'lum') or (1 if extraer_carga_robusta(p, 'poste')==0 and extraer_carga_robusta(p, 'cable')==0 else 0)
                                    p['Cant_Postes'] = extraer_carga_robusta(p, 'poste')
                                    p['Cant_Cable_m'] = extraer_carga_robusta(p, 'cable')
                                    p['Maps'] = f"https://www.google.com/maps?q={p['lat_aux']},{p['lon_aux']}"
                                    
                                    r_lums += p['Cant_Luminarias']
                                    r_postes += p['Cant_Postes']
                                    r_cable += p['Cant_Cable_m']

                                    # --- REPLICACIÓN DE VISTA PREMIUM CDATA EN EL KML MULTI-RUTA ---
                                    pnt = folder.newpoint(name=f"[{r_id}-#{idx_r}] {p['ID_Pangea_Nombre']}", coords=[(p['lon_aux'], p['lat_aux'])])
                                    h = "<![CDATA[<table border='1' style='width:300px; border-collapse:collapse; font-family:Arial; font-size:12px;'>"
                                    h += f"<tr><td bgcolor='#767171' colspan='2' align='center'><b style='color:white;'>DATOS DEL REPORTE ({r_id})</b></td></tr>"
                                    for col in cols_orig:
                                        val = str(p.get(col, '')).strip()
                                        if val: h += f"<tr><td bgcolor='#F2F2F2'><b>{col}:</b></td><td>{val}</td></tr>"
                                    h += "<tr><td bgcolor='#1F4E78' colspan='2' align='center'><b style='color:white;'>DESGLOSE OPERATIVO</b></td></tr>"
                                    h += f"<tr><td bgcolor='#D9EAD3'><b>Punto de Ruta:</b></td><td>{p['No_Ruta']}</td></tr>"
                                    h += f"<tr><td bgcolor='#D9EAD3'><b>Luminarias:</b></td><td>{p['Cant_Luminarias']}</td></tr>"
                                    h += f"<tr><td bgcolor='#D9EAD3'><b>Postes:</b></td><td>{p['Cant_Postes']}</td></tr>"
                                    h += f"<tr><td bgcolor='#D9EAD3'><b>Cable:</b></td><td>{p['Cant_Cable_m']} m</td></tr>"
                                    # ENCABEZADO CORREGIDO: Ahora jala el nombre dinámico del KML idéntico al del Excel Pro
                                    h += f"<tr><td bgcolor='#C00000' colspan='2' align='center'><b style='color:white;'>--- RESUMEN OPERATIVO DINÁMICO ({r_id}) ---</b></td></tr>"
                                    h += f"<tr><td><b>Total Puntos Ruta:</b></td><td>{len(r_pts)}</td></tr>"
                                    h += f"<tr><td><b>Total Luminarias:</b></td><td>{r_lums}</td></tr>"
                                    h += f"<tr><td><b>Total Postes:</b></td><td>{r_postes}</td></tr>"
                                    h += f"<tr><td><b>Total Cable:</b></td><td>{r_cable} m</td></tr>"
                                    h += f"<tr><td><b>Distancia Tramo:</b></td><td>{round(dist_real_km,2)} km</td></tr>"
                                    h += "</table>]]>"
                                    pnt.description = h

                                if geo_trazo:
                                    ls = folder.newlinestring(name=f"Trayecto Vial {r_id}")
                                    ls.coords = [(float(c[0]), float(c[1])) for c in geo_trazo]
                                    ls.style.linestyle.width = 6
                                    ls.style.linestyle.color = 'ff00cc00' 
                                else:
                                    ls = folder.newlinestring(name=f"Trayecto Directo {r_id}")
                                    ls.coords = [(float(c[1]), float(c[0])) for c in route_coords]
                                    ls.style.linestyle.width = 4
                                    ls.style.linestyle.color = 'ff00ffff'
                                
                                min_totales = ((r_lums + r_postes) * t_por_punto) + (dist_real_km / v_promedio * 60)
                                t_estimado_r = f"{int(min_totales // 60)} h {int(min_totales % 60)} m"

                                metricas_por_ruta[r_id] = {
                                    "puntos": len(r_pts), "distancia": round(dist_real_km, 2), "tiempo": t_estimado_r
                                }

                                tot_puntos_global += len(r_pts)
                                tot_lums_global += r_lums
                                tot_postes_global += r_postes
                                tot_cable_global += r_cable
                                tot_dist_global += dist_real_km
                                
                                resumen_global_texto += f"**• {r_id}:** {len(r_pts)} Pts | 💡 {r_lums} Lums | 🏗️ {r_postes} Postes | 🛣️ {round(dist_real_km,1)} km | ⏱️ {t_estimado_r}\n\n"
                                
                                df_temp_r = pd.DataFrame(r_pts)
                                cols_vits = ['Ruta_Asignada', 'No_Ruta', 'ID_Pangea_Nombre', 'Cant_Luminarias', 'Cant_Postes', 'Cant_Cable_m', 'Maps']
                                columnas_finales = cols_vits + [c for c in df_raw.columns if c != id_col and c not in ['lat_aux', 'lon_aux', 'ï»¿No_Ruta', 'Maps', 'Ruta_Asignada']]
                                df_temp_r = df_temp_r[columnas_finales]
                                excel_rutas_desglose.append(df_temp_r)

                            st.subheader("📊 Resumen Global Consolidado")
                            mg1, mg2, mg3, mg4, mg5 = st.columns(5)
                            mg1.metric("🚚 Total Rutas", len(lista_rutas_finales))
                            mg2.metric("📍 Total Puntos", tot_puntos_global)
                            mg3.metric("💡 Total Luminarias", tot_lums_global)
                            mg4.metric("🏗️ Total Postes", tot_postes_global)
                            mg5.metric("🛣️ Kilometraje Total", f"{round(tot_dist_global, 1)} km")

                            st.markdown("### 📋 Desglose Técnico por Ruta")
                            st.markdown(resumen_global_texto)

                            # --- GENERACIÓN DEL EXCEL PRO MULTI-RESUMEN ---
                            buf_xlsx = io.BytesIO()
                            with pd.ExcelWriter(buf_xlsx, engine='openpyxl') as writer:
                                df_maestro_completo = pd.concat(excel_rutas_desglose, ignore_index=True)
                                df_maestro_completo.to_excel(writer, index=False, sheet_name='Plan_De_Rutas_SF')
                                ws = writer.sheets['Plan_De_Rutas_SF']
                                
                                ws.delete_rows(2, ws.max_row)
                                r_actual = 2
                                fg, fa = PatternFill(start_color="E2E2E2", end_color="E2E2E2", fill_type="solid"), PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
                                
                                for df_r in excel_rutas_desglose:
                                    r_id = df_r.iloc[0]['Ruta_Asignada']
                                    m_info = metricas_por_ruta[r_id]
                                    inicio_bloque = r_actual
                                    
                                    for _, fila in df_r.iterrows():
                                        for c_idx, valor in enumerate(fila, 1):
                                            ws.cell(row=r_actual, column=c_idx, value=valor)
                                        
                                        if int(fila['Cant_Postes']) > 0:
                                            for cell in ws[r_actual]: cell.fill = fg
                                        elif int(fila['Cant_Cable_m']) > 0:
                                            for cell in ws[r_actual]: cell.fill = fa
                                        r_actual += 1
                                    
                                    fin_bloque = r_actual - 1
                                    
                                    r_actual += 1
                                    ws.cell(row=r_actual, column=2, value=f"--- RESUMEN OPERATIVO DINÁMICO ({r_id}) ---")
                                    ws.cell(row=r_actual+1, column=1, value="Total Puntos:"); ws.cell(row=r_actual+1, column=2, value=m_info["puntos"])
                                    ws.cell(row=r_actual+2, column=1, value="Total Luminarias:"); ws.cell(row=r_actual+2, column=2, value=f"=SUM(D{inicio_bloque}:D{fin_bloque})")
                                    ws.cell(row=r_actual+3, column=1, value="Total Postes:"); ws.cell(row=r_actual+3, column=2, value=f"=SUM(E{inicio_bloque}:E{fin_bloque})")
                                    ws.cell(row=r_actual+4, column=1, value="Total Cable:"); ws.cell(row=r_actual+4, column=2, value=f"=SUM(F{inicio_bloque}:F{fin_bloque})")
                                    ws.cell(row=r_actual+5, column=1, value="Distancia Tramo:"); ws.cell(row=r_actual+5, column=2, value=f"{m_info['distancia']} km")
                                    ws.cell(row=r_actual+6, column=1, value="Tiempo Estimado:"); ws.cell(row=r_actual+6, column=2, value=m_info["tiempo"])
                                    
                                    r_actual += 9 

                            st.dataframe(df_maestro_completo, use_container_width=True, hide_index=True)

                            # --- GENERACIÓN DEL CSV MULTI-RUTA ESTÁTICO ---
                            for df_r in excel_rutas_desglose:
                                df_r.to_csv(csv_multi_buffer, index=False)
                                r_id = df_r.iloc[0]['Ruta_Asignada']
                                m_info = metricas_por_ruta[r_id]
                                csv_multi_buffer.write(f"\n--- RESUMEN OPERATIVO DINÁMICO ({r_id}) ---\n")
                                csv_multi_buffer.write(f"Total Puntos:,{m_info['puntos']}\n")
                                csv_multi_buffer.write(f"Distancia Total:,{m_info['distancia']} km\n")
                                csv_multi_buffer.write(f"Tiempo Estimado:,{m_info['tiempo']}\n\n")

                            st.write("---")
                            c1, c2, c3, c4 = st.columns(4)
                            c1.download_button("📗 Excel Multi-Resumen Pro", buf_xlsx.getvalue(), file_name=f"SF_MULTI_PRO_{up_name}.xlsx", use_container_width=True)
                            c2.download_button("📊 CSV Multi-Estático", csv_multi_buffer.getvalue().encode('utf-8-sig'), file_name=f"SF_MULTI_PRO_{up_name}.csv", use_container_width=True)
                            c3.download_button("🗺️ KML Maestro Detallado", kml.kml(), file_name=f"SF_MULTI_PRO_{up_name}.kml", use_container_width=True)
                            c4.link_button("🚀 My Maps", "https://www.google.com/maps/d/", use_container_width=True)

                            if st.button("💾 REGISTRAR LOTE EN BITÁCORA", use_container_width=True, key="reg_m"):
                                try:
                                    conn = st.connection("gsheets", type=GSheetsConnection)
                                    hist = conn.read(spreadsheet=URL_DB, worksheet=HOJA_PRINCIPAL, ttl=0).dropna(how='all')
                                    info_j = f"Modo: Multi-Ruta Pro, Rutas: {len(lista_rutas_finales)}, Pts: {tot_puntos_global}, Lums: {tot_lums_global}, Dist: {round(tot_dist_global,1)}km"
                                    n_f = pd.DataFrame([{"Fecha": pd.Timestamp.now().strftime("%d/%m/%Y %H:%M"), "Nombre_Ruta": f"MULTIRUTA_PRO_{up_name}", "Usuario_Generador": st.session_state.usuario_nombre, "Datos_JSON": info_j}])
                                    conn.update(spreadsheet=URL_DB, worksheet=HOJA_PRINCIPAL, data=pd.concat([hist, n_f], ignore_index=True))
                                    st.balloons(); st.success("¡Bitácora actualizada!")
                                except Exception as e: st.error(f"Error GSheets: {e}")
                        else:
                            st.error("No se pudieron extraer coordenadas válidas en Modo Multi-Ruta.")
                    except Exception as e: st.error(f"Error en Motor Multi-Ruta V24: {e}")

        # ==========================================
        # PESTAÑA 3: BITÁCORA DE PROCESOS
        # ==========================================
        with tab2:
            try:
                conn = st.connection("gsheets", type=GSheetsConnection)
                df_bt = conn.read(spreadsheet=URL_DB, worksheet=HOJA_PRINCIPAL, ttl=0).dropna(how='all')
                if not df_bt.empty:
                    df_bt_v = df_bt.copy()
                    df_bt_v.insert(0, "ID_Reg", range(1, len(df_bt_v) + 1))
                    if st.session_state.perfil == "ADMIN":
                        c_sel, c_del = st.columns([3, 1])
                        ids_e = st.multiselect("ID para mover a papelera:", df_bt_v["ID_Reg"].tolist())
                        if c_del.button("🗑️ Mover"):
                            if ids_e:
                                idx_e = df_bt_v[df_bt_v["ID_Reg"].isin(ids_e)].index
                                df_tr = conn.read(spreadsheet=URL_DB, worksheet=HOJA_PAPELERA, ttl=0).dropna(how='all')
                                conn.update(spreadsheet=URL_DB, worksheet=HOJA_PAPELERA, data=pd.concat([df_tr, df_bt.loc[idx_e]], ignore_index=True))
                                conn.update(spreadsheet=URL_DB, worksheet=HOJA_PRINCIPAL, data=df_bt.drop(idx_e))
                                st.success("Movido."); time.sleep(1); st.rerun()
                    st.dataframe(df_bt_v.sort_values("ID_Reg", ascending=False), hide_index=True, use_container_width=True)
                else: 
                    st.info("Bitácora vacía.")
            except Exception as e: 
                st.info(f"Sincronizando bitácora... {e}")

        # ==========================================
        # PESTAÑA 4: PAPELERA
        # ==========================================
        with tab3:
            if st.session_state.perfil == "ADMIN":
                try:
                    conn = st.connection("gsheets", type=GSheetsConnection)
                    df_tr = conn.read(spreadsheet=URL_DB, worksheet=HOJA_PAPELERA, ttl=0).dropna(how='all')
                    if not df_tr.empty:
                        df_tr_v = df_tr.copy()
                        df_tr_v.insert(0, "ID_Reg", range(1, len(df_tr_v) + 1))
                        col_r1, col_r2, col_r3 = st.columns([2, 1, 1])
                        with col_r1: ids_r = st.multiselect("ID para restaurar:", df_tr_v["ID_Reg"].tolist())
                        with col_r2: 
                            if st.button("♻️ Restaurar"):
                                if ids_r:
                                    idx_r = df_tr_v[df_tr_v["ID_Reg"].isin(ids_r)].index
                                    df_pr = conn.read(spreadsheet=URL_DB, worksheet=HOJA_PRINCIPAL, ttl=0).dropna(how='all')
                                    conn.update(spreadsheet=URL_DB, worksheet=HOJA_PRINCIPAL, data=pd.concat([df_pr, df_tr.loc[idx_r]], ignore_index=True))
                                    conn.update(spreadsheet=URL_DB, worksheet=HOJA_PAPELERA, data=df_tr.drop(idx_r))
                                    st.success("Restaurado."); time.sleep(1); st.rerun()
                        with col_r3:
                            if st.button("🔥 VACIAR PAPELERA"):
                                df_vacio = pd.DataFrame(columns=df_tr.columns)
                                conn.update(spreadsheet=URL_DB, worksheet=HOJA_PAPELERA, data=df_vacio)
                                st.success("¡Papelera purgada!"); time.sleep(1); st.rerun()
                        st.dataframe(df_tr_v, hide_index=True, use_container_width=True)
                    else: 
                        st.info("Papelera vacía.")
                except Exception as e: 
                    st.info(f"Cargando papelera... {e}")
            else:
                st.warning("🔒 Área restringida para administradores.")

    elif st.session_state.menu == "SF4":
        st.title("🏗️ SF4 - Arquitecto de Procesos & Oficios")

        # --- 1. CONEXIÓN A GOOGLE SHEETS ---
        scope = ["https://www.googleapis.com/auth/spreadsheets"]
        creds_dict = {
            "type": st.secrets["connections"]["gsheets"]["type"],
            "project_id": st.secrets["connections"]["gsheets"]["project_id"],
            "private_key_id": st.secrets["connections"]["gsheets"]["private_key_id"],
            "private_key": st.secrets["connections"]["gsheets"]["private_key"].replace('\\n', '\n'),
            "client_email": st.secrets["connections"]["gsheets"]["client_email"],
            "client_id": st.secrets["connections"]["gsheets"]["client_id"],
            "auth_uri": st.secrets["connections"]["gsheets"]["auth_uri"],
            "token_uri": st.secrets["connections"]["gsheets"]["token_uri"],
            "auth_provider_x509_cert_url": st.secrets["connections"]["gsheets"]["auth_provider_x509_cert_url"],
            "client_x509_cert_url": st.secrets["connections"]["gsheets"]["client_x509_cert_url"]
        }
        creds = Credentials.from_service_account_info(creds_dict, scopes=scope)
        client = gspread.authorize(creds)
        SHEET_ID = "14_fewol5DiFXoiO102wviiWR08Lw3PKHzEjSbMwxUm8"
        
        try:
            sh = client.open_by_key(SHEET_ID)
            ws = sh.worksheet("Boveda_Bajas")
        except:
            ws = sh.get_worksheet(0)

        # --- SINCRO-BÓVEDA RESILIENTE (ANTI-DESCONEXIONES) ---
        bovedas_completas = ["db_oficios", "boveda_mmd", "db_depuracion", "vales_historial"]
        if any(b not in st.session_state for b in bovedas_completas):
            for _ in range(3):
                try:
                    if "db_oficios" not in st.session_state: st.session_state.db_oficios = {}
                    if "boveda_mmd" not in st.session_state: st.session_state.boveda_mmd = {}
                    if "db_depuracion" not in st.session_state: st.session_state.db_depuracion = {}
                    if "vales_historial" not in st.session_state: st.session_state.vales_historial = []
                    
                    # Descarga de matriz pura de celdas para ignorar problemas de encabezados
                    filas_raw = ws.get_all_values()
                    if len(filas_raw) > 1:
                        for row in filas_raw[1:]: # Saltamos la fila 1 de encabezados
                            if len(row) >= 5:
                                reg_id = str(row[0]).strip()
                                origen = str(row[2]).strip() if str(row[2]).strip() else "Sin Nombre"
                                datos_raw = str(row[4]).strip() if row[4] else "{}"
                                
                                if reg_id.startswith("SF4-PRY-"):
                                    try: st.session_state.boveda_mmd[origen] = json.loads(datos_raw) if isinstance(datos_raw, str) else datos_raw
                                    except: pass
                                elif reg_id.startswith("SF4-OFC-"):
                                    try: st.session_state.db_oficios[origen] = json.loads(datos_raw) if isinstance(datos_raw, str) else datos_raw
                                    except: pass
                                elif reg_id.startswith("SF5-DEP-"):
                                    try: st.session_state.db_depuracion[reg_id] = json.loads(datos_raw) if isinstance(datos_raw, str) else datos_raw
                                    except: pass
                                elif reg_id.startswith("SF6-VAL-"):
                                    try:
                                        vale_parsed = json.loads(datos_raw) if isinstance(datos_raw, str) else datos_raw
                                        if vale_parsed not in st.session_state.vales_historial:
                                            st.session_state.vales_historial.append(vale_parsed)
                                    except: pass
                    break
                except Exception:
                    time.sleep(1)

        tab_c, tab_b, tab_i, tab_o, tab_j = st.tabs(["🆕 Constructor Inteligente", "🗄️ Bóveda de Proyectos", "📥 Importación Externa", "📄 GENERADOR DE OFICIOS", "📝 JUSTIFICACIONES OPERATIVAS"])
        with tab_c:
            with st.expander("📝 CONFIGURAR PASO", expanded=True):
                idx = st.session_state.edit_index
                editando = (idx != -1)
                paso_actual = st.session_state.pasos_sf4[idx] if editando else {}
                
                txt = st.text_input("Actividad o Pregunta (usa '?' para bifurcar):", value=paso_actual.get('texto', ""), key=f"txt_sf4_{idx}")
                is_decision = txt.strip().endswith('?')
                destinos = ["Siguiente", "Fin"] + [f"Paso {i+1}" for i in range(len(st.session_state.pasos_sf4))]

                c1, c2, c3 = st.columns(3)
                if not is_decision:
                    with c1: tipo = st.selectbox("Forma:", ["Proceso", "Inicio/Fin"], index=0 if paso_actual.get('tipo') == "Proceso" else (1 if paso_actual.get('tipo') == "Inicio/Fin" else 0))
                    with c2: 
                        d_val = paso_actual.get('conecta_a', "Siguiente")
                        destino = st.selectbox("Conecta a:", destinos, index=destinos.index(d_val) if d_val in destinos else 0)
                    with c3: label = st.text_input("Etiqueta flecha:", value=paso_actual.get('etiqueta_flecha', ""), placeholder="Ej: Ok")
                else:
                    with c1: 
                        label_si = st.text_input("Etiqueta SÍ:", value=paso_actual.get('label_si', "SÍ"))
                        d_si_val = paso_actual.get('dest_si', "Siguiente")
                        dest_si = st.selectbox("Destino SÍ:", destinos, index=destinos.index(d_si_val) if d_si_val in destinos else 0)
                    with c2: 
                        label_no = st.text_input("Etiqueta NO:", value=paso_actual.get('label_no', "NO"))
                        d_no_val = paso_actual.get('dest_no', "Siguiente")
                        dest_no = st.selectbox("Destino NO (Salto):", destinos, index=destinos.index(d_no_val) if d_no_val in destinos else 0)
                    with c3: st.info("Las decisiones requieren dos salidas obligatorias.")

                if not editando:
                    if st.button("➕ Agregar al Flujo", use_container_width=True):
                        if txt:
                            nuevo = {"texto": txt, "is_decision": is_decision}
                            if is_decision: nuevo.update({"label_si": label_si, "dest_si": dest_si, "label_no": label_no, "dest_no": dest_no, "tipo": "Decisión"})
                            else: nuevo.update({"tipo": tipo, "conecta_a": destino, "etiqueta_flecha": label})
                            st.session_state.pasos_sf4.append(nuevo)
                            st.rerun()
                else:
                    cs, cc = st.columns(2)
                    if cs.button("💾 Guardar Cambios", use_container_width=True):
                        nuevo = {"texto": txt, "is_decision": is_decision}
                        if is_decision: nuevo.update({"label_si": label_si, "dest_si": dest_si, "label_no": label_no, "dest_no": dest_no, "tipo": "Decisión"})
                        else: nuevo.update({"tipo": tipo, "conecta_a": destino, "etiqueta_flecha": label})
                        st.session_state.pasos_sf4[idx] = nuevo
                        st.session_state.edit_index = -1
                        st.rerun()
                    if cc.button("❌ Cancelar", use_container_width=True):
                        st.session_state.edit_index = -1
                        st.rerun()

            if st.session_state.pasos_sf4:
                col_l, col_p = st.columns([1, 1.2])
                with col_l:
                    st.subheader("📋 Pasos")
                    for i, p in enumerate(st.session_state.pasos_sf4):
                        with st.container(border=True):
                            cx, cy, cz = st.columns([0.5, 3, 1])
                            cx.write(f"#{i+1}"); cy.write(p['texto'])
                            if cz.button("✏️", key=f"e_{i}"): st.session_state.edit_index = i; st.rerun()
                            if cz.button("🗑️", key=f"d_{i}"): st.session_state.pasos_sf4.pop(i); st.rerun()
                    
                    st.write("---")
                    seguro_reiniciar = st.checkbox("🔐 Confirmar vaciado total de la mesa", key="reiniciar_seguro_sf4")
                    if st.button("🔥 Reiniciar Mesa", use_container_width=True, disabled=not seguro_reiniciar): 
                        st.session_state.pasos_sf4 = []
                        st.rerun()

                with col_p:
                    st.subheader("📊 Visualización Premium")
                    def clean(t): return re.sub(r'[^a-zA-Z0-9 áéíóúÁÉÍÓÚñÑ]', '', str(t))
                    
                    mmd_head = ["graph TD", "classDef decision fill:#f9f,stroke:#333,stroke-width:2px;", "classDef proceso fill:#bbf,stroke:#333,stroke-width:2px;"]
                    mmd_nodos, mmd_conexiones = [], []

                    for i, p in enumerate(st.session_state.pasos_sf4):
                        id_n = f"N{i}"
                        t_c = clean(p.get('texto', ''))
                        if p.get('tipo') == "Decisión": mmd_nodos.append(f'    {id_n}{{\"{t_c}\"}}:::decision')
                        elif p.get('tipo') == "Inicio/Fin": mmd_nodos.append(f'    {id_n}((\"{t_c}\"))')
                        else: mmd_nodos.append(f'    {id_n}[\"{t_c}\"]:::proceso')

                    for i, p in enumerate(st.session_state.pasos_sf4):
                        id_n = f"N{i}"
                        if not p.get('is_decision', False):
                            tgt = p.get('conecta_a', "Siguiente")
                            lab = p.get('etiqueta_flecha', "")
                            f_style = f'-- "{lab}" -->' if lab else "-->"
                            if tgt == "Siguiente" and i < len(st.session_state.pasos_sf4)-1: mmd_conexiones.append(f'    {id_n} {f_style} N{i+1}')
                            elif tgt == "Fin": mmd_conexiones.append(f'    {id_n} {f_style} Fin([Fin])')
                            elif "Paso" in str(tgt):
                                p_num = int(re.search(r'\d+', str(tgt)).group()) - 1
                                mmd_conexiones.append(f'    {id_n} {f_style} N{p_num}')
                        else:
                            for l_key, d_key in [('label_si', 'dest_si'), ('label_no', 'dest_no')]:
                                dst = p.get(d_key, "Siguiente")
                                lab_f = p.get(l_key, "Opción")
                                f_style = f'-- "{lab_f}" -->'
                                if dst == "Siguiente" and i < len(st.session_state.pasos_sf4)-1: mmd_conexiones.append(f'    {id_n} {f_style} N{i+1}')
                                elif dst == "Fin": mmd_conexiones.append(f'    {id_n} {f_style} Fin([Fin])')
                                elif "Paso" in str(dst):
                                    p_num = int(re.search(r'\d+', str(dst)).group()) - 1
                                    mmd_conexiones.append(f'    {id_n} {f_style} N{p_num}')

                    full_m = "\n".join(mmd_head + mmd_nodos + mmd_conexiones)
                    st.code(full_m, language="mermaid")
                    
                    if st.session_state.pasos_sf4:
                        tema = st.session_state.pasos_sf4[0]['texto'].replace('?', '')
                        st.markdown("---")
                        st.subheader("📝 Objetivos del Proceso")
                        adm, tec = st.columns(2)
                        with adm:
                            st.info("**Administrativo-Normativo**")
                            st.caption(f"Establecer el marco procedimental de '{tema}', asegurando el cumplimiento de los criterios de validación.")
                        with tec:
                            st.success("**Técnico-Operativo**")
                            st.caption(f"Optimizar la respuesta de las cuadrillas en '{tema}', mediante la estandarización técnica.")

                    b64 = base64.b64encode(full_m.encode('utf-8')).decode('utf-8')
                    st.link_button("🚀 LIVE EDITOR", f"https://mermaid.live/edit#base64:{b64}", use_container_width=True)
                    st.write("---")
                    nom_p = st.text_input("Nombre para Bóveda:")
                    if st.button("💾 Guardar en Bóveda Pangea"):
                        if nom_p:
                            # --- CONTROL DE DUPLICADOS EN DIAGRAMAS ---
                            if nom_p in st.session_state.boveda_mmd:
                                st.error(f"⚠️ Error: El nombre '{nom_p}' ya existe en la bóveda. Use un nombre diferente para evitar confusiones.")
                            else:
                                payload = {"code": full_m, "struct": list(st.session_state.pasos_sf4)}
                                st.session_state.boveda_mmd[nom_p] = payload
                                
                                tz_mx = timezone(timedelta(hours=-6))
                                ahora = datetime.now(tz_mx)
                                id_reg = f"SF4-PRY-{ahora.strftime('%Y%m%d-%H%M%S')}"
                                fecha_mx = ahora.strftime("%d/%m/%Y %H:%M:%S")
                                
                                ws.append_row([id_reg, fecha_mx, nom_p, len(st.session_state.pasos_sf4), json.dumps(payload), ""])
                                st.success("Guardado en Bóveda Permanente Correctamente.")

        with tab_b:
            if not st.session_state.boveda_mmd: st.info("Bóveda vacía.")
            else:
                for k, v in list(st.session_state.boveda_mmd.items()):
                    with st.expander(f"📁 {k}"):
                        # --- ESCUDO ANTI-CRASH PARA CONTROL DE COMPONENTES CORRUPTOS ---
                        if isinstance(v, dict) and 'code' in v and 'struct' in v:
                            st.code(v['code'], language="mermaid")
                            b1, b2, b3 = st.columns(3)
                            if b1.button("🛠️ RECUPERAR", key=f"r_{k}"): st.session_state.pasos_sf4 = list(v['struct']); st.rerun()
                            b_u = base64.b64encode(v['code'].encode('utf-8')).decode('utf-8')
                            b2.link_button("🚀 Live", f"https://mermaid.live/edit#base64:{b_u}")
                        else:
                            st.error("⚠️ Los datos de este diagrama están corruptos o incompletos en Google Sheets.")
                            b1, b2, b3 = st.columns(3) # Mantiene las columnas alineadas para el botón de borrado

                        if k.strip().upper() != "PASTEL VERDE":
                            seguro_borrado_p = st.checkbox("🔐 Confirmar borrado permanente del diagrama", key=f"chk_del_mmd_{k}")
                            if b3.button("🗑️", key=f"x_{k}", use_container_width=True, disabled=not seguro_borrado_p):
                                try:
                                    cell = ws.find(k, in_column=3)
                                    if cell: ws.delete_rows(cell.row)
                                    del st.session_state.boveda_mmd[k]
                                    st.success("¡Eliminado de la nube!")
                                    time.sleep(1); st.rerun()
                                except Exception as e: st.error(f"Error al borrar: {e}")

        with tab_i:
            st.subheader("📥 Importación Externa")
            raw_import = st.text_area("Pega el código Mermaid aquí:", height=300, key="area_import_sf", placeholder="graph TD\nN0((\"Inicio\")) --> N1[\"Proceso\"]")
            
            if st.button("🚀 REDISEÑAR PROCESO", use_container_width=True):
                if raw_import:
                    try:
                        nuevos_pasos = []
                        lineas = [l.strip() for l in raw_import.split('\n') if l.strip()]
                        
                        for linea in lineas:
                            m_io = re.search(r'(\w+)\(\("(.+?)"\)\)', linea)
                            if m_io:
                                nuevos_pasos.append({"texto": m_io.group(2), "tipo": "Inicio/Fin", "is_decision": False, "conecta_a": "Siguiente", "etiqueta_flecha": ""})
                                continue
                            
                            m_dec = re.search(r'(\w+)\{"(.+?)"\}', linea)
                            if m_dec:
                                nuevos_pasos.append({"texto": m_dec.group(2) + "?", "tipo": "Decisión", "is_decision": True, "label_si": "SÍ", "dest_si": "Siguiente", "label_no": "NO", "dest_no": "Fin"})
                                continue

                            m_proc = re.search(r'(\w+)\["(.+?)"\]', linea)
                            if m_proc:
                                nuevos_pasos.append({"texto": m_proc.group(2), "tipo": "Proceso", "is_decision": False, "conecta_a": "Siguiente", "etiqueta_flecha": ""})

                        if nuevos_pasos:
                            st.session_state.pasos_sf4 = nuevos_pasos
                            st.success(f"✅ Se han importado {len(nuevos_pasos)} pasos correctamente.")
                            time.sleep(1); st.rerun()
                        else:
                            st.error("❌ No se detectaron nodos válidos en el código. Verifica el formato (N0[\"Texto\"]).")
                    except Exception as e:
                        st.error(f"Error en el motor de importación: {e}")
                else:
                    st.warning("⚠️ El área de texto está vacía.")

        with tab_o:
            st.subheader("📄 Correspondencia Oficial y Control de Bóveda")

            try:
                from fpdf import FPDF
                motor_pdf_listo = True
            except ImportError:
                motor_pdf_listo = False
                st.warning("⚠️ Motor PDF (fpdf) no detectado. Las descargas están deshabilitadas.")

            plantillas_maestras = {
                "✅ Atención Exitosa": "Por medio de la presente, se hace de su conocimiento que la petición con folio [FOLIO] ha sido atendida exitosamente por las brigadas de esta Dirección, quedando el servicio en óptimas condiciones de operación.",
                "💡 Ya en Servicio": "Tras la inspección realizada por el personal técnico, se hace de su conocimiento que la luminaria correspondiente al folio [FOLIO] ya se encuentra en servicio y funcionando correctamente.",
                "⏳ Programado/Parcial": "Se informa que la atención al folio [FOLIO] se encuentra en estado parcial; los trabajos continuarán conforme a la disponibilidad de material specialized en el programa de mantenimiento.",
                "🔌 Bajadas de Luz": "Se autoriza la maniobra de bajada de luz solicitada mediante el folio [FOLIO], misma que será coordinada por el personal asignado a la zona correspondiente.",
                "❌ Atención Negativa": "Respecto a la petición [FOLIO], se informa que tras el análisis técnico, la solicitud ha sido determinada como improcedente debido a restricciones normativas o técnicas vigentes.",
                "✏️ Libre (Escribir desde cero)": ""
            }

            c_config, c_preview = st.columns([1, 1.1])

            with c_config:
                modo_of = st.radio("Operación:", ["✨ Crear Nuevo", "📂 Consultar Bóveda"], horizontal=True)
                
                data_previa = {}
                id_sel = "nuevo"
                if modo_of == "📂 Consultar Bóveda":
                    if st.session_state.db_oficios:
                        id_sel = st.selectbox("Seleccionar Oficio:", list(st.session_state.db_oficios.keys())[::-1])
                        data_previa = st.session_state.db_oficios[id_sel]
                        
                        # --- TABLA RESUMEN CON RESALTADO DINÁMICO ---
                        lista_oficios_tabla = [{"Oficio ID": k, "Fecha": v.get("fecha", "N/A"), "Folio Ref": v.get("folio", "N/A")} for k, v in st.session_state.db_oficios.items()]
                        df_oficios_vista = pd.DataFrame(lista_oficios_tabla)
                        
                        def resaltar_oficio_seleccionado(row):
                            color = '#d1e7dd' if row['Oficio ID'] == id_sel else ''
                            return [f'background-color: {color}'] * len(row)
                            
                        st.dataframe(df_oficios_vista.style.apply(resaltar_oficio_seleccionado, axis=1), use_container_width=True, hide_index=True)
                        
                        # Proximidad Máxima: Checkbox y Botón acoplados juntos abajo de la tabla summary
                        col_chk_del, col_btn_del = st.columns([1.8, 1.2])
                        with col_chk_del:
                            seguro_borrado = st.checkbox("🔐 Confirmar eliminación permanente", key=f"chk_del_ofc_{id_sel}")
                        with col_btn_del:
                            if st.button("🗑️ BORRAR OFICIO", use_container_width=True, disabled=not seguro_borrado, type="primary", key=f"btn_del_ofc_{id_sel}"):
                                try:
                                    cell = ws.find(id_sel, in_column=3)
                                    if cell: ws.delete_rows(cell.row)
                                    del st.session_state.db_oficios[id_sel]
                                    st.warning(f"Registro {id_sel} eliminado de la nube.")
                                    time.sleep(1); st.rerun()
                                except Exception as e: st.error(f"Error al eliminar: {e}")
                    else:
                        st.info("La bóveda está vacía.")
                
                # Generamos una llave de identidad única para forzar el refresco de datos
                pk = f"{modo_of}_{id_sel}"
                
                with st.container(border=True):
                    st.markdown("**📌 Configuración**")
                    
                    # Recuperar la plantilla de origen para que el selector no se mueva de su lugar
                    plantilla_guardada = data_previa.get("plantilla", list(plantillas_maestras.keys())[0])
                    lista_plantillas = list(plantillas_maestras.keys())
                    idx_plantilla = lista_plantillas.index(plantilla_guardada) if plantilla_guardada in lista_plantillas else 0
                    
                    tipo_p = st.selectbox("Plantilla:", lista_plantillas, index=idx_plantilla, key=f"tipo_p_{pk}")
                    c1, c2 = st.columns(2)
                    n_oficio = c1.text_input("No. Oficio:", value=data_previa.get("num", "DAP/___/2026"), key=f"num_{pk}")
                    
                    if data_previa.get("fecha"):
                        try: default_date = pd.to_datetime(data_previa.get("fecha")).date()
                        except: default_date = pd.Timestamp.now().date()
                    else:
                        default_date = pd.Timestamp.now().date()
                        
                    f_oficio = c2.date_input("Fecha:", value=default_date, key=f"fecha_{pk}")
                    dest = st.text_area("Destinatario:", value=data_previa.get("dest", ""), height=70, key=f"dest_{pk}", kwargs={"spellcheck": "true"})
                    cargo = st.text_input("Cargo:", value=data_previa.get("cargo", "P R E S E N T E"), key=f"cargo_{pk}")
                    f_ref = st.text_input("Folio Ref:", value=data_previa.get("folio", ""), key=f"folio_{pk}")

                with st.container(border=True):
                    st.markdown("**📝 Mensaje**")
                    v_cuerpo = data_previa.get("cuerpo", plantillas_maestras[tipo_p])
                    cuerpo_txt = st.text_area("Cuerpo:", value=v_cuerpo, height=150, key=f"cuerpo_{pk}_{tipo_p}", kwargs={"spellcheck": "true"})
                    firm = st.text_input("Firma (Nombre):", value=data_previa.get("firma", "NOMBRE DEL DIRECTOR"), key=f"firma_{pk}")
                    cargo_firm = st.text_input("Cargo del Firmante:", value=data_previa.get("cargo_f", "DIRECTOR DE ALUMBRADO PÚBLICO"), key=f"cargo_f_{pk}")
                    ccp = st.text_area("C.c.p.:", value=data_previa.get("ccp", "Archivo, Minutario."), height=65, key=f"ccp_{pk}", kwargs={"spellcheck": "true"})

                h_membrete = st.toggle("🛰️ Modo Hoja Membretada", value=False, key=f"memb_{pk}")

            with c_preview:
                st.markdown("### 👁️ Vista Previa")
                c_final = cuerpo_txt.replace("[FOLIO]", f"**{f_ref}**" if f_ref else "**_______**")
                e_sup = "100px" if h_membrete else "20px"

                st.markdown(f"""
                <div style="background: white; color: black; padding: 40px; border: 1px solid #ddd; font-family: 'Arial'; line-height: 1.6; min-height: 550px;">
                    <div style="height: {e_sup};"></div>
                    <div style="text-align: right; font-weight: bold;">Toluca, México; a {f_oficio.strftime('%d/%m/%Y')}<br>Oficio: {n_oficio}</div><br>
                    <div style="text-align: left; font-weight: bold; white-space: pre-line;">{dest.upper()}<br>{cargo.upper()}</div><br>
                    <div style="text-align: justify;"> {c_final} </div><br><br>
                    <div style="text-align: center;"><b>A T E N T A M E N T E</b><br><br><br>__________________________<br><b>{firm.upper()}</b><br>{cargo_firm.upper()}</div>
                    <div style="font-size: 10px; border-top: 1px solid #eee; margin-top: 20px; white-space: pre-line;">C.c.p. {ccp}</div>
                </div>
                """, unsafe_allow_html=True)

                st.divider()
                
                ejecutar_guardado = False
                
                # Separación e identificación de flujos de guardado independientes
                if modo_of == "📂 Consultar Bóveda":
                    seguro_actualizar = st.checkbox("🔐 Confirmar cambios y actualización del oficio histórico", key="chk_seguro_actualizar_ofc")
                    if st.button("🔄 ACTUALIZAR REGISTRO EXISTENTE", use_container_width=True, disabled=not seguro_actualizar, type="primary"):
                        ejecutar_guardado = True
                else:
                    if st.button("💾 GUARDAR NUEVO OFICIO", use_container_width=True, type="primary"):
                        ejecutar_guardado = True
                        
                if ejecutar_guardado:
                    id_r = n_oficio.replace("/", "-")
                    
                    # --- CONTROL DE DUPLICADOS EN OFICIOS (SOLO AL CREAR NUEVOS) ---
                    if modo_of == "✨ Crear Nuevo" and id_r in st.session_state.db_oficios:
                        st.error(f"⚠️ Error: El oficio '{n_oficio}' ya existe en el registro histórico de la Dirección. Por favor verifique el número.")
                    else:
                        payload_oficio = {
                            "num": n_oficio, "fecha": str(f_oficio), "dest": dest, 
                            "cargo": cargo, "folio": f_ref, "cuerpo": cuerpo_txt, 
                            "firma": firm, "cargo_f": cargo_firm, "ccp": ccp,
                            "plantilla": tipo_p
                        }
                        st.session_state.db_oficios[id_r] = payload_oficio
                        
                        tz_mx = timezone(timedelta(hours=-6))
                        ahora = datetime.now(tz_mx)
                        id_reg = f"SF4-OFC-{ahora.strftime('%Y%m%d-%H%M%S')}"
                        fecha_mx = ahora.strftime("%d/%m/%Y %H:%M:%S")
                        
                        try:
                            cell = ws.find(id_r, in_column=3)
                            if cell: ws.delete_rows(cell.row)
                        except: pass
                        
                        ws.append_row([id_reg, fecha_mx, id_r, f_ref, json.dumps(payload_oficio), ""])
                        st.success("✅ Bóveda Nube de Oficios Actualizada Exitosamente."); time.sleep(1); st.rerun()
                if motor_pdf_listo:
                    pdf = FPDF(orientation='P', unit='mm', format='Letter')
                    pdf.set_margins(30, 25, 20)
                    pdf.set_auto_page_break(auto=True, margin=25) 
                    pdf.add_page()
                    
                    if h_membrete: 
                        pdf.ln(15)
                    pdf.set_font("Arial", 'B', 11)
                    pdf.cell(0, 5, txt=f"Toluca, México; a {f_oficio.strftime('%d/%m/%Y')}", ln=True, align='R')
                    pdf.cell(0, 5, txt=f"Oficio No: {n_oficio}", ln=True, align='R')
                    pdf.ln(15); pdf.cell(0, 5, txt=dest.upper() if dest else "A QUIEN CORRESPONDA", ln=True)
                    pdf.cell(0, 5, txt=cargo.upper(), ln=True)
                    pdf.ln(15); pdf.set_font("Arial", '', 11)
                    c_pdf = cuerpo_txt.replace("[FOLIO]", f_ref)
                    pdf.multi_cell(0, 7, txt=c_pdf.encode('latin-1', 'replace').decode('latin-1'), align='J')
                    pdf.ln(25); pdf.set_font("Arial", 'B', 11); pdf.cell(0, 5, txt="A T E N T A M E N T E", ln=True, align='C')
                    pdf.ln(20); pdf.cell(0, 5, txt="__________________________", ln=True, align='C')
                    pdf.cell(0, 5, txt=firm.upper(), ln=True, align='C')
                    pdf.cell(0, 5, txt=cargo_firm.upper(), ln=True, align='C')
                    pdf.set_y(-30); pdf.set_font("Arial", '', 8); pdf.cell(0, 5, txt=f"C.c.p. {ccp}", ln=True)
                    
                    pdf_data = pdf.output(dest='S').encode('latin-1', 'replace')
                    st.download_button(label="🚀 DESCARGAR OFICIO PDF", data=pdf_data, file_name=f"Oficio_{n_oficio.replace('/','-')}.pdf", mime="application/pdf", use_container_width=True)
                else:
                    st.error("❌ Función PDF no disponible.")
# ==================================================================================
# 📝 AQUÍ TERMINA OFICIOS (TAB_O) Y EMPIEZA INDEPENDIENTE LA NUEVA PESTAÑA (TAB_J)
# ==================================================================================
        with tab_j:
            st.subheader("📝 Control de Justificaciones e Incidencias de Personal")
            
            # --- SINCRO-BÓVEDA LOCAL (EXCLUSIVA PARA JUSTIFICACIONES) ---
            if "db_justificaciones" not in st.session_state:
                st.session_state.db_justificaciones = {}
                try:
                    filas_raw = ws.get_all_values()
                    if len(filas_raw) > 1:
                        for row in filas_raw[1:]:
                            if len(row) >= 5:
                                reg_id = str(row[0]).strip()
                                datos_raw = str(row[4]).strip() if row[4] else "{}"
                                if reg_id.startswith("SF4-JST-"):
                                    try: st.session_state.db_justificaciones[reg_id] = json.loads(datos_raw)
                                    except: pass
                except: pass
    
            c_config, c_preview = st.columns([1, 1.1])

            with c_config:
                modo_j = st.radio("Operación Justificaciones:", ["✨ Crear Nuevo", "📂 Consultar Bóveda"], horizontal=True, key="radio_modo_justificaciones")
                
                data_previa_j = {}
                id_sel_j = "nuevo"
                
                if modo_j == "📂 Consultar Bóveda":
                    if st.session_state.db_justificaciones:
                        id_sel_j = st.selectbox("Seleccionar Registro:", list(st.session_state.db_justificaciones.keys())[::-1], key="select_boveda_justificaciones")
                        data_previa_j = st.session_state.db_justificaciones[id_sel_j]
                        
                        # Tabla resumen de control histórico
                        lista_j_tabla = [{"ID Registro": k, "Empleado": v.get("nombre", "N/A"), "No. Emp": v.get("num_emp", "N/A"), "Concepto": v.get("clave_concepto", "N/A")} for k, v in st.session_state.db_justificaciones.items()]
                        df_j_vista = pd.DataFrame(lista_j_tabla)
                        
                        def resaltar_j_seleccionada(row):
                            color = '#d1e7dd' if row['ID Registro'] == id_sel_j else ''
                            return [f'background-color: {color}'] * len(row)
                            
                        st.dataframe(df_j_vista.style.apply(resaltar_j_seleccionada, axis=1), use_container_width=True, hide_index=True)
                        
                        col_chk_j, col_btn_j = st.columns([1.8, 1.2])
                        with col_chk_j:
                            seguro_borrado_j = st.checkbox("🔐 Confirmar eliminación permanente del registro", key=f"chk_del_just_{id_sel_j}")
                        with col_btn_j:
                            if st.button("🗑️ BORRAR FORMATO", use_container_width=True, disabled=not seguro_borrado_j, type="primary", key=f"btn_del_just_{id_sel_j}"):
                                try:
                                    cell = ws.find(id_sel_j, in_column=1)
                                    if cell: ws.delete_rows(cell.row)
                                    del st.session_state.db_justificaciones[id_sel_j]
                                    st.warning(f"Registro {id_sel_j} eliminado de la nube.")
                                    time.sleep(1); st.rerun()
                                except Exception as e: st.error(f"Error al eliminar: {e}")
                    else:
                        st.info("La bóveda de justificaciones está vacía.")

                pk_j = f"{modo_j}_{id_sel_j}"

                with st.container(border=True):
                    st.markdown("**📌 Secuencia Oficial de Llenado**")
                    
                    # 1. Fecha del Documento
                    try:
                        f_doc_raw = data_previa_j.get("fecha_doc")
                        if f_doc_raw and str(f_doc_raw) != "NaT" and str(f_doc_raw).strip() != "":
                            t_doc = pd.to_datetime(f_doc_raw)
                            def_f_doc = t_doc.date() if not pd.isna(t_doc) else pd.Timestamp.now().date()
                        else: def_f_doc = pd.Timestamp.now().date()
                    except: def_f_doc = pd.Timestamp.now().date()
                        
                    f_doc = st.date_input("1. Fecha de Emisión del Documento:", value=def_f_doc, key=f"f_doc_{pk_j}")
                    
                    # 2. Solicita
                    solicita_raw = st.text_input("2. Solicita (Área o Persona):", value=data_previa_j.get("solicita", ""), key=f"txt_sol_{pk_j}")
                    
                    # 3. Acción Administrativa
                    accion_idx = ["JUSTIFICAR", "SANCIONAR"].index(data_previa_j.get("accion", "JUSTIFICAR")) if data_previa_j.get("accion") in ["JUSTIFICAR", "SANCIONAR"] else 0
                    accion = st.selectbox("3. Acción Administrativa:", ["JUSTIFICAR", "SANCIONAR"], index=accion_idx, key=f"sel_accion_{pk_j}")
                    
                    # 4. No. de Empleado y Nombre
                    c_emp1, c_emp2 = st.columns([1, 2])
                    num_emp_raw = c_emp1.text_input("4. No. de Empleado:", value=data_previa_j.get("num_emp", ""), key=f"txt_num_emp_{pk_j}")
                    nombre_emp_raw = c_emp2.text_input("Nombre Completo del Empleado:", value=data_previa_j.get("nombre", ""), key=f"txt_nom_emp_{pk_j}")
                    
                    # 5. Adscrito a
                    adscrito_raw = st.text_input("5. Adscrito a (Departamento/Área):", value=data_previa_j.get("adscrito", ""), key=f"txt_ads_{pk_j}")
                    
                    # 6. F. Registro
                    f_reg_idx = ["LISTA", "HAND PUNCH"].index(data_previa_j.get("f_registro", "LISTA")) if data_previa_j.get("f_registro") in ["LISTA", "HAND PUNCH"] else 0
                    f_registro = st.selectbox("6. F. Registro:", ["LISTA", "HAND PUNCH"], index=f_reg_idx, key=f"sel_f_reg_{pk_j}")
                    
                    # 7. Clave y Concepto
                    cat_hand_punch = [
                        "3 | FALTA INJUSTIFICADA", "4 | FALTA JUSTIFICADA (TIEMPO X TIEMPO)", 
                        "9 | LICENCIA CON GOCE DE SUELDO", "10 | LICENCIA SIN GOCE DE SUELDO", 
                        "11 | VACACIONES", "12 | INCAPACIDAD", "14 | COMISIÓN", 
                        "16 | LICENCIA POR MATRIMONIO(SINDICALIZADO)", "17 | LICENCIA POR GRAVIDEZ", 
                        "18 | HORA DE LACTANCIA", "19 | LICENCIA POR FALLECIMIENTO DE FAMILIAR", 
                        "20 | LICENCIA POR NACIMIENTO", "23 | OMISIÓN DE CHECADA", 
                        "24 | DÍA ECONÓMICO (SINDICALIZADO)", "34 | CUMPLEAÑOS (SINDICALIZADO)", 
                        "CM | CUIDADOS MÉDICOS"
                    ]
                    concept_def = data_previa_j.get("clave_concepto", cat_hand_punch[0])
                    concept_idx = cat_hand_punch.index(concept_def) if concept_def in cat_hand_punch else 0
                    clave_concepto = st.selectbox("7. Clave y Concepto (Escribe clave o texto para buscar):", cat_hand_punch, index=concept_idx, key=f"sel_concept_{pk_j}")

                    # 8. Modalidad de Fechas
                    tipo_fecha_j = st.radio("8. Modalidad de Fecha de la Incidencia:", ["Día Único", "Rango de Fechas"], index=0 if data_previa_j.get("tipo_fecha", "Día Único") == "Día Único" else 1, horizontal=True, key=f"radio_tipo_f_{pk_j}")
                    
                    try:
                        f1_raw = data_previa_j.get("fecha_inicio")
                        if f1_raw and str(f1_raw) != "NaT" and str(f1_raw).strip() != "":
                            t1 = pd.to_datetime(f1_raw)
                            def_f1 = t1.date() if not pd.isna(t1) else pd.Timestamp.now().date()
                        else: def_f1 = pd.Timestamp.now().date()
                    except: def_f1 = pd.Timestamp.now().date()

                    try:
                        f2_raw = data_previa_j.get("fecha_fin")
                        if f2_raw and str(f2_raw) != "NaT" and str(f2_raw).strip() != "":
                            t2 = pd.to_datetime(f2_raw)
                            def_f2 = t2.date() if not pd.isna(t2) else pd.Timestamp.now().date()
                        else: def_f2 = pd.Timestamp.now().date()
                    except: def_f2 = pd.Timestamp.now().date()

                    if tipo_fecha_j == "Día Único":
                        f_inicio = st.date_input("Fecha de Incidencia:", value=def_f1, key=f"date_ini_{pk_j}")
                        f_fin = f_inicio
                    else:
                        c_f1, c_f2 = st.columns(2)
                        f_inicio = c_f1.date_input("Fecha Inicio:", value=def_f1, key=f"date_ini_r_{pk_j}")
                        f_fin = c_f2.date_input("Fecha Fin:", value=def_f2, key=f"date_fin_r_{pk_j}")

                with st.container(border=True):
                    st.markdown("**📝 Sección de Cierre Técnico**")
                    # 9. Motivo
                    motivo_raw = st.text_area("9. Motivo de la Incidencia / Justificación:", value=data_previa_j.get("motivo", ""), height=100, key=f"area_mot_{pk_j}")
                    
                    st.markdown("**✏️ Bloques de Validación y Firmas (Selección de Cargo)**")
                    # 10. Solicitante
                    firma_solicita_raw = st.text_input("10. Solicitante (Nombre):", value=data_previa_j.get("firma_solicita", ""), key=f"txt_firm_sol_{pk_j}")
                    
                    # 11. Autoriza (Nombre y Selección de Cargo)
                    c_aut1, c_aut2 = st.columns(2)
                    autoriza_n_raw = c_aut1.text_input("11. Autoriza (Nombre):", value=data_previa_j.get("autoriza_n", ""), key=f"txt_aut_{pk_j}")
                    aut_opts = ["Vo. Bo. DIRECTOR DE ALUMBRADO PÚBLICO", "Vo. Bo. DIRECTORA DE ALUMBRADO PÚBLICO"]
                    prev_aut = data_previa_j.get("autoriza_c", aut_opts[0])
                    idx_aut = aut_opts.index(prev_aut) if prev_aut in aut_opts else 0
                    autoriza_c_raw = c_aut2.selectbox("Cargo Autoriza:", aut_opts, index=idx_aut, key=f"sel_aut_c_{pk_j}")
                    
                    # 12. Vo. Bo. / Revisa (Nombre y Selección de Cargo)
                    c_vob1, c_vob2 = st.columns(2)
                    revisa_n_raw = c_vob1.text_input("12. Vo. Bo. / Revisa (Nombre):", value=data_previa_j.get("revisa_n", ""), key=f"txt_rev_{pk_j}")
                    rev_opts = ["DELEGADO ADMINISTRATIVO", "DELEGADA ADMINISTRATIVA"]
                    prev_rev = data_previa_j.get("revisa_c", rev_opts[1])
                    idx_rev = rev_opts.index(prev_rev) if prev_rev in rev_opts else 1
                    revisa_c_raw = c_vob2.selectbox("Cargo Revisa:", rev_opts, index=idx_rev, key=f"sel_rev_c_{pk_j}")
                    
                    # 13. Recursos Humanos (Nombre y Selección de Cargo)
                    c_rh1, c_rh2 = st.columns(2)
                    recibe_n_raw = c_rh1.text_input("13. Recursos Humanos (Nombre):", value=data_previa_j.get("recibe_n", ""), key=f"txt_rec_{pk_j}")
                    rec_opts = ["DIRECTOR DE RECURSOS HUMANOS", "DIRECTORA DE RECURSOS HUMANOS"]
                    prev_rec = data_previa_j.get("recibe_c", rec_opts[1])
                    idx_rec = rec_opts.index(prev_rec) if prev_rec in rec_opts else 1
                    recibe_c_raw = c_rh2.selectbox("Cargo RRHH:", rec_opts, index=idx_rec, key=f"sel_rec_c_{pk_j}")

            # --- CONTROL DE DOBLE CAPA INTERCAMBIABLE EN INTERFAZ ---
            h_membrete_j = st.toggle("🛰️ Modo Hoja Membretada (Ocultar Encabezado Impreso)", value=False, key=f"memb_j_{pk_j}")

            # --- FILTRO DE FUERZA BRUTA: PROCESAMIENTO ESTRICTO EN MAYÚSCULAS ---
            solicita = solicita_raw.upper().strip()
            adscrito = adscrito_raw.upper().strip()
            nombre_emp = nombre_emp_raw.upper().strip()
            num_emp = num_emp_raw.upper().strip()
            motivo = motivo_raw.upper().strip()
            
            firma_solicita = firma_solicita_raw.upper().strip() if firma_solicita_raw else solicita
            autoriza_n = autoriza_n_raw.upper().strip()
            autoriza_c = autoriza_c_raw.upper().strip()
            revisa_n = revisa_n_raw.upper().strip()
            revisa_c = revisa_c_raw.upper().strip()
            recibe_n = recibe_n_raw.upper().strip()
            recibe_c = recibe_c_raw.upper().strip()

        with c_preview:
            st.markdown("### 👁️ Vista Previa del Formato")
            
            # --- TRADUCCIÓN DE MESES A ESPAÑOL Y MAYÚSCULAS ---
            meses_mx = {1: "ENERO", 2: "FEBRERO", 3: "MARZO", 4: "ABRIL", 5: "MAYO", 6: "JUNIO", 7: "JULIO", 8: "AGOSTO", 9: "SEPTIEMBRE", 10: "OCTUBRE", 11: "NOVIEMBRE", 12: "DICIEMBRE"}
            f_doc_str = f"{f_doc.day}/{meses_mx[f_doc.month]}/{f_doc.year}"
            f_ini_str = f"DEL {f_inicio.day} DE {meses_mx[f_inicio.month]} DEL {f_inicio.year}"
            f_fin_str = f"AL {f_fin.day} DE {meses_mx[f_fin.month]} DEL {f_fin.year}"

            cod_sel = clave_concepto.split(" | ")[0].strip()
            
            if accion == "JUSTIFICAR":
                just_line = f"[{cod_sel}] A: {nombre_emp}"
                sanc_line = ""
            else:
                just_line = ""
                sanc_line = f"[{cod_sel}] A: {nombre_emp}"

            bg_reg_lista = "background: #bcbcbc; font-weight: bold;" if f_registro == "LISTA" else ""
            bg_reg_hp = "background: #bcbcbc; font-weight: bold;" if f_registro == "HAND PUNCH" else ""

            conceptos_grid = [
                ("3", "FALTA INJUSTIFICADA", "17", "LICENCIA POR GRAVIDEZ"),
                ("4", "FALTA JUSTIFICADA (TIEMPO X TIEMPO)", "18", "HORA DE LACTANCIA"),
                ("9", "LICENCIA CON GOCE DE SUELDO", "19", "LICENCIA POR FALLECIMIENTO DE FAMILIAR"),
                ("10", "LICENCIA SIN GOCE DE SUELDO", "20", "LICENCIA POR NACIMIENTO"),
                ("11", "VACACIONES", "23", "OMISIÓN DE CHECADA"),
                ("12", "INCAPACIDAD", "24", "DÍA ECONÓMICO (SINDICALIZADO)"),
                ("14", "COMISIÓN", "34", "CUMPLEAÑOS (SINDICALIZADO)"),
                ("16", "LICENCIA POR MATRIMONIO(SINDICALIZADO)", "CM", "CUIDADOS MÉDICOS")
            ]

            tabla_html_conceptos = ""
            for c1, n1, c2, n2 in conceptos_grid:
                style_c1 = "background: #bcbcbc; font-weight: bold;" if cod_sel == c1 else ""
                style_c2 = "background: #bcbcbc; font-weight: bold;" if cod_sel == c2 else ""
                
                tabla_html_conceptos += f"""
                <tr>
                    <td style="border: 1px solid black; text-align: center; width: 8%; font-size: 10px; {style_c1}">{c1}</td>
                    <td style="border: 1px solid black; padding-left: 5px; width: 42%; font-size: 9px; {style_c1}">{n1}</td>
                    <td style="border: 1px solid black; text-align: center; width: 8%; font-size: 10px; {style_c2}">{c2}</td>
                    <td style="border: 1px solid black; padding-left: 5px; width: 42%; font-size: 9px; {style_c2}">{n2}</td>
                </tr>
                """

            html_formato = f"""
            <div style="background: white; color: black; padding: 25px; border: 1px solid #aaa; font-family: 'Arial', sans-serif; line-height: 1.3; width: 100%; box-sizing: border-box;">
                <!-- CAPA HTML 1: MEMBRETE DINÁMICO -->
                <table style="width: 100%; border-collapse: collapse; margin-bottom: 5px; {'visibility: hidden; opacity: 0;' if h_membrete_j else ''}">
                    <tr>
                        <td style="width: 25%; font-size: 22px; font-family: 'Arial Black', Gadget, sans-serif; font-weight: 900; vertical-align: top; line-height: 0.9;">
                            Toluca<br><span style="font-size: 9px; font-family: Arial; letter-spacing: 3px; font-weight: bold;">CAPITAL</span>
                            <div style="font-size: 6px; letter-spacing: 0.5px; font-weight: normal; margin-top: 2px; color:#555;">DE OPORTUNIDADES Y PROGRESO</div>
                        </td>
                        <td style="text-align: center; width: 55%; vertical-align: top; font-size: 11px; font-weight: bold; font-family: Arial;">
                            DIRECCIÓN GENERAL DE SERVICIOS PÚBLICOS<br>
                            DIRECCIÓN DE ALUMBRADO PÚBLICO<br><br>
                            <span style="font-size: 9.5px; font-weight: normal; font-style: italic;">"2026. Año del Humanismo Mexicano en el Estado de México"</span>
                        </td>
                        <td style="width: 20%; text-align: right; vertical-align: top; font-size: 8px; color: #777;">
                            <div style="border: 1px dashed #999; padding: 5px; text-align: center; height: 35px;">ESCUDO TIMBRE</div>
                        </td>
                    </tr>
                </table>
                
                <!-- CAPA HTML 2: CUERPO FIJO -->
                <div style="text-align: center; font-weight: bold; font-size: 13px; margin-top: 10px; margin-bottom: 15px; letter-spacing: 0.5px;">FORMATO ÚNICO DE JUSTIFICACIÓN</div>
                
                <table style="width: 100%; border-collapse: collapse; margin-bottom: 8px; font-size: 11px;">
                    <tr>
                        <td style="font-weight: bold; width: 12%; padding: 4px 0;">SOLICITA:</td>
                        <td style="border-bottom: 1px solid black; color: blue; width: 53%; padding: 4px 0;">{solicita}</td>
                        <td style="font-weight: bold; width: 10%; text-align: right; padding: 4px 0;">FECHA:</td>
                        <td style="border-bottom: 1px solid black; color: blue; text-align: center; width: 25%; padding: 4px 0;">{f_doc_str}</td>
                    </tr>
                    <tr>
                        <td style="font-weight: bold; padding: 4px 0;">JUSTIFICAR:</td>
                        <td style="border-bottom: 1px solid black; color: blue; padding: 4px 0;" colspan="3">{just_line}</td>
                    </tr>
                    <tr>
                        <td style="font-weight: bold; padding: 4px 0;">SANCIONAR:</td>
                        <td style="border-bottom: 1px solid black; color: blue; padding: 4px 0;">{sanc_line}</td>
                        <td style="font-weight: bold; text-align: right; padding: 4px 0;">No. DE EMP.:</td>
                        <td style="border-bottom: 1px solid black; color: blue; text-align: center; padding: 4px 0;">{num_emp}</td>
                    </tr>
                    <tr>
                        <td style="font-weight: bold; padding: 4px 0;">ADSCRITO A:</td>
                        <td style="border-bottom: 1px solid black; color: blue; padding: 4px 0;">{adscrito}</td>
                        <td style="font-weight: bold; text-align: right; padding: 4px 0;" colspan="2">
                            F. REGISTRO &nbsp;
                            <span style="border: 1px solid black; padding: 1px 4px; margin-right: 3px; {bg_reg_lista}">LISTA</span>
                            <span style="border: 1px solid black; padding: 1px 4px; {bg_reg_hp}">HAND PUNCH</span>
                        </td>
                    </tr>
                </table>
                
                <table style="width: 100%; border-collapse: collapse; margin-top: 10px; margin-bottom: 8px;">
                    <thead>
                        <tr style="background: #e1e1e1; font-weight: bold; font-size: 10px; text-align: center;">
                            <td style="border: 1px solid black; padding: 4px; width: 8%;">CLAVE</td>
                            <td style="border: 1px solid black; padding: 4px; width: 42%;">CONCEPTO</td>
                            <td style="border: 1px solid black; padding: 4px; width: 8%;">CLAVE</td>
                            <td style="border: 1px solid black; padding: 4px; width: 42%;">CONCEPTO</td>
                        </tr>
                    </thead>
                    <tbody>
                        {tabla_html_conceptos}
                    </tbody>
                </table>
                
                <table style="width: 100%; border-collapse: collapse; border: 1px solid black; font-size: 11px; margin-bottom: 10px; text-align: center;">
                    <tr>
                        <td style="font-weight: bold; text-align: left; padding: 4px; border-bottom: 1px solid black; background: #f5f5f5;" colspan="2">&nbsp;FECHA:</td>
                    </tr>
                    <tr>
                        <td style="width: 50%; padding: 8px; border-right: 1px solid black; color: red; font-weight: bold; font-size: 11.5px;">{f_ini_str}</td>
                        <td style="width: 50%; padding: 8px; color: red; font-weight: bold; font-size: 11.5px;">{f_fin_str}</td>
                    </tr>
                </table>
                
                <div style="border: 1px solid black; margin-bottom: 15px; font-size: 11px;">
                    <div style="font-weight: bold; text-align: center; background: #e1e1e1; padding: 3px; border-bottom: 1px solid black; letter-spacing: 1px;">MOTIVO</div>
                    <div style="padding: 10px; text-align: center; min-height: 40px; font-weight: bold; line-height: 1.4; color: #333;">
                        {motivo if motivo else "ASUNTO OPERATIVO ASIGNADO EN CAMPO."}
                    </div>
                </div>
                
                <table style="width: 100%; border-collapse: collapse; text-align: center; font-size: 9.5px; margin-top: 25px; margin-bottom: 15px;">
                    <tr>
                        <td style="width: 23%; vertical-align: bottom; padding: 0 5px;">
                            <div style="border-bottom: 1px solid black; padding-bottom: 3px; font-weight: bold; color: blue;">{firma_solicita}</div>
                            <div style="margin-top: 4px; font-weight: bold; font-size: 9px; height: 24px; display: flex; align-items: center; justify-content: center;">SOLICITANTE</div>
                        </td>
                        <td style="width: 23%; vertical-align: bottom; padding: 0 5px;">
                            <div style="border-bottom: 1px solid black; padding-bottom: 3px; font-weight: bold; color: blue;">{autoriza_n}</div>
                            <div style="margin-top: 4px; font-weight: bold; font-size: 9px; height: 24px; display: flex; align-items: center; justify-content: center; line-height: 1.1;">{autoriza_c}</div>
                        </td>
                        <td style="width: 23%; vertical-align: bottom; padding: 0 5px;">
                            <div style="border-bottom: 1px solid black; padding-bottom: 3px; font-weight: bold; color: blue;">{revisa_n}</div>
                            <div style="margin-top: 4px; font-weight: bold; font-size: 9px; height: 24px; display: flex; align-items: center; justify-content: center; line-height: 1.1;">{revisa_c}</div>
                        </td>
                        <td style="width: 23%; vertical-align: bottom; padding: 0 5px;">
                            <div style="border-bottom: 1px solid black; padding-bottom: 3px; font-weight: bold; color: blue;">{recibe_n}</div>
                            <div style="margin-top: 4px; font-weight: bold; font-size: 9px; height: 24px; display: flex; align-items: center; justify-content: center; line-height: 1.1;">{recibe_c}</div>
                        </td>
                    </tr>
                </table>
                
                <div style="border-top: 1px solid black; text-align: center; font-size: 10px; font-weight: bold; padding-top: 3px; margin-top: 10px;">
                    H. Ayuntamiento de Toluca
                </div>
                <div style="text-align: center; font-size: 8px; color: #444; margin-top: 2px;">
                    Rafael Alducin s/n esquina Primero de Mayo, Col. Reforma y Ferrocarriles | Tel: 7223171747
                </div>
                <div style="background: black; height: 10px; width: 100%; margin-top: 6px;"></div>
            </div>
            """
            
            st.components.v1.html(html_formato, height=800, scrolling=True)
            st.divider()
            
            ejecutar_guardado_j = False
            if modo_j == "📂 Consultar Bóveda":
                seguro_actualizar_j = st.checkbox("🔐 Confirmar actualización del formato guardado", key="chk_seguro_actualizar_incidencias")
                if st.button("🔄 ACTUALIZAR JUSTIFICACIÓN HISTÓRICA", use_container_width=True, disabled=not seguro_actualizar_j, type="primary"):
                    ejecutar_guardado_j = True
            else:
                if st.button("💾 GUARDAR NUEVA JUSTIFICACIÓN", use_container_width=True, type="primary"):
                    ejecutar_guardado_j = True

            if ejecutar_guardado_j:
                tz_mx = timezone(timedelta(hours=-6))
                ahora = datetime.now(tz_mx)
                
                if modo_j == "✨ Crear Nuevo":
                    id_reg_j = f"SF4-JST-{ahora.strftime('%Y%m%d-%H%M%S')}"
                else:
                    id_reg_j = id_sel_j
                    
                fecha_mx = ahora.strftime("%d/%m/%Y %H:%M:%S")
                
                payload_j = {
                    "f_registro": f_registro, "accion": accion, "solicita": solicita,
                    "adscrito": adscrito, "nombre": nombre_emp, "num_emp": num_emp,
                    "tipo_fecha": tipo_fecha_j, "fecha_inicio": str(f_inicio), "fecha_fin": str(f_fin),
                    "clave_concepto": clave_concepto, "motivo": motivo, "revisa_n": revisa_n,
                    "autoriza_n": autoriza_n, "recibe_n": recibe_n, "fecha_doc": str(f_doc),
                    "firma_solicita": firma_solicita, "autoriza_c": autoriza_c, "revisa_c": revisa_c, "recibe_c": recibe_c
                }
                
                st.session_state.db_justificaciones[id_reg_j] = payload_j
                
                try:
                    cell = ws.find(id_reg_j, in_column=1)
                    if cell: ws.delete_rows(cell.row)
                except: pass
                
                ws.append_row([id_reg_j, fecha_mx, nombre_emp, num_emp, json.dumps(payload_j), ""])
                st.success("✅ Formato Único de Justificación Sincronizado en la Nube."); time.sleep(1); st.rerun()

            # --- CONSTRUCTOR DEL DOCUMENTO PDF OFICIAL (MÁXIMA FIDELIDAD IMPRESA) ---
            if motor_pdf_listo:
                X_START = 24.0  # Control rígido lateral izquierdo
                W_TOTAL = 178.0  # Ancho de celdas estandarizado
                
                pdf_j = FPDF(orientation='P', unit='mm', format='Letter')
                pdf_j.set_margins(X_START, 26, 12)
                pdf_j.set_auto_page_break(auto=False)
                pdf_j.add_page()
                
                # Filtro defensivo anti-crash para caracteres especiales
                solicita_enc = solicita.encode('latin-1', 'replace').decode('latin-1')
                nombre_emp_enc = nombre_emp.encode('latin-1', 'replace').decode('latin-1')
                just_line_enc = just_line.encode('latin-1', 'replace').decode('latin-1')
                sanc_line_enc = sanc_line.encode('latin-1', 'replace').decode('latin-1')
                adscrito_enc = adscrito.encode('latin-1', 'replace').decode('latin-1')
                motivo_enc = (motivo if motivo else "ASUNTO OPERATIVO ASIGNADO EN CAMPO.").encode('latin-1', 'replace').decode('latin-1')
                
                firma_solicita_enc = firma_solicita.encode('latin-1', 'replace').decode('latin-1')
                autoriza_n_enc = autoriza_n.encode('latin-1', 'replace').decode('latin-1')
                revisa_n_enc = revisa_n.encode('latin-1', 'replace').decode('latin-1')
                recibe_n_enc = recibe_n.encode('latin-1', 'replace').decode('latin-1')
                
                autoriza_c_enc = autoriza_c.encode('latin-1', 'replace').decode('latin-1')
                revisa_c_enc = revisa_c.encode('latin-1', 'replace').decode('latin-1')
                recibe_c_enc = recibe_c.encode('latin-1', 'replace').decode('latin-1')
                
                # ==========================================
                # 🛰️ CAPA 1: LAYER HEADER (MEMBRETE OPCIONAL)
                # ==========================================
                if not h_membrete_j:
                    pdf_j.set_font("Arial", 'B', 18)
                    pdf_j.set_xy(X_START, 20)
                    pdf_j.cell(45, 6, txt="Toluca", ln=False)
                    
                    pdf_j.set_font("Arial", 'B', 9)
                    pdf_j.set_xy(X_START + 45, 20)
                    pdf_j.cell(95, 4, txt="DIRECCIÓN GENERAL DE SERVICIOS PÚBLICOS", ln=False, align='C')
                    
                    pdf_j.set_font("Arial", '', 8)
                    pdf_j.set_xy(X_START + 140, 20)
                    pdf_j.cell(38, 4, txt="[ TIMBRE ]", ln=True, align='R')
                    
                    pdf_j.set_font("Arial", 'B', 8)
                    pdf_j.set_xy(X_START, 27)
                    pdf_j.cell(45, 4, txt="CAPITAL DE OPORTUNIDADES", ln=False)
                    
                    pdf_j.set_font("Arial", 'B', 9)
                    pdf_j.set_xy(X_START + 45, 27)
                    pdf_j.cell(95, 4, txt="DIRECCIÓN DE ALUMBRADO PÚBLICO", ln=True, align='C')
                    
                    pdf_j.set_font("Arial", 'I', 8.5)
                    pdf_j.set_xy(X_START, 35)
                    pdf_j.cell(W_TOTAL, 4, txt='"2026. Año del Humanismo Mexicano en el Estado de México"', ln=True, align='C')
                
                # ==========================================
                # 🏗️ CAPA 2: LAYER BODY (CUERPO ANCLADO INMUTABLE)
                # ==========================================
                Y_START_BODY = 46.0  # El ancla inmutable del documento
                
                pdf_j.set_font("Arial", 'B', 11)
                pdf_j.set_xy(X_START, Y_START_BODY)
                pdf_j.cell(W_TOTAL, 5, txt="FORMATO ÚNICO DE JUSTIFICACIÓN", ln=True, align='C')
                
                # Renglón 1: Solicita y Fecha
                pdf_j.set_font("Arial", 'B', 10)
                pdf_j.set_xy(X_START, Y_START_BODY + 8)
                pdf_j.cell(22, 7, txt="SOLICITA: ", border=0, ln=False)
                pdf_j.set_font("Arial", '', 10)
                pdf_j.cell(105, 7, txt=solicita_enc, border='B', ln=False)
                pdf_j.set_font("Arial", 'B', 10)
                pdf_j.cell(18, 7, txt=" FECHA: ", border=0, ln=False)
                pdf_j.set_font("Arial", '', 10)
                pdf_j.cell(33, 7, txt=f_doc_str, border='B', ln=True, align='C')
                
                # Renglón 2: Justificar
                pdf_j.set_font("Arial", 'B', 10)
                pdf_j.set_xy(X_START, Y_START_BODY + 16)
                pdf_j.cell(26, 7, txt="JUSTIFICAR: ", border=0, ln=False)
                pdf_j.set_font("Arial", '', 10)
                pdf_j.cell(152, 7, txt=just_line_enc, border='B', ln=True)
                
                # Renglón 3: Sancionar y No. de Empleado
                pdf_j.set_font("Arial", 'B', 10)
                pdf_j.set_xy(X_START, Y_START_BODY + 24)
                pdf_j.cell(26, 7, txt="SANCIONAR: ", border=0, ln=False)
                pdf_j.set_font("Arial", '', 10)
                pdf_j.cell(93, 7, txt=sanc_line_enc, border='B', ln=False)
                pdf_j.set_font("Arial", 'B', 10)
                pdf_j.cell(26, 7, txt=" No. DE EMP.: ", border=0, ln=False)
                pdf_j.set_font("Arial", '', 10)
                pdf_j.cell(33, 7, txt=num_emp, border='B', ln=True, align='C')
                
                # Renglón 4: Adscrito y Tipo de Registro
                pdf_j.set_font("Arial", 'B', 10)
                pdf_j.set_xy(X_START, Y_START_BODY + 32)
                pdf_j.cell(26, 7, txt="ADSCRITO A: ", border=0, ln=False)
                pdf_j.set_font("Arial", '', 10)
                pdf_j.cell(93, 7, txt=adscrito_enc, border='B', ln=False)
                pdf_j.set_font("Arial", 'B', 8.5)
                pdf_j.cell(26, 7, txt=" F. REGISTRO: ", border=0, ln=False)
                
                fill_lista = (f_registro == "LISTA")
                pdf_j.set_fill_color(188, 188, 188)
                pdf_j.cell(14, 5.5, txt="LISTA", border=1, ln=False, align='C', fill=fill_lista)
                pdf_j.cell(2, 5.5, txt="", border=0, ln=False)
                fill_hp = (f_registro == "HAND PUNCH")
                pdf_j.cell(17, 5.5, txt="H.P.", border=1, ln=True, align='C', fill=fill_hp)
                
                # --- TABLA DE CONCEPTOS TOTALMENTE GEOMETRIZADA ---
                Y_TABLE = Y_START_BODY + 43
                pdf_j.set_font("Arial", 'B', 9)
                pdf_j.set_fill_color(225, 225, 225)
                pdf_j.set_xy(X_START, Y_TABLE)
                pdf_j.cell(12, 5.5, txt="CLAVE", border=1, ln=False, align='C', fill=True)
                pdf_j.cell(77, 5.5, txt="CONCEPTO", border=1, ln=False, align='C', fill=True)
                pdf_j.cell(12, 5.5, txt="CLAVE", border=1, ln=False, align='C', fill=True)
                pdf_j.cell(77, 5.5, txt="CONCEPTO", border=1, ln=True, align='C', fill=True)
                
                pdf_j.set_font("Arial", '', 8)
                current_y = Y_TABLE + 5.5
                for c1, n1, c2, n2 in conceptos_grid:
                    fill_l = (cod_sel == c1)
                    fill_r = (cod_sel == c2)
                    pdf_j.set_fill_color(188, 188, 188)
                    
                    t_n1 = f" {n1}".encode('latin-1', 'replace').decode('latin-1')
                    t_n2 = f" {n2}".encode('latin-1', 'replace').decode('latin-1')
                    
                    pdf_j.set_xy(X_START, current_y)
                    pdf_j.cell(12, 6.2, txt=c1, border=1, ln=False, align='C', fill=fill_l)
                    pdf_j.cell(77, 6.2, txt=t_n1, border=1, ln=False, fill=fill_l)
                    pdf_j.cell(12, 6.2, txt=c2, border=1, ln=False, align='C', fill=fill_r)
                    pdf_j.cell(77, 6.2, txt=t_n2, border=1, ln=True, fill=fill_r)
                    current_y += 6.2
                    
                # Recuadro Periodo / Fechas
                Y_FECHAS = current_y + 3.0
                pdf_j.set_font("Arial", 'B', 9.5)
                pdf_j.set_fill_color(245, 245, 245)
                pdf_j.set_xy(X_START, Y_FECHAS)
                pdf_j.cell(W_TOTAL, 5.5, txt=" FECHA:", border=1, ln=True, fill=True)
                
                pdf_j.set_font("Arial", 'B', 10.5)
                pdf_j.set_text_color(220, 0, 0)
                pdf_j.set_xy(X_START, Y_FECHAS + 5.5)
                pdf_j.cell(89, 9, txt=f_ini_str, border=1, ln=False, align='C')
                pdf_j.cell(89, 9, txt=f_fin_str, border=1, ln=True, align='C')
                pdf_j.set_text_color(0, 0, 0)
                
                # Recuadro Ampliado de Motivo
                Y_MOTIVO = Y_FECHAS + 17.5
                pdf_j.set_font("Arial", 'B', 9.5)
                pdf_j.set_fill_color(225, 225, 225)
                pdf_j.set_xy(X_START, Y_MOTIVO)
                pdf_j.cell(W_TOTAL, 5.5, txt="MOTIVO", border=1, ln=True, align='C', fill=True)
                
                pdf_j.rect(X_START, Y_MOTIVO + 5.5, W_TOTAL, 22)
                pdf_j.set_xy(X_START, Y_MOTIVO + 7.5)
                pdf_j.multi_cell(W_TOTAL, 5, txt=motivo_enc, align='C')
                
                # --- GRID RIGIDO DE FIRMAS ---
                Y_FIRMAS = Y_MOTIVO + 32.0  
                h_grid = 54
                w_col = W_TOTAL / 4
                
                pdf_j.rect(X_START, Y_FIRMAS, W_TOTAL, h_grid)
                pdf_j.line(X_START + w_col, Y_FIRMAS, X_START + w_col, Y_FIRMAS + h_grid)
                pdf_j.line(X_START + 2*w_col, Y_FIRMAS, X_START + 2*w_col, Y_FIRMAS + h_grid)
                pdf_j.line(X_START + 3*w_col, Y_FIRMAS, X_START + 3*w_col, Y_FIRMAS + h_grid)
                
                pdf_j.line(X_START, Y_FIRMAS + 36, X_START + W_TOTAL, Y_FIRMAS + 36)
                pdf_j.line(X_START, Y_FIRMAS + 45, X_START + W_TOTAL, Y_FIRMAS + 45)
                
                pdf_j.set_font("Arial", 'B', 8.5)
                pdf_j.set_xy(X_START + w_col, Y_FIRMAS + 2)
                pdf_j.cell(w_col, 4, txt="AUTORIZACIÓN", align='C')
                pdf_j.set_xy(X_START + 2*w_col, Y_FIRMAS + 2)
                pdf_j.cell(w_col, 4, txt="Vo. Bo.", align='C')
                
                pdf_j.set_font("Arial", 'B', 8)
                pdf_j.set_xy(X_START, Y_FIRMAS + 37.5)
                pdf_j.multi_cell(w_col, 3.2, txt=firma_solicita_enc, align='C')
                pdf_j.set_xy(X_START + w_col, Y_FIRMAS + 37.5)
                pdf_j.multi_cell(w_col, 3.2, txt=autoriza_n_enc, align='C')
                pdf_j.set_xy(X_START + 2*w_col, Y_FIRMAS + 37.5)
                pdf_j.multi_cell(w_col, 3.2, txt=revisa_n_enc, align='C')
                pdf_j.set_xy(X_START + 3*w_col, Y_FIRMAS + 37.5)
                pdf_j.multi_cell(w_col, 3.2, txt=recibe_n_enc, align='C')
                
                pdf_j.set_font("Arial", 'B', 7.5)
                pdf_j.set_xy(X_START, Y_FIRMAS + 46.5)
                pdf_j.multi_cell(w_col, 3, txt="SOLICITANTE", align='C')
                pdf_j.set_xy(X_START + w_col, Y_FIRMAS + 46.5)
                pdf_j.multi_cell(w_col, 3, txt=autoriza_c_enc, align='C')
                pdf_j.set_xy(X_START + 2*w_col, Y_FIRMAS + 46.5)
                pdf_j.multi_cell(w_col, 3, txt=revisa_c_enc, align='C')
                pdf_j.set_xy(X_START + 3*w_col, Y_FIRMAS + 46.5)
                pdf_j.multi_cell(w_col, 3, txt=recibe_c_enc, align='C')
                
                # Pie de página institucional inamovible
                pdf_j.set_xy(X_START, Y_FIRMAS + 56)
                pdf_j.set_font("Arial", 'B', 9)
                pdf_j.cell(W_TOTAL, 4, txt="H. Ayuntamiento de Toluca", ln=True, align='C')
                pdf_j.set_font("Arial", '', 7.5)
                pdf_j.cell(W_TOTAL, 3, txt="Rafael Alducin s/n esquina Primero de Mayo, Col. Reforma y Ferrocarriles | Tel: 7223171747", ln=True, align='C')
                pdf_j.set_fill_color(0, 0, 0)
                pdf_j.rect(X_START, pdf_j.get_y() + 1.5, W_TOTAL, 2.5, 'F')
                
                pdf_data_j = pdf_j.output(dest='S').encode('latin-1', 'replace')
                st.download_button(label="🚀 DESCARGAR JUSTIFICACIÓN PDF", data=pdf_data_j, file_name=f"Justificacion_{num_emp}_{f_inicio.strftime('%Y%m%d')}.pdf", mime="application/pdf", use_container_width=True)
            else:
                st.error("❌ Función PDF no disponible.")
    elif st.session_state.menu == "SF5":
        st.title("🛡️ SF5 - Centro de Depuración Inteligente")

        # --- 1. CONEXIÓN A GOOGLE SHEETS ---
        scope = ["https://www.googleapis.com/auth/spreadsheets"]
        creds_dict = {
            "type": st.secrets["connections"]["gsheets"]["type"],
            "project_id": st.secrets["connections"]["gsheets"]["project_id"],
            "private_key_id": st.secrets["connections"]["gsheets"]["private_key_id"],
            "private_key": st.secrets["connections"]["gsheets"]["private_key"].replace('\\n', '\n'),
            "client_email": st.secrets["connections"]["gsheets"]["client_email"],
            "client_id": st.secrets["connections"]["gsheets"]["client_id"],
            "auth_uri": st.secrets["connections"]["gsheets"]["auth_uri"],
            "token_uri": st.secrets["connections"]["gsheets"]["token_uri"],
            "auth_provider_x509_cert_url": st.secrets["connections"]["gsheets"]["auth_provider_x509_cert_url"],
            "client_x509_cert_url": st.secrets["connections"]["gsheets"]["client_x509_cert_url"]
        }
        creds = Credentials.from_service_account_info(creds_dict, scopes=scope)
        client = gspread.authorize(creds)
        SHEET_ID = "14_fewol5DiFXoiO102wviiWR08Lw3PKHzEjSbMwxUm8"
        
        try:
            sh = client.open_by_key(SHEET_ID)
            ws = sh.worksheet("Boveda_Bajas")
        except:
            ws = sh.get_worksheet(0)

        # --- SINCRO-BÓVEDA RESILIENTE (ANTI-DESCONEXIONES) ---
        bovedas_completas = ["db_oficios", "boveda_mmd", "db_depuracion", "vales_historial"]
        if any(b not in st.session_state for b in bovedas_completas):
            for _ in range(3):
                try:
                    if "db_oficios" not in st.session_state: st.session_state.db_oficios = {}
                    if "boveda_mmd" not in st.session_state: st.session_state.boveda_mmd = {}
                    if "db_depuracion" not in st.session_state: st.session_state.db_depuracion = {}
                    if "vales_historial" not in st.session_state: st.session_state.vales_historial = []
                    
                    # Lectura directa por matriz de celdas para blindar el Módulo 5
                    filas_raw = ws.get_all_values()
                    if len(filas_raw) > 1:
                        for row in filas_raw[1:]: # Saltamos los encabezados de la fila 1
                            if len(row) >= 5:
                                reg_id = str(row[0]).strip()
                                origen = str(row[2]).strip() if str(row[2]).strip() else "Sin Nombre"
                                datos_raw = str(row[4]).strip() if row[4] else "{}"
                                
                                if reg_id.startswith("SF4-PRY-"):
                                    try: st.session_state.boveda_mmd[origen] = json.loads(datos_raw) if isinstance(datos_raw, str) else datos_raw
                                    except: pass
                                elif reg_id.startswith("SF4-OFC-"):
                                    try: st.session_state.db_oficios[origen] = json.loads(datos_raw) if isinstance(datos_raw, str) else datos_raw
                                    except: pass
                                elif reg_id.startswith("SF5-DEP-"):
                                    try: st.session_state.db_depuracion[reg_id] = json.loads(datos_raw) if isinstance(datos_raw, str) else datos_raw
                                    except: pass
                                elif reg_id.startswith("SF6-VAL-"):
                                    try:
                                        vale_parsed = json.loads(datos_raw) if isinstance(datos_raw, str) else datos_raw
                                        if vale_parsed not in st.session_state.vales_historial:
                                            st.session_state.vales_historial.append(vale_parsed)
                                    except: pass
                    break
                except Exception:
                    time.sleep(1)

        # Inicialización de variables de sesión para la mesa de trabajo actual
        if "da_actual" not in st.session_state: st.session_state.da_actual = None
        if "h1_actual" not in st.session_state: st.session_state.h1_actual = None
        if "h2_actual" not in st.session_state: st.session_state.h2_actual = None
        if "tipo_depuracion_actual" not in st.session_state: st.session_state.tipo_depuracion_actual = ""

        # --- MOTOR DE PROCESAMIENTO REFINADO ---
        def motor_sf5(df_total):
            df_total = df_total.astype(str).fillna("")
            
            # Extract GPS
            def motor_gps_v25(fila):
                texto = " ".join(fila.values).lower()
                numeros = re.findall(r'-?\d+\.\d{4,}', texto)
                if len(numeros) >= 2: return float(numeros[0]), float(numeros[1])
                return None, None

            coords = df_total.apply(motor_gps_v25, axis=1)
            df_total['lat_aux'] = [c[0] for c in coords]
            df_total['lon_aux'] = [c[1] for c in coords]
            df_total = df_total.dropna(subset=['lat_aux', 'lon_aux']).reset_index(drop=True)
            df_total['Grupo_Duplicado'] = 0
            
            if df_total.empty: return None, None, None

            # Cluster detection (20 meters)
            umbral = 20 / 111111.0
            coords_arr = df_total[['lat_aux', 'lon_aux']].values
            marcador = [0] * len(df_total)
            color_id = 1
            for i in range(len(coords_arr)):
                if marcador[i] != 0: continue
                encontrado = False
                for j in range(i + 1, len(coords_arr)):
                    if np.linalg.norm(coords_arr[i] - coords_arr[j]) < umbral:
                        marcador[j] = color_id
                        encontrado = True
                if encontrado:
                    marcador[i] = color_id
                    color_id += 1
            
            df_total['Grupo_Duplicado'] = marcador
            
            # --- SEPARATION LOGIC ---
            indices_reps = []
            seen_groups = set()
            for idx, row in df_total.iterrows():
                gid = row['Grupo_Duplicado']
                if gid == 0 or gid not in seen_groups:
                    indices_reps.append(idx)
                    if gid > 0: seen_groups.add(gid)
            
            h1 = df_total.loc[indices_reps].copy()
            h2 = df_total[(df_total['Grupo_Duplicado'] > 0) & (~df_total.index.isin(indices_reps))].copy()
            
            return df_total, h1, h2

        # --- INTERFACE DE DASHBOARD Y PRODUCTO FINAL ---
        def renderizar_interfaz(da, h1, h2, suffix):
            st.markdown("### 📈 Dashboard de Depuración SF5")
            m_cols = st.columns(5)
            metricas = [
                ("🔍 PROCESADOS", len(da), "#1f4e78"),
                ("🚨 DUPLICADOS", len(h2), "#e67e22"),
                ("🗑️ REMOVIDOS", len(h2), "#95a5a6"),
                ("✅ ÚNICOS (H1)", len(h1), "#28a745"),
                ("⏱️ AHORRO", f"{len(h2) * 5} min", "#dc3545")
            ]
            for col, (label, val, colr) in zip(m_cols, metricas):
                col.markdown(f"<div style='text-align: center; background-color: #f0f2f6; padding: 10px; border-radius: 10px; border-left: 5px solid {colr};'><b style='font-size: 11px;'>{label}</b><br><span style='font-size: 18px;'>{val}</span></div>", unsafe_allow_html=True)

            out = io.BytesIO()
            from openpyxl.styles import PatternFill
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

            with pd.ExcelWriter(out, engine='openpyxl') as w:
                h1.to_excel(w, index=False, sheet_name='PARA_MODULO_1')
                ws_sheet = w.sheets['PARA_MODULO_1']
                for r_num, (_, row) in enumerate(h1.iterrows(), 2):
                    if int(row['Grupo_Duplicado']) > 0:
                        for cell in ws_sheet[r_num]: cell.fill = yellow_fill
                
                h2.to_excel(w, index=False, sheet_name='REPORTE_DUPLICADOS')
            
            excel_data = out.getvalue()
            
            st.write("---")
            c_down, c_trans = st.columns(2)
            c_down.download_button("🚀 DESCARGAR PRODUCTO FINAL v25", excel_data, "SF_PANGEA_DEPURADO.xlsx", use_container_width=True, key=f"dl_{suffix}")
            
            if c_trans.button("➡️ ENVIAR AL GENERADOR (SF1)", use_container_width=True, type="primary", key=f"btn_{suffix}"):
                st.session_state.df_transferido = h1.copy()
                st.session_state.nombre_archivo_transferido = "DEPURADO_SF5.xlsx"
                st.session_state.menu = "SF1"
                st.rerun()

            st.write("---")
            st.subheader("💾 Guardado en Bóveda Histórica Nube")
            col_txt_nom, col_btn_json = st.columns([2.5, 1.5])
            nombre_depuracion = col_txt_nom.text_input("Asigna un nombre a esta depuración para tu archivo permanente:", placeholder="Ej: Depuración Oriente 01/06", key=f"txt_save_{suffix}")
            
            if col_btn_json.button("📁 Guardar en Historial", use_container_width=True, key=f"save_bov_{suffix}"):
                if nombre_depuracion.strip():
                    # --- CONTROL DE DUPLICADOS EN NOMBRES ---
                    nombre_repetido = any(v.get("nombre", "").strip().upper() == nombre_depuracion.strip().upper() for v in st.session_state.db_depuracion.values())
                    
                    if nombre_repetido:
                        st.error(f"⚠️ El nombre '{nombre_depuracion}' ya existe en el registro permanente de la Dirección. Use una nomenclatura diferente.")
                    else:
                        tz_mx = timezone(timedelta(hours=-6))
                        ahora = datetime.now(tz_mx)
                        id_dep = f"SF5-DEP-{ahora.strftime('%Y%m%d-%H%M%S')}"
                        fecha_mx = ahora.strftime("%d/%m/%Y %H:%M:%S")
                        
                        payload = {
                            "nombre": nombre_depuracion.strip(),
                            "fecha": fecha_mx,
                            "tipo": st.session_state.tipo_depuracion_actual,
                            "procesados": len(da),
                            "duplicados": len(h2),
                            "unicos": len(h1),
                            "da_json": da.to_json(orient='split'),
                            "h1_json": h1.to_json(orient='split'),
                            "h2_json": h2.to_json(orient='split')
                        }
                        
                        st.session_state.db_depuracion[id_dep] = payload
                        ws.append_row([id_dep, fecha_mx, nombre_depuracion.strip(), len(da), json.dumps(payload), ""])
                        st.success(f"✅ Depuración guardada con éxito en Google Sheets. ID: {id_dep}")
                        time.sleep(0.5)
                        st.rerun()
                else:
                    st.warning("⚠️ Ingresa un nombre para poder guardar el reporte físico.")

        # --- ARQUITECTURA DE PESTAÑAS (TABS) ---
        tab_multi, tab_auditoria, tab_boveda = st.tabs(["🔄 Comparar Varios Archivos", "🔍 Auditoría Interna (1 archivo)", "🗄️ Bóveda de Historial Permanente"])
        
        with tab_multi:
            f_in = st.file_uploader("📂 Archivos", accept_multiple_files=True, key="m_in_sf5")
            if f_in:
                if st.button("⚡ Ejecutar Depuración Masiva", use_container_width=True, type="primary"):
                    try:
                        dfs = []
                        for f in f_in:
                            df = pd.read_excel(f, dtype=str) if f.name.endswith('.xlsx') else pd.read_csv(f, encoding='latin-1', dtype=str)
                            df['archivo_origen'] = f.name
                            dfs.append(df)
                        da, h1, h2 = motor_sf5(pd.concat(dfs, ignore_index=True))
                        if da is not None:
                            st.session_state.da_actual = da
                            st.session_state.h1_actual = h1
                            st.session_state.h2_actual = h2
                            st.session_state.tipo_depuracion_actual = "Masiva (Varios Archivos)"
                            st.rerun()
                    except Exception as e:
                        st.error(f"Error al procesar archivos masivos: {e}")

        with tab_auditoria:
            f_in_s = st.file_uploader("📂 Archivo Único", accept_multiple_files=False, key="s_in_sf5")
            if f_in_s:
                if st.button("⚡ Ejecutar Auditoría Única", use_container_width=True, type="primary"):
                    try:
                        df = pd.read_excel(f_in_s, dtype=str) if f_in_s.name.endswith('.xlsx') else pd.read_csv(f_in_s, encoding='latin-1', dtype=str)
                        df['archivo_origen'] = f_in_s.name
                        da, h1, h2 = motor_sf5(df)
                        if da is not None:
                            st.session_state.da_actual = da
                            st.session_state.h1_actual = h1
                            st.session_state.h2_actual = h2
                            st.session_state.tipo_depuracion_actual = "Individual (Auditoría Única)"
                            st.rerun()
                    except Exception as e:
                        st.error(f"Error al procesar archivo único: {e}")

        with tab_boveda:
            st.subheader("🗄️ Historial de Reportes de Depuración (Nube)")
            if st.session_state.db_depuracion:
                lista_boveda_dep = []
                for k, v in st.session_state.db_depuracion.items():
                    lista_boveda_dep.append({
                        "ID Registro": k,
                        "Nombre": v.get("nombre", "Sin nombre"),
                        "Fecha": v.get("fecha", "N/A"),
                        "Tipo": v.get("tipo", "N/A"),
                        "Total Mapeado": v.get("procesados", 0),
                        "Duplicados": v.get("duplicados", 0),
                        "Únicos": v.get("unicos", 0)
                    })
                df_bov_vista = pd.DataFrame(lista_boveda_dep).sort_values(by="ID Registro", ascending=False)
                
                # --- CONTROL DE SELECCIÓN DE RECOPILACIÓN ---
                id_recuperar = st.selectbox("Selecciona Depuración del Historial Histórico:", list(st.session_state.db_depuracion.keys())[::-1], key="sb_rec_dep")
                
                # --- APLICACIÓN DEL MARCADOR VERDE PARA FILA SELECCIONADA ---
                def resaltar_depuracion_activa(row):
                    color = '#d1e7dd' if row['ID Registro'] == id_recuperar else ''
                    return [f'background-color: {color}'] * len(row)
                
                st.dataframe(df_bov_vista.style.apply(resaltar_depuracion_activa, axis=1), use_container_width=True, hide_index=True)
                
                st.markdown("---")
                col_rec, col_el = st.columns([2.5, 1.5])
                
                with col_rec:
                    if id_recuperar:
                        if st.button("🔄 Recuperar a Mesa de Trabajo", use_container_width=True):
                            data_h = st.session_state.db_depuracion[id_recuperar]
                            st.session_state.da_actual = pd.read_json(io.StringIO(data_h["da_json"]), orient='split')
                            st.session_state.h1_actual = pd.read_json(io.StringIO(data_h["h1_json"]), orient='split')
                            st.session_state.h2_actual = pd.read_json(io.StringIO(data_h["h2_json"]), orient='split')
                            st.session_state.tipo_depuracion_actual = data_h["tipo"]
                            st.toast(f"Reporte {id_recuperar} cargado en pantalla", icon="🔄")
                            time.sleep(0.3)
                            st.rerun()
                
                with col_el:
                    seguro_borrado_f = st.checkbox("🔐 Confirmar eliminación física permanente", key="chk_seg_bov_dep")
                    if st.button("🗑️ BORRAR DE BÓVEDA", use_container_width=True, type="secondary", disabled=not seguro_borrado_f):
                        if id_recuperar:
                            try:
                                cell = ws.find(id_recuperar, in_column=1)
                                if cell: ws.delete_rows(cell.row)
                                del st.session_state.db_depuracion[id_recuperar]
                                st.warning(f"Reporte {id_recuperar} eliminado permanentemente de la nube.")
                                time.sleep(0.5)
                                st.rerun()
                            except Exception as e: st.error(f"Error al borrar de Sheets: {e}")
            else:
                st.info("La bóveda histórica de depuración está vacía.")

        # --- DESPLIEGUE CONTINUO DE LA MESA DE TRABAJO ACTUAL ---
        if st.session_state.da_actual is not None:
            st.write("---")
            st.markdown(f"### 📍 MESA DE TRABAJO EN OPERACIÓN: **{st.session_state.tipo_depuracion_actual}**")
            renderizar_interfaz(st.session_state.da_actual, st.session_state.h1_actual, st.session_state.h2_actual, "mesa")
            
            st.write("---")
            st.write("⚠️ **Zona de Peligro Interna**")
            seguro_limpieza_mesa = st.checkbox("🔐 Confirmar vaciado completo de la mesa de trabajo actual", key="chk_seg_mesa")
            if st.button("🗑️ LIMPIAR MESA DE TRABAJO (Reiniciar)", use_container_width=True, type="secondary", disabled=not seguro_limpieza_mesa):
                st.session_state.da_actual = None
                st.session_state.h1_actual = None
                st.session_state.h2_actual = None
                st.session_state.tipo_depuracion_actual = ""
                st.toast("Mesa de trabajo vaciada", icon="🗑️")
                time.sleep(0.5)
                st.rerun()
    elif st.session_state.menu == "SF6":
        # ==========================================
        # --- FILTRO 1: INICIO MAESTRO DEL SISTEMA (PIN 1827) ---
        # ==========================================
        if "maestro_auth" not in st.session_state:
            st.session_state.maestro_auth = False

        if not st.session_state.maestro_auth:
            st.title("🔒 SF6 - Suite de Gestión Municipal (DAP)")
            st.markdown("### **Control de Acceso Institucional**")
            st.caption("Por seguridad de la Dirección de Alumbrado Público, ingrese el PIN Maestro para inicializar el módulo.")
            
            pass_maestro = st.text_input("🔑 Ingrese PIN Maestro:", type="password", key="pass_maestro_root")
            if st.button("🔓 Inicializar Sistema", use_container_width=True, type="primary"):
                if pass_maestro == config.PIN_MAESTRO:
                    st.session_state.maestro_auth = True
                    st.success("🎉 Acceso Maestro concedido. Inicializando entorno...")
                    time.sleep(0.5)
                    st.rerun()
                else:
                    st.error("❌ PIN Maestro incorrecto. Acceso denegado al módulo SF6.")
            st.stop()

        # --- El sistema continúa con todo su poder si el PIN Maestro es correcto ---
        st.title("📦 SF6 - Sistema de Gestión de Almacén (DAP)")
        
        LEYENDA_OFICIAL = "Este material es propiedad del Ayuntamiento de Toluca y se genera en la Dirección de Alumbrado Público"

        # --- CONEXIÓN DE SEGURIDAD CON GOOGLE SHEETS ---
        scope = ["https://www.googleapis.com/auth/spreadsheets"]
        creds_dict = {
            "type": st.secrets["connections"]["gsheets"]["type"],
            "project_id": st.secrets["connections"]["gsheets"]["project_id"],
            "private_key_id": st.secrets["connections"]["gsheets"]["private_key_id"],
            "private_key": st.secrets["connections"]["gsheets"]["private_key"].replace('\\n', '\n'),
            "client_email": st.secrets["connections"]["gsheets"]["client_email"],
            "client_id": st.secrets["connections"]["gsheets"]["client_id"],
            "auth_uri": st.secrets["connections"]["gsheets"]["auth_uri"],
            "token_uri": st.secrets["connections"]["gsheets"]["token_uri"],
            "auth_provider_x509_cert_url": st.secrets["connections"]["gsheets"]["auth_provider_x509_cert_url"],
            "client_x509_cert_url": st.secrets["connections"]["gsheets"]["client_x509_cert_url"]
        }
        creds = Credentials.from_service_account_info(creds_dict, scopes=scope)
        client = gspread.authorize(creds)
        SHEET_ID = "14_fewol5DiFXoiO102wviiWR08Lw3PKHzEjSbMwxUm8"
        
        try:
            sh = client.open_by_key(SHEET_ID)
            ws = sh.worksheet("Boveda_Bajas")
        except:
            ws = sh.get_worksheet(0)
            # --- CONEXIÓN CON PESTAÑA DE INVENTARIO EN LA NUBE ---
        try:
            ws_inv = sh.worksheet("Inventario_Catalogo")
        except:
            try:
                # Si la pestaña no existe en tu Google Sheets, el sistema la crea automáticamente
                ws_inv = sh.add_worksheet(title="Inventario_Catalogo", rows="100", cols="5")
                ws_inv.append_row(["ID", "Material", "Stock", "Min", "Unidad"])
            except:
                ws_inv = sh.get_worksheet(0)

        # Redefinición dinámica de data_manager para centralizar el catálogo en la nube
        def cloud_cargar_inventario():
            try:
                rows = ws_inv.get_all_values()
                if len(rows) > 1:
                    headers = rows[0]
                    df = pd.DataFrame(rows[1:], columns=headers)
                    # Forzamos conversión a números para evitar errores en las alertas críticas de Streamlit
                    if 'Stock' in df.columns:
                        df['Stock'] = pd.to_numeric(df['Stock'], errors='coerce').fillna(0).astype(int)
                    if 'Min' in df.columns:
                        df['Min'] = pd.to_numeric(df['Min'], errors='coerce').fillna(0).astype(int)
                    return df
                return None
            except:
                return None

        def cloud_guardar_inventario(df):
            try:
                ws_inv.clear()
                ws_inv.append_row(df.columns.tolist())
                ws_inv.append_rows(df.values.tolist())
            except Exception as e:
                st.error(f"❌ Error de sincronización de inventario en la nube: {e}")

        # Inyectamos las nuevas funciones en el motor local
        data_manager.cargar_inventario = cloud_cargar_inventario
        data_manager.guardar_inventario = cloud_guardar_inventario

        # Lector ultra-robusto territorial
        try:
            try:
                df_territorial = pd.read_csv('DELEUTB2.csv', encoding='utf-8')
            except Exception:
                try:
                    df_territorial = pd.read_csv('DELEUTB2.csv', encoding='latin-1')
                except Exception:
                    df_territorial = pd.read_csv('DELEUTB2.csv', encoding='utf-8-sig')
            
            df_territorial.columns = [str(c).strip().upper() for c in df_territorial.columns]
            df_territorial = df_territorial.dropna(subset=['DELEGACION', 'UTB'])
            df_territorial['DELEGACION'] = df_territorial['DELEGACION'].astype(str).str.strip().str.upper()
            df_territorial['UTB'] = df_territorial['UTB'].astype(str).str.strip().str.upper()
            
            DELEGACIONES_TOLUCA = sorted(df_territorial['DELEGACION'].unique().tolist())
        except Exception as e:
            st.sidebar.info("⚙️ Sincronizando base territorial del repositorio...")
            DELEGACIONES_TOLUCA = [
                "CENTRO HISTORICO", "BARRIOS TRADICIONALES", "ARBOL DE LAS MANITAS", "LA MAQUINITA", 
                "INDEPENDENCIA", "SAN SEBASTIAN", "UNIVERSIDAD", "SANTA MARIA DE LAS ROSAS", 
                "DEL PARQUE", "DELEGACION DEL PARQUE 18 DE MARZO", "ADOLFO LOPEZ MATEOS", 
                "CIUDAD UNIVERSITARIA", "NUEVA OXTOTITLAN", "SAN BUENAVENTURA", "CAPULTITLAN", 
                "TLACOTEPEC", "SAN JUAN TILAPA", "SAN FELIPE TLALMIMILOLPAN", "PINO SUAREZ", 
                "SAN MATEO OXTOTITLAN", "SANTA CRUZ ATZCAPOTZALTONGO", "SAN PEDRO TOTOLTEPEC", 
                "SAN LORENZO TEPALTITLAN", "SAN ANDRES CUEXCONTITLAN", "SAN CRISTOBAL HUICHOCHITLAN", 
                "SAN PABLO AUTOPAN", "SAN JUAN AUTOPAN", "SAN MARTIN TOLTEPEC", "SAUCES", 
                "CALIXTLAHUACA", "TECAXIC", "SAN ANTONIO BUENAVISTA"
            ]
            respaldo_datos = []
            for d in DELEGACIONES_TOLUCA:
                for i in range(1, 6):
                    respaldo_datos.append({"DELEGACION": d, "UTB": f"UTB {i}"})
            df_territorial = pd.DataFrame(respaldo_datos)

        # ==========================================
        # --- MOTOR DE PERSISTENCIA INMUTABLE NUBE ---
        # ==========================================
        if "db_inventario" not in st.session_state:
            inv_datos = data_manager.cargar_inventario()
            if inv_datos is not None:
                st.session_state.db_inventario = inv_datos
            else:
                df_base = pd.DataFrame(STOCK_INICIAL)
                df_base['Stock'] = 100 
                df_base['Min'] = 10    
                data_manager.guardar_inventario(df_base)
                st.session_state.db_inventario = df_base

        # --- SINCRO-BÓVEDA RESILIENTE (ANTI-DESCONEXIONES) ---
        bovedas_completas = ["db_oficios", "boveda_mmd", "db_depuracion", "vales_historial"]
        if any(b not in st.session_state for b in bovedas_completas):
            for _ in range(3):
                try:
                    if "db_oficios" not in st.session_state: st.session_state.db_oficios = {}
                    if "boveda_mmd" not in st.session_state: st.session_state.boveda_mmd = {}
                    if "db_depuracion" not in st.session_state: st.session_state.db_depuracion = {}
                    if "vales_historial" not in st.session_state: st.session_state.vales_historial = []
                    
                    # Motor de lectura por coordenadas físicas puras para el almacén (SF6)
                    filas_raw = ws.get_all_values()
                    if len(filas_raw) > 1:
                        for row in filas_raw[1:]: # Saltamos los encabezados de la fila 1
                            if len(row) >= 5:
                                reg_id = str(row[0]).strip()
                                origen = str(row[2]).strip() if str(row[2]).strip() else "Sin Nombre"
                                datos_raw = str(row[4]).strip() if row[4] else "{}"
                                
                                if reg_id.startswith("SF4-PRY-"):
                                    try: st.session_state.boveda_mmd[origen] = json.loads(datos_raw) if isinstance(datos_raw, str) else datos_raw
                                    except: pass
                                elif reg_id.startswith("SF4-OFC-"):
                                    try: st.session_state.db_oficios[origen] = json.loads(datos_raw) if isinstance(datos_raw, str) else datos_raw
                                    except: pass
                                elif reg_id.startswith("SF5-DEP-"):
                                    try: st.session_state.db_depuracion[reg_id] = json.loads(datos_raw) if isinstance(datos_raw, str) else datos_raw
                                    except: pass
                                elif reg_id.startswith("SF6-VAL-"):
                                    try:
                                        vale_parsed = json.loads(datos_raw) if isinstance(datos_raw, str) else datos_raw
                                        if vale_parsed not in st.session_state.vales_historial:
                                            st.session_state.vales_historial.append(vale_parsed)
                                    except: pass
                    break
                except Exception:
                    time.sleep(1)

        if "carrito_vale" not in st.session_state:
            st.session_state.carrito_vale = []
        if "admin_auth" not in st.session_state:
            st.session_state.admin_auth = False

        tab_inv, tab_vales, tab_seguimiento, tab_admin = st.tabs([
            "📊 Existencias y Resumen", 
            "🚚 Salida (Vale Oficial)", 
            "🎯 Seguimiento y Consumos",
            "⚙️ Gestión Almacén"
        ])

        # ==========================================
        # --- PESTAÑA 1: EXISTENCIAS Y RESUMEN ---
        # ==========================================
        with tab_inv:
            df_inv = st.session_state.db_inventario
            st.subheader("🚨 Dashboard de Inventario")
            
            if "filtrar_critico" not in st.session_state:
                st.session_state.filtrar_critico = False

            criticos = len(df_inv[df_inv['Stock'] <= df_inv['Min']])
            
            c_met1, c_met2 = st.columns(2)
            c_met1.metric("📦 Materiales en Catálogo", len(df_inv))
            c_met2.metric("⚠️ Alertas de Stock Crítico", criticos, delta=-criticos, delta_color="inverse")

            text_filtro = "👁️ Mostrar SOLO materiales con Stock Crítico" if not st.session_state.filtrar_critico else "🔄 Mostrar TODO el Catálogo"
            type_filtro = "secondary" if not st.session_state.filtrar_critico else "primary"
            
            if st.button(text_filtro, type=type_filtro, use_container_width=True):
                st.session_state.filtrar_critico = not st.session_state.filtrar_critico
                st.rerun()

            if st.session_state.filtrar_critico:
                st.markdown("### ⚠️ Mostrando Materiales en Stock Crítico o Alerta")
                df_mostrar = df_inv[df_inv['Stock'] <= df_inv['Min']].reset_index(drop=True)
                if df_mostrar.empty:
                    st.success("🎉 ¡Excelente! No hay ningún material con stock crítico en este momento.")
            else:
                st.write("### Inventario Actual de Materiales e Insumos")
                df_mostrar = df_inv

            st.dataframe(df_mostrar, use_container_width=True, hide_index=True)
            
            if st.button("📄 GENERAR RESUMEN EJECUTIVO (EXCEL)", use_container_width=True):
                output_res = io.BytesIO()
                with pd.ExcelWriter(output_res, engine='openpyxl') as writer:
                    df_mostrar.to_excel(writer, index=False, sheet_name='Estado_Almacen')
                st.download_button(
                    label="📥 Descargar Reporte de Stock", 
                    data=output_res.getvalue(), 
                    file_name=f"Resumen_Almacen_{pd.Timestamp.now().strftime('%d-%m-%Y')}.xlsx", 
                    use_container_width=True
                )

        # ==========================================
        # --- PESTAÑA 2: SALIDA (VALE OFICIAL) ---
        # ==========================================
        with tab_vales:
            df_inv = st.session_state.db_inventario
            st.subheader("🧾 Generador de Vale de Salida Oficial")
            
            prox_folio_num = len(st.session_state.vales_historial) + 1
            folio_actual = f"DAP-{prox_folio_num}"
            
            st.markdown(
                """
                <style>
                @media print {
                    .no-imprimir {
                        display: none !important;
                    }
                }
                </style>
                <div class='no-imprimir' style='background-color: #eef4f8; padding: 15px; border-left: 6px solid #1f4e78; border-radius: 6px; margin-bottom: 15px;'>
                    <h4 style='margin:0; color: #1f4e78;'>🚚 ASIGNACIÓN DE UNIDAD Y UBICACIÓN DE DESTINO</h4>
                    <p style='margin:0; font-size:13px; color:#555;'>Asigne la brigada y la delimitación territorial para el seguimiento posterior del material.</p>
                </div>
                """, 
                unsafe_allow_html=True
            )
            
            if "sb_vale_delegacion" not in st.session_state:
                st.session_state.sb_vale_delegacion = "TODAS"
            if "sb_vale_utb" not in st.session_state:
                st.session_state.sb_vale_utb = "TODAS"
            if "prev_vale_del" not in st.session_state:
                st.session_state.prev_vale_del = "TODAS"
            if "prev_vale_utb" not in st.session_state:
                st.session_state.prev_vale_utb = "TODAS"

            val_del_current = st.session_state.sb_vale_delegacion
            val_utb_current = st.session_state.sb_vale_utb

            if val_del_current != st.session_state.prev_vale_del:
                if val_del_current == "TODAS":
                    st.session_state.sb_vale_utb = "TODAS"
                    val_utb_current = "TODAS"
                else:
                    if not df_territorial.empty:
                        utbs_validas = sorted(df_territorial[df_territorial['DELEGACION'] == val_del_current]['UTB'].unique().tolist())
                        if val_utb_current != "TODAS" and val_utb_current not in utbs_validas:
                            st.session_state.sb_vale_utb = "TODAS"
                            val_utb_current = "TODAS"
                st.session_state.prev_vale_del = val_del_current
                st.session_state.prev_vale_utb = val_utb_current

            elif val_utb_current != st.session_state.prev_vale_utb:
                if val_utb_current != "TODAS" and not df_territorial.empty:
                    deles_de_utb = df_territorial[df_territorial['UTB'] == val_utb_current]['DELEGACION'].unique().tolist()
                    if val_del_current == "TODAS" or val_del_current not in deles_de_utb:
                        if deles_de_utb:
                            st.session_state.sb_vale_delegacion = sorted(deles_de_utb)[0]
                            val_del_current = sorted(deles_de_utb)[0]
                            st.session_state.prev_vale_del = val_del_current
                st.session_state.prev_vale_utb = val_utb_current

            if val_utb_current != "TODAS" and not df_territorial.empty:
                delegaciones_dispo_vale = ["TODAS"] + sorted(df_territorial[df_territorial['UTB'] == val_utb_current]['DELEGACION'].unique().tolist())
            else:
                delegaciones_dispo_vale = ["TODAS"] + DELEGACIONES_TOLUCA

            if val_del_current != "TODAS" and not df_territorial.empty:
                utbs_dispo_vale = ["TODAS"] + sorted(df_territorial[df_territorial['DELEGACION'] == val_del_current]['UTB'].unique().tolist())
            else:
                utbs_dispo_vale = ["TODAS"] + (sorted(df_territorial['UTB'].unique().tolist()) if not df_territorial.empty else [])

            if val_del_current not in delegaciones_dispo_vale: val_del_current = "TODAS"
            if val_utb_current not in utbs_dispo_vale: val_utb_current = "TODAS"

            idx_del_vale = delegaciones_dispo_vale.index(val_del_current)
            idx_utb_vale = utbs_dispo_vale.index(val_utb_current)

            c_bri1, c_bri2, c_bri3 = st.columns(3)
            with c_bri1:
                bri_sel = st.selectbox(
                    "Brigada Asignada:", 
                    [f"Brigada {i}" for i in range(1, 18)] + ["Personal de Mantenimiento Interno", "Cuadrilla de Alumbrado Especial"],
                    key="brigada_salida_vales"
                )
            with c_bri2:
                delegacion_sel = st.selectbox(
                    "Delegación Destino:", 
                    delegaciones_dispo_vale, 
                    index=idx_del_vale, 
                    key="sb_vale_delegacion"
                )
            with c_bri3:
                utb_sel = st.selectbox(
                    "UTB Destino:", 
                    utbs_dispo_vale, 
                    index=idx_utb_vale, 
                    key="sb_vale_utb"
                )

            st.info(f"Folio del Vale en Proceso: **{folio_actual}**")

            with st.container(border=True):
                st.markdown("**🛒 Adición de Materiales al Lote**")
                c2, c3 = st.columns([2, 1])
                mat_sel = c2.selectbox("Seleccione Insumo:", df_inv['Material'].tolist())
                can_sel = c3.number_input("Cantidad a Entregar:", min_value=1, step=1)
                
                if st.button("➕ Agregar Insumo al Carrito", use_container_width=True):
                    stock_real = df_inv.loc[df_inv['Material'] == mat_sel, 'Stock'].values[0]
                    if stock_real >= can_sel:
                        existe = False
                        for item in st.session_state.carrito_vale:
                            if item['Material'] == mat_sel:
                                if (item['Cantidad'] + can_sel) <= stock_real:
                                    item['Cantidad'] += can_sel
                                    existe = True
                                else:
                                    st.error("⚠️ La cantidad agregada supera las existencias físicas en el almacén.")
                                    existe = True
                        if not existe:
                            st.session_state.carrito_vale.append({
                                "Material": mat_sel, 
                                "Cantidad": can_sel, 
                                "Unidad": df_inv.loc[df_inv['Material'] == mat_sel, 'Unidad'].values[0]
                            })
                        
                        st.session_state.vale_listo_descarga = None
                        st.toast(f"✅ {mat_sel} sumado al vale actual.")
                        time.sleep(0.3)
                        st.rerun()
                    else:
                        st.error(f"⚠️ No hay suficiente material disponible. Existencia actual: {stock_real}")

            st.markdown("### 📝 Control y Notas de Entrega")
            obs_digital = st.text_area(
                "Observaciones del Responsable de Almacén (Se captura en sistema):", 
                placeholder="Ej: Material destinado a la rehabilitación de luminarias en San Martín Toltepec. Se entrega cable con empalmes de fábrica.",
                key="obs_responsable_entrega"
            )

            # --- VISTA DEL CARRITO ACTIVO MIENTRAS CAPTURAS ---
            if st.session_state.carrito_vale:
                st.write("---")
                st.markdown("### **🛒 Resumen del Lote a Entregar:**")
                
                df_carrito = pd.DataFrame(st.session_state.carrito_vale)
                st.dataframe(df_carrito, use_container_width=True, hide_index=True)
                
                col_v1, col_v2 = st.columns(2)
                
                with col_v1:
                    seguro_cancelar_vale = st.checkbox("🔐 Confirmar vaciado del carrito actual", key="chk_seg_cancel_vale")
                    if st.button("❌ Cancelar Vale Completo", use_container_width=True, disabled=not seguro_cancelar_vale):
                        st.session_state.carrito_vale = []
                        st.rerun()
                
                with col_v2:
                    if st.button(f"🚀 PROCESAR Y EMITIR VALE ({folio_actual})", type="primary", use_container_width=True):
                        try:
                            # --- CÁLCULO DINÁMICO DE FOLIO CORREGIDO (APUNTANDO A SHEET1) ---
                            conn = st.connection("gsheets", type=GSheetsConnection)
                            df_vales_db = conn.read(spreadsheet=URL_DB, worksheet="Sheet1", ttl=0).dropna(how='all')
                            
                            nums_existentes = []
                            if not df_vales_db.empty and df_vales_db.shape[1] > 3:
                                col_id_vals = df_vales_db.iloc[:, 0].astype(str).tolist()
                                for c_val in col_id_vals:
                                    if "SF6-VAL-DAP-" in c_val:
                                        try:
                                            nums_existentes.append(int(c_val.split("SF6-VAL-DAP-")[-1]))
                                        except:
                                            pass
                            
                            # Ajuste dinámico del folio anti-duplicados
                            true_num = max(nums_existentes) + 1 if nums_existentes else (len(st.session_state.vales_historial) + 1)
                            folio_actual = f"DAP-{true_num}"
                            id_reg_vale = f"SF6-VAL-{folio_actual}"
                            
                            if any(v["Folio"] == folio_actual for v in st.session_state.vales_historial):
                                local_max = max([int(v["Folio"].replace("DAP-", "")) for v in st.session_state.vales_historial if "DAP-" in v["Folio"]], default=0)
                                true_num = max(true_num, local_max + 1)
                                folio_actual = f"DAP-{true_num}"
                                id_reg_vale = f"SF6-VAL-{folio_actual}"

                            # --- 1. CONSTRUCCIÓN DEL PDF ---
                            from fpdf import FPDF
                            pdf = FPDF()
                            pdf.add_page()
                            
                            pdf.set_font("Arial", 'B', 14)
                            pdf.cell(0, 10, "AYUNTAMIENTO DE TOLUCA", ln=True, align='C')
                            pdf.set_font("Arial", 'B', 12)
                            pdf.cell(0, 8, "DIRECCION DE ALUMBRADO PUBLICO", ln=True, align='C')
                            pdf.cell(0, 8, f"VALE OFICIAL DE SALIDA: {folio_actual}", ln=True, align='C')
                            pdf.ln(8)
                            
                            pdf.set_font("Arial", '', 10)
                            pdf.cell(0, 6, f"Fecha y Hora de Emision: {pd.Timestamp.now().strftime('%d/%m/%Y %H:%M')}", ln=True)
                            pdf.cell(0, 6, f"Ubicacion Geografica: {delegacion_sel} - {utb_sel}", ln=True)
                            pdf.set_font("Arial", 'B', 10)
                            pdf.cell(0, 6, f"UNIDAD / BRIGADA DESTINO: {bri_sel.upper()}", ln=True)
                            pdf.ln(4)
                            
                            pdf.set_fill_color(230, 235, 240)
                            pdf.set_font("Arial", 'B', 10)
                            pdf.cell(110, 8, " Descripcion del Material / Insumo", 1, 0, 'L', True)
                            pdf.cell(36, 8, "Entregado", 1, 0, 'C', True)
                            pdf.cell(22, 8, "Utilizado", 1, 0, 'C', True)
                            pdf.cell(22, 8, "Devuelto", 1, 1, 'C', True)
                            
                            pdf.set_font("Arial", '', 10)
                            for it in st.session_state.carrito_vale:
                                pdf.cell(110, 8, f" {str(it['Material'])}", 1, 0, 'L')
                                pdf.cell(36, 8, f"{it['Cantidad']} {it['Unidad']}", 1, 0, 'C')
                                pdf.cell(22, 8, "", 1, 0, 'C')
                                pdf.cell(22, 8, "", 1, 1, 'C')
                            pdf.ln(6)
                            
                            pdf.set_font("Arial", 'B', 10)
                            pdf.cell(0, 6, "Observaciones del Responsable de Almacen (Sistema):", ln=True)
                            pdf.set_font("Arial", 'I', 9)
                            msg_obs = obs_digital if obs_digital.strip() else "Ninguna anotada en sistema al momento de la salida."
                            pdf.multi_cell(0, 5, msg_obs.encode('latin-1', 'replace').decode('latin-1'), 1)
                            pdf.ln(4)
                            
                            pdf.set_font("Arial", 'B', 10)
                            pdf.cell(0, 6, "Observaciones de la Brigada al Recibir (Llenar en Fisico a Mano):", ln=True)
                            pdf.set_fill_color(255, 255, 255)
                            pdf.cell(0, 15, "", 1, ln=True, fill=True)
                            pdf.ln(10)
                            
                            pdf.set_font("Arial", 'I', 9)
                            pdf.multi_cell(0, 5, LEYENDA_OFICIAL.encode('latin-1', 'replace').decode('latin-1'), align='C')
                            pdf.ln(12)
                            
                            y_pos_firmas = pdf.get_y()
                            pdf.set_font("Arial", 'B', 8)
                            pdf.set_xy(15, y_pos_firmas)
                            pdf.cell(75, 4, "_____________________________________", ln=False, align='C')
                            pdf.set_xy(15, y_pos_firmas + 4)
                            pdf.cell(75, 4, "RESPONSABLE DE ENTREGA DE MATERIAL", ln=False, align='C')
                            pdf.set_xy(15, y_pos_firmas + 8)
                            pdf.set_font("Arial", '', 7)
                            pdf.cell(75, 4, "(Firma y Sello de Almacen DAP)", ln=False, align='C')
                            pdf.set_font("Arial", 'B', 8)
                            pdf.set_xy(115, y_pos_firmas)
                            pdf.cell(75, 4, "_____________________________________", ln=False, align='C')
                            pdf.set_xy(115, y_pos_firmas + 4)
                            pdf.cell(75, 4, "RESPONSABLE QUE RECIBE MATERIAL", ln=False, align='C')
                            pdf.set_xy(115, y_pos_firmas + 8)
                            pdf.set_font("Arial", '', 7)
                            pdf.cell(75, 4, f"({bri_sel.upper()})", ln=False, align='C')
                            
                            pdf_output = pdf.output(dest='S')
                            pdf_bytes = pdf_output.encode('latin-1', 'replace') if isinstance(pdf_output, str) else pdf_output

                            # --- 2. DEDUCCIÓN LOCAL DEL STOCK ---
                            for item in st.session_state.carrito_vale:
                                idx = df_inv[df_inv['Material'] == item['Material']].index[0]
                                st.session_state.db_inventario.at[idx, 'Stock'] -= item['Cantidad']
                            
                            data_manager.guardar_inventario(st.session_state.db_inventario)
                            
                            materiales_con_cierres = []
                            for m_car in st.session_state.carrito_vale:
                                m_car["Utilizado"] = 0
                                m_car["Devuelto"] = 0
                                materiales_con_cierres.append(m_car)
                            
                            payload_vale = {
                                "Folio": folio_actual, 
                                "Fecha": pd.Timestamp.now().strftime('%Y-%m-%d'),
                                "FechaHora": pd.Timestamp.now().strftime('%d/%m/%Y %H:%M'),
                                "Brigada": bri_sel,
                                "Delegacion": delegacion_sel,      
                                "UTB": utb_sel,               
                                "Delegacion_Real": delegacion_sel, 
                                "UTB_Real": utb_sel,             
                                "Estado": "Pendiente",
                                "Materiales": materiales_con_cierres,
                                "Observaciones": obs_digital if obs_digital.strip() else "Sin observaciones"
                            }
                            
                            st.session_state.vales_historial.append(payload_vale)
                            
                            # --- 3. INSERCIÓN EN SHEET1 ---
                            fecha_actual_mx = pd.Timestamp.now().strftime("%d/%m/%Y %H:%M:%S")
                            nuevo_vale_fila = pd.DataFrame([{"Fecha": fecha_actual_mx, "Nombre_Ruta": f"VALE_{folio_actual}", "Usuario_Generador": st.session_state.usuario_nombre, "Datos_JSON": json.dumps(payload_vale)}])
                            conn.update(spreadsheet=URL_DB, worksheet="Sheet1", data=pd.concat([df_vales_db, nuevo_vale_fila], ignore_index=True))
                            
                            st.session_state.vale_listo_descarga = {"folio": folio_actual, "bytes": pdf_bytes}
                            st.session_state.carrito_vale = []
                            st.success("✅ Vale Oficial sincronizado con éxito.")
                        except Exception as e:
                            st.error(f"❌ Error al procesar datos: {e}")
                        st.rerun()

            if "vale_listo_descarga" in st.session_state and st.session_state.vale_listo_descarga:
                st.write("---")
                vale_data = st.session_state.vale_listo_descarga
                st.success(f"🎉 ¡Vale Oficial **{vale_data['folio']}** registrado exitosamente!")
                
                c_dl1, c_dl2 = st.columns([3, 1])
                c_dl1.download_button(
                    label=f"📥 DESCARGAR COMPROBANTE PDF ({vale_data['folio']})",
                    data=vale_data["bytes"],
                    file_name=f"Vale_Oficial_Salida_{vale_data['folio']}.pdf",
                    mime="application/pdf",
                    use_container_width=True,
                    type="primary"
                )
                if c_dl2.button("🔄 Ocultar Notificación", use_container_width=True):
                    st.session_state.vale_listo_descarga = None
                    st.rerun()

        # ==========================================
        # --- PESTAÑA 3: SEGUIMIENTO Y CONSUMOS COMPLETO ---
        # ==========================================
        with tab_seguimiento:
            st.subheader("🎯 Control de Material Aplicado en Campo y Devoluciones")
            
            # --- MÉTRICAS ANALÍTICAS ORIGINALES (INTACTAS) ---
            total_vales = len(st.session_state.vales_historial)
            pendientes = sum(1 for v in st.session_state.vales_historial if v.get("Estado", "Pendiente") == "Pendiente")
            redireccionados = sum(1 for v in st.session_state.vales_historial if v.get("Delegacion") != v.get("Delegacion_Real"))
            
            c_seg1, c_seg2, c_seg3 = st.columns(3)
            c_seg1.metric("📋 Total Vales Emitidos", total_vales)
            c_seg2.metric("⏳ Pendientes de Conciliar", pendientes, delta=f"-{total_vales - pendientes} cerrados", delta_color="normal" if pendientes > 0 else "off")
            c_seg3.metric("⚠️ Cambios de Ruta (Alertas)", redireccionados, delta="Incidencias en campo" if redireccionados > 0 else "Operación Nominal", delta_color="inverse" if redireccionados > 0 else "normal")
            
            if total_vales == 0:
                st.info("📂 No hay vales registrados en la Bóveda para darles seguimiento.")
            else:
                st.write("---")
                
                # --- ARQUITECTURA VISUAL PARALELA (SIEMPRE VISIBLE) ---
                col_tabla_fija, col_operacion_activa = st.columns([1.1, 1.2])
                
                with col_tabla_fija:
                    st.markdown("### 📊 Monitoreo Técnico de Despliegue")
                    st.caption("Selecciona un folio de la lista para auditarlo o capturar cierres:")
                    
                    # Generación de la tabla fija de control
                    tabla_resumen_seguimiento = []
                    for v in st.session_state.vales_historial:
                        loc_original = f"{v['Delegacion']} ({v['UTB']})"
                        loc_real = f"{v.get('Delegacion_Real', v['Delegacion'])} ({v.get('UTB_Real', v['UTB'])})"
                        status_ruta = "📢 NOMINAL" if v['Delegacion'] == v.get('Delegacion_Real') else "🔄 REDIRECCIONADO"
                        status_cierre = "🔴 PENDIENTE" if v.get("Estado", "Pendiente") == "Pendiente" else "🟢 CONCILIADO"
                        
                        tabla_resumen_seguimiento.append({
                            "Folio": v["Folio"],
                            "Brigada": v["Brigada"],
                            "Destino Almacén": loc_original,
                            "Destino Real Aplicado": loc_real,
                            "Ruta": status_ruta,
                            "Estatus": status_cierre
                        })
                    df_seg_vista = pd.DataFrame(tabla_resumen_seguimiento)
                    
                    # Selector integrado como disparador de fila
                    folios_dispo = [v["Folio"] for v in st.session_state.vales_historial]
                    v_select = st.selectbox("👉 Folio Activo en Mesa:", folios_dispo, key="sb_seguimiento_folios")
                    
                    # Lógica matemática de marcado en color verde claro (#d1e7dd)
                    def resaltar_vale_seguimiento(row):
                        color = '#d1e7dd' if row['Folio'] == v_select else ''
                        return [f'background-color: {color}'] * len(row)
                        
                    st.dataframe(df_seg_vista.style.apply(resaltar_vale_seguimiento, axis=1), use_container_width=True, hide_index=True)
                
                with col_operacion_activa:
                    st.markdown("### 🛠️ Mesa de Conciliación en Tiempo Real")
                    
                    idx_vale = next(i for i, item in enumerate(st.session_state.vales_historial) if item["Folio"] == v_select)
                    vale_obj = st.session_state.vales_historial[idx_vale]
                    
                    if vale_obj.get("Estado", "Pendiente") == "Conciliado":
                        st.success(f"✅ Folio **{v_select}** consolidado. Al guardar actualizarás los datos en la nube.")
                    
                    # --- FILTRADO CRUZADO TERRITORIAL EN CONCILIACIÓN ---
                    if "sb_rec_delegacion" not in st.session_state or st.session_state.get("current_v_select") != v_select:
                        st.session_state.current_v_select = v_select
                        st.session_state.sb_rec_delegacion = vale_obj.get("Delegacion_Real", vale_obj["Delegacion"])
                        st.session_state.sb_rec_utb = vale_obj.get("UTB_Real", vale_obj["UTB"])
                        st.session_state.prev_rec_del = vale_obj.get("Delegacion_Real", vale_obj["Delegacion"])

                    rec_del_current = st.session_state.sb_rec_delegacion
                    rec_utb_current = st.session_state.sb_rec_utb

                    if rec_del_current != st.session_state.prev_rec_del:
                        if not df_territorial.empty:
                            utbs_validas_rec = sorted(df_territorial[df_territorial['DELEGACION'] == rec_del_current]['UTB'].unique().tolist())
                            if utbs_validas_rec:
                                st.session_state.sb_rec_utb = utbs_validas_rec[0]
                                rec_utb_current = utbs_validas_rec[0]
                        st.session_state.prev_rec_del = rec_del_current

                    delegaciones_dispo_rec = DELEGACIONES_TOLUCA
                    utbs_dispo_rec = sorted(df_territorial[df_territorial['DELEGACION'] == rec_del_current]['UTB'].unique().tolist()) if not df_territorial.empty else ["UTB 1"]

                    idx_del_rec = delegaciones_dispo_rec.index(rec_del_current) if rec_del_current in delegaciones_dispo_rec else 0
                    idx_utb_rec = utbs_dispo_rec.index(rec_utb_current) if rec_utb_current in utbs_dispo_rec else 0

                    with st.container(border=True):
                        st.markdown(f"📋 **Expediente de Salida:** `{vale_obj['Folio']}` | **Brigada:** `{vale_obj['Brigada']}`")
                        st.markdown("⚠️ **¿Fue redireccionada la ruta en campo? Corrige aquí:**")
                        col_re1, col_re2 = st.columns(2)
                        real_del_sel = col_re1.selectbox("Delegación REAL aplicada:", delegaciones_dispo_rec, index=idx_del_rec, key="sb_rec_delegacion")
                        real_utb_sel = col_re2.selectbox("UTB REAL aplicada:", utbs_dispo_rec, index=idx_utb_rec, key="sb_rec_utb")
                    
                    st.write("")
                    df_materiales_vale = pd.DataFrame(vale_obj["Materiales"])
                    
                    # LLAVE DE SOLUCIÓN AL DOBLE CLIC: Callback on_change para actualizar la memoria al instante
                    key_editor_live = f"editor_live_{v_select}"
                    
                    def callback_actualizar_edicion():
                        if key_editor_live in st.session_state:
                            cambios = st.session_state[key_editor_live]
                            if "edited_rows" in cambios:
                                for r_idx, r_changes in cambios["edited_rows"].items():
                                    for col_key, col_val in r_changes.items():
                                        df_materiales_vale.at[int(r_idx), col_key] = col_val

                    df_editado = st.data_editor(
                        df_materiales_vale,
                        column_config={
                            "Material": st.column_config.TextColumn("Insumo", disabled=True),
                            "Cantidad": st.column_config.NumberColumn("Entregado", disabled=True),
                            "Unidad": st.column_config.TextColumn("Unidad", disabled=True),
                            "Utilizado": st.column_config.NumberColumn("Instalado (Calle)", min_value=0, required=True),
                            "Devuelto": st.column_config.NumberColumn("Devuelto (Almacén)", min_value=0, required=True),
                        },
                        hide_index=True, use_container_width=True, key=key_editor_live, on_change=callback_actualizar_edicion
                    )
                    
                    if st.button("💾 Guardar Conciliación de Materiales", use_container_width=True, type="primary"):
                        error_cantidades = False
                        for index, row in df_editado.iterrows():
                            if (row["Utilizado"] + row["Devuelto"]) > row["Cantidad"]:
                                st.error(f"⚠️ Error en {row['Material']}: La suma supera las piezas entregadas originalmente.")
                                error_cantidades = True
                        
                        if not error_cantidades:
                            try:
                                # 1. Ajuste matemático de inventarios locales
                                for index, row in df_editado.iterrows():
                                    dif_devolucion = row["Devuelto"] - df_materiales_vale.loc[index, "Devuelto"]
                                    if dif_devolucion > 0:
                                        idx_inv = st.session_state.db_inventario[st.session_state.db_inventario['Material'] == row['Material']].index[0]
                                        st.session_state.db_inventario.at[idx_inv, 'Stock'] += dif_devolucion
                                
                                data_manager.guardar_inventario(st.session_state.db_inventario)
                                
                                # 2. Sincronización estructural en memoria
                                st.session_state.vales_historial[idx_vale]["Materiales"] = df_editado.to_dict(orient='records')
                                st.session_state.vales_historial[idx_vale]["Delegacion_Real"] = real_del_sel
                                st.session_state.vales_historial[idx_vale]["UTB_Real"] = real_utb_sel
                                st.session_state.vales_historial[idx_vale]["Estado"] = "Conciliado"
                                
                                # 3. Sincronización en la Bóveda de Google Sheets
                                id_reg_buscar = f"SF6-VAL-{v_select}"
                                cell = ws.find(id_reg_buscar, in_column=1)
                                if cell:
                                    ws.update_cell(cell.row, 5, json.dumps(st.session_state.vales_historial[idx_vale]))
                                
                                st.success(f"📊 Conciliación del folio {v_select} inyectada en la nube con éxito.")
                                time.sleep(0.5)
                                st.rerun()
                            except Exception as e:
                                st.error(f"❌ Error en los canales de comunicación de la nube: {e}")

                # --- REPORTE COMPLETO DINÁMICO DE CONSUMO HISTÓRICO ---
                st.divider()
                st.markdown("### 📈 Reporte Dinámico Global de Consumo Real e Histórico")
                
                datos_reporte_completo = []
                for v in st.session_state.vales_historial:
                    for mat in v["Materiales"]:
                        datos_reporte_completo.append({
                            "Vale": v["Folio"], "Fecha": v["Fecha"], "Brigada": v["Brigada"],
                            "Del. Programada": v["Delegacion"], "UTB Programada": v["UTB"],
                            "Del. Real (Aplicada)": v.get("Delegacion_Real", v["Delegacion"]),
                            "UTB Real (Aplicada)": v.get("UTB_Real", v["UTB"]),
                            "Material": mat["Material"], "Entregado": mat["Cantidad"],
                            "Utilizado": mat.get("Utilizado", 0), "Devuelto": mat.get("Devuelto", 0), "Unidad": mat["Unidad"]
                        })
                
                df_reporte_base = pd.DataFrame(datos_reporte_completo)
                
                c_f1, c_f2, c_f3 = st.columns(3)
                with c_f1:
                    fechas_dispo = sorted(df_reporte_base["Fecha"].unique()) if not df_reporte_base.empty else []
                    fecha_filtro = st.selectbox("1. Filtrar por Fecha Determinante:", ["TODAS"] + fechas_dispo, key="sb_fecha_rep_fijo")
                
                df_filtrado = df_reporte_base if fecha_filtro == "TODAS" else df_reporte_base[df_reporte_base["Fecha"] == fecha_filtro]
                
                if "sb_filtrar_delegacion_rep" not in st.session_state:
                    st.session_state.sb_filtrar_delegacion_rep = "TODAS"
                if "sb_filtrar_utb_rep" not in st.session_state:
                    st.session_state.sb_filtrar_utb_rep = "TODAS"

                rep_del_current = st.session_state.sb_filtrar_delegacion_rep
                rep_utb_current = st.session_state.sb_filtrar_utb_rep

                if rep_utb_current != "TODAS" and not df_territorial.empty:
                    delegaciones_dispo = ["TODAS"] + sorted(df_territorial[df_territorial['UTB'] == rep_utb_current]['DELEGACION'].unique().tolist())
                else:
                    delegaciones_dispo = ["TODAS"] + DELEGACIONES_TOLUCA

                if rep_del_current != "TODAS" and not df_territorial.empty:
                    utbs_dispo = ["TODAS"] + sorted(df_territorial[df_territorial['DELEGACION'] == rep_del_current]['UTB'].unique().tolist())
                else:
                    utbs_dispo = ["TODAS"] + (sorted(df_territorial['UTB'].unique().tolist()) if not df_territorial.empty else [])

                idx_del = delegaciones_dispo.index(rep_del_current) if rep_del_current in delegaciones_dispo else 0
                idx_utb = utbs_dispo.index(rep_utb_current) if rep_utb_current in utbs_dispo else 0
                    
                with c_f2:
                    del_filtro = st.selectbox("2. Filtrar por Delegación Real Aplicada:", delegaciones_dispo, index=idx_del, key="sb_filtrar_delegacion_rep")
                with c_f3:
                    utb_filtro = st.selectbox("3. Filtrar por UTB Real Aplicada:", utbs_dispo, index=idx_utb, key="sb_filtrar_utb_rep")
                
                if not df_filtrado.empty:
                    if del_filtro != "TODAS":
                        df_filtrado = df_filtrado[df_filtrado["Del. Real (Aplicada)"] == del_filtro]
                    if utb_filtro != "TODAS":
                        df_filtrado = df_filtrado[df_filtrado["UTB Real (Aplicada)"] == utb_filtro]
                
                st.markdown("#### **Informe Resultante de Material en Calle**")
                st.dataframe(df_filtrado[["Vale", "Fecha", "Brigada", "Del. Programada", "Del. Real (Aplicada)", "UTB Real (Aplicada)", "Material", "Entregado", "Utilizado", "Devuelto", "Unidad"]], use_container_width=True, hide_index=True)
                
                st.markdown("**📊 Totales del Filtro de Aplicación Física:**")
                c_m1, c_m2, c_m3 = st.columns(3)
                c_m1.metric("📦 Total Entregado", f"{int(df_filtrado['Entregado'].sum()) if not df_filtrado.empty else 0} pzas")
                c_m2.metric("✅ Total Utilizado", f"{int(df_filtrado['Utilizado'].sum()) if not df_filtrado.empty else 0} pzas")
                c_m3.metric("🔄 Total Devuelto", f"{int(df_filtrado['Devuelto'].sum()) if not df_filtrado.empty else 0} pzas")
                
                output_rep = io.BytesIO()
                with pd.ExcelWriter(output_rep, engine='openpyxl') as writer:
                    df_filtrado.to_excel(writer, index=False, sheet_name='Reporte_Consumos_DAP')
                st.download_button(label="📄 DESCARGAR REPORTE FILTRADO EN EXCEL", data=output_rep.getvalue(), file_name=f"Reporte_Consumos_DAP_{pd.Timestamp.now().strftime('%d-%m-%Y')}.xlsx", use_container_width=True, key="btn_download_rep_fijo")

        # ==========================================
        # --- PESTAÑA 4: GESTIÓN ALMACÉN COMPLETA (PIN DAP-2026) ---
        # ==========================================
        with tab_admin:
            df_inv = st.session_state.db_inventario
            st.subheader("⚙️ Panel de Control de Existencias")
            if not st.session_state.admin_auth:
                st.warning("🔒 Este apartado requiere clave de acceso de Almacén.")
                pass_in = st.text_input("🔑 Ingrese PIN de Almacén:", type="password")
                if st.button("🔓 Conceder Acceso"):
                    if pass_in == config.PIN_ALMACEN:
                        st.session_state.admin_auth = True
                        st.rerun()
                    else:
                        st.error("❌ Contraseña incorrecta para el área DAP.")
            else:
                st.success("✅ Acceso de Administrador de Almacén Concedido")
                if st.button("🔒 Bloquear Panel de Gestión", type="secondary"):
                    st.session_state.admin_auth = False
                    st.rerun()
                
                st.divider()
                
                with st.expander("📥 Aumentar Existencias Físicas (Abastecimiento)", expanded=False):
                    m_in = st.selectbox("Seleccione Material a Abastecer:", df_inv['Material'].tolist(), key="sb_abastecer")
                    stock_previo = df_inv.loc[df_inv['Material'] == m_in, 'Stock'].values[0]
                    unidad_medida = df_inv.loc[df_inv['Material'] == m_in, 'Unidad'].values[0]
                    
                    with st.container(border=True):
                        st.markdown(f"📋 **Material:** {m_in} | **Cantidad Actual:** `{stock_previo} {unidad_medida}`")
                        c_in = st.number_input("Cantidad a ingresar:", min_value=1, step=1, value=1, key="cant_ingresar_live")
                        
                        if st.button("✅ ACTUALIZAR STOCK", use_container_width=True, key="btn_actualizar_stock_live"):
                            idx = df_inv[df_inv['Material'] == m_in].index[0]
                            st.session_state.db_inventario.at[idx, 'Stock'] += c_in
                            try:
                                data_manager.guardar_inventario(st.session_state.db_inventario)
                                st.success(f"📦 Entrada registrada correctamente.")
                            except Exception as e:
                                st.error(f"❌ Error al guardar archivo: {e}")
                            st.rerun()
                
                with st.expander("➕ Dar de Alta Nuevo Ítem en el Catálogo Oficial", expanded=False):
                    with st.container(border=True):
                        nuevo_nombre = st.text_input("Nombre del Material / Insumo:", placeholder="Ej: FOTOCELDA MULTIVOLTAJE").upper().strip()
                        c1, c2 = st.columns(2)
                        nueva_unidad = c1.selectbox("Unidad de Medida:", ["Piezas", "Metros", "Kilos", "Cajas"])
                        stock_inicial_item = c2.number_input("Stock Inicial Físico:", min_value=0, value=100, step=1)
                        minimo_alerta = c2.number_input("Stock Mínimo (Alerta Crítica):", min_value=1, value=10, step=1)
                        
                        if st.button("🚀 REGISTRAR NUEVO MATERIAL INMUEBLE", use_container_width=True):
                            if nuevo_nombre and nuevo_nombre not in df_inv['Material'].tolist():
                                nuevo_registro = {"ID": f"MAT-{len(df_inv)+1}", "Material": nuevo_nombre, "Stock": stock_inicial_item, "Min": minimo_alerta, "Unidad": nueva_unidad}
                                st.session_state.db_inventario = pd.concat([st.session_state.db_inventario, pd.DataFrame([nuevo_registro])], ignore_index=True)
                                try:
                                    data_manager.guardar_inventario(st.session_state.db_inventario)
                                    st.success("✅ Alta registrada.")
                                except Exception as e:
                                    st.error(f"❌ Error al guardar archivo: {e}")
                                time.sleep(0.5)
                                st.rerun()
                            else:
                                st.warning("⚠️ Nombre vacío o el material ya existe.")

                with st.expander("🗑️ Eliminar Ítem del Catálogo (Baja Definitiva)", expanded=False):
                    with st.container(border=True):
                        mat_a_eliminar = st.selectbox("Seleccione el Ítem a eliminar:", df_inv['Material'].tolist(), key="del_item_catalog")
                        check_seguro_item = st.checkbox("Confirmar que deseo borrar de forma permanente este concepto", key="chk_seguro_item_del")
                        
                        if st.button("💥 ELIMINAR CONCEPTO DEL INVENTARIO", use_container_width=True, type="secondary", disabled=not check_seguro_item):
                            st.session_state.db_inventario = st.session_state.db_inventario[st.session_state.db_inventario['Material'] != mat_a_eliminar].reset_index(drop=True)
                            try:
                                data_manager.guardar_inventario(st.session_state.db_inventario)
                                st.success("🗑️ El ítem ha sido borrado.")
                            except Exception as e:
                                st.error(f"❌ Error al guardar archivo: {e}")
                            time.sleep(0.5)
                            st.rerun()
                
                st.write("")
                st.markdown("---")
                st.subheader("🔒 Bóveda de Vales Emitidos (Historial Antirrobos)")
                
                if not st.session_state.vales_historial:
                    st.info("📂 La bóveda se encuentra vacía.")
                else:
                    tabla_boveda = []
                    for v in st.session_state.vales_historial:
                        tabla_boveda.append({"Folio": v["Folio"], "Fecha/Hora": v.get("FechaHora", "N/A"), "Brigada": v["Brigada"], "Total Insumos": len(v.get("Materiales", []))})
                    df_bov_vales = pd.DataFrame(tabla_boveda)
                    
                    folio_select = st.selectbox("Seleccione Folio para auditoría interna:", [v["Folio"] for v in st.session_state.vales_historial], key="sb_auditoria_boveda")
                    
                    def resaltar_vale_admin(row):
                        color = '#d1e7dd' if row['Folio'] == folio_select else ''
                        return [f'background-color: {color}'] * len(row)
                        
                    st.dataframe(df_bov_vales.style.apply(resaltar_vale_admin, axis=1), use_container_width=True, hide_index=True)
                    
                    vale_auditado = next(item for item in st.session_state.vales_historial if item["Folio"] == folio_select)
                    
                    with st.container(border=True):
                        st.markdown(f"### 📄 Expediente: {vale_auditado['Folio']}")
                        st.write(f"**Asignado a:** {vale_auditado['Brigada']} | **Ubicación:** {vale_auditado.get('Delegacion')} - {vale_auditado.get('UTB')}")
                        df_mat_auditoria = pd.DataFrame(vale_auditado.get("Materiales", []))
                        st.dataframe(df_mat_auditoria, use_container_width=True, hide_index=True)
                        
                        # --- RECONSTRUCCIÓN DE PDF EN VIVO DESDE EL JSON DEL HISTORIAL ---
                        try:
                            from fpdf import FPDF
                            pdf_h = FPDF()
                            pdf_h.add_page()
                            
                            pdf_h.set_font("Arial", 'B', 14)
                            pdf_h.cell(0, 10, "AYUNTAMIENTO DE TOLUCA", ln=True, align='C')
                            pdf_h.set_font("Arial", 'B', 12)
                            pdf_h.cell(0, 8, "DIRECCION DE ALUMBRADO PUBLICO", ln=True, align='C')
                            pdf_h.cell(0, 8, f"VALE OFICIAL DE SALIDA: {vale_auditado['Folio']}", ln=True, align='C')
                            pdf_h.ln(8)
                            
                            pdf_h.set_font("Arial", '', 10)
                            pdf_h.cell(0, 6, f"Fecha y Hora de Emision: {vale_auditado.get('FechaHora', vale_auditado['Fecha'])}", ln=True)
                            pdf_h.cell(0, 6, f"Ubicacion Geografica: {vale_auditado.get('Delegacion')} - {vale_auditado.get('UTB')}", ln=True)
                            pdf_h.set_font("Arial", 'B', 10)
                            pdf_h.cell(0, 6, f"UNIDAD / BRIGADA DESTINO: {vale_auditado['Brigada'].upper()}", ln=True)
                            pdf_h.ln(4)
                            
                            pdf_h.set_fill_color(230, 235, 240)
                            pdf_h.set_font("Arial", 'B', 10)
                            pdf_h.cell(110, 8, " Descripcion del Material / Insumo", 1, 0, 'L', True)
                            pdf_h.cell(36, 8, "Entregado", 1, 0, 'C', True)
                            pdf_h.cell(22, 8, "Utilizado", 1, 0, 'C', True)
                            pdf_h.cell(22, 8, "Devuelto", 1, 1, 'C', True)
                            
                            pdf_h.set_font("Arial", '', 10)
                            for it in vale_auditado.get("Materiales", []):
                                pdf_h.cell(110, 8, f" {str(it['Material'])}", 1, 0, 'L')
                                pdf_h.cell(36, 8, f"{it['Cantidad']} {it['Unidad']}", 1, 0, 'C')
                                pdf_h.cell(22, 8, f"{it.get('Utilizado', 0)}", 1, 0, 'C')
                                pdf_h.cell(22, 8, f"{it.get('Devuelto', 0)}", 1, 1, 'C')
                            pdf_h.ln(6)
                            
                            pdf_h.set_font("Arial", 'B', 10)
                            pdf_h.cell(0, 6, "Observaciones del Responsable de Almacen (Sistema):", ln=True)
                            pdf_h.set_font("Arial", 'I', 9)
                            msg_obs_h = vale_auditado.get("Observaciones", "Sin observaciones")
                            pdf_h.multi_cell(0, 5, msg_obs_h.encode('latin-1', 'replace').decode('latin-1'), 1)
                            pdf_h.ln(4)
                            
                            pdf_h.set_font("Arial", 'B', 10)
                            pdf_h.cell(0, 6, "Observaciones de la Brigada al Recibir (Llenar en Fisico a Mano):", ln=True)
                            pdf_h.cell(0, 15, "", 1, ln=True)
                            pdf_h.ln(10)
                            
                            pdf_h.set_font("Arial", 'I', 9)
                            pdf_h.multi_cell(0, 5, LEYENDA_OFICIAL.encode('latin-1', 'replace').decode('latin-1'), align='C')
                            pdf_h.ln(12)
                            
                            y_pos_f = pdf_h.get_y()
                            pdf_h.set_font("Arial", 'B', 8)
                            pdf_h.set_xy(15, y_pos_f)
                            pdf_h.cell(75, 4, "_____________________________________", ln=False, align='C')
                            pdf_h.set_xy(15, y_pos_f + 4)
                            pdf_h.cell(75, 4, "RESPONSABLE DE ENTREGA DE MATERIAL", ln=False, align='C')
                            pdf_h.set_xy(15, y_pos_f + 8)
                            pdf_h.set_font("Arial", '', 7)
                            pdf_h.cell(75, 4, "(Firma y Sello de Almacen DAP)", ln=False, align='C')
                            
                            pdf_h.set_font("Arial", 'B', 8)
                            pdf_h.set_xy(115, y_pos_f)
                            pdf_h.cell(75, 4, "_____________________________________", ln=False, align='C')
                            pdf_h.set_xy(115, y_pos_f + 4)
                            pdf_h.cell(75, 4, "RESPONSABLE QUE RECIBE MATERIAL", ln=False, align='C')
                            pdf_h.set_xy(115, y_pos_f + 8)
                            pdf_h.set_font("Arial", '', 7)
                            pdf_h.cell(75, 4, f"({vale_auditado['Brigada'].upper()})", ln=False, align='C')
                            
                            pdf_output_h = pdf_h.output(dest='S')
                            pdf_bytes_h = pdf_output_h.encode('latin-1', 'replace') if isinstance(pdf_output_h, str) else pdf_output_h
                            
                            st.download_button(
                                label=f"📄 RECONSTRUIR Y DESCARGAR PDF ({vale_auditado['Folio']})",
                                data=pdf_bytes_h,
                                file_name=f"Vale_Oficial_Salida_{vale_auditado['Folio']}.pdf",
                                mime="application/pdf",
                                use_container_width=True,
                                key=f"dl_h_{vale_auditado['Folio']}"
                            )
                            st.write("")
                        except Exception as e_pdf:
                            st.error(f"Error al reconstruir reporte PDF: {e_pdf}")
                                
                        check_seguro = st.checkbox(f"Confirmar destrucción física del Folio {folio_select}", key=f"chk_del_{folio_select}")
                        if st.button(f"🔥 ELIMINAR VALE {folio_select} PERMANENTEMENTE", use_container_width=True, type="secondary", disabled=not check_seguro):
                            try:
                                # --- 1. ROLLBACK DE MATERIALES: DEVOLVER EL NETO AL ALMACÉN ---
                                for mat_item in vale_auditado.get("Materiales", []):
                                    m_name = mat_item["Material"]
                                    cant_a_devolver = mat_item["Cantidad"] - mat_item.get("Devuelto", 0)
                                    if cant_a_devolver > 0:
                                        if m_name in st.session_state.db_inventario['Material'].tolist():
                                            idx_inv = st.session_state.db_inventario[st.session_state.db_inventario['Material'] == m_name].index[0]
                                            st.session_state.db_inventario.at[idx_inv, 'Stock'] += cant_a_devolver
                                
                                # Guardar inventario actualizado en la nube inmediatamente
                                data_manager.guardar_inventario(st.session_state.db_inventario)

                                # --- 2. DESTRUCCIÓN DEL REGISTRO EN LA NUBE ---
                                id_reg_borrar = f"SF6-VAL-{folio_select}"
                                cell = ws.find(id_reg_borrar, in_column=1)
                                if cell:
                                    ws.delete_rows(cell.row)
                                
                                st.session_state.vales_historial = [item for item in st.session_state.vales_historial if item["Folio"] != folio_select]
                                st.warning("Vale purgado de la bóveda permanente.")
                                time.sleep(0.5)
                                st.rerun()
                            except Exception as e:
                                st.error(f"Error al eliminar de Google Sheets: {e}")

                st.write("")
                st.markdown("---")
                st.subheader("🚨 Zona de Despliegue Oficial (Pasar a Producción)")
                st.caption("Usa este panel SOLO cuando termine la presentación para formatear las bases y arrancar en ceros.")
                
                pin_produccion = st.text_input("🔑 Ingrese PIN Maestro de Despliegue:", type="password", key="pin_despliegue_oficial")
                check_reset_total = st.checkbox("Confirmar que deseo borrar toda la simulación y dejar stock en 0", key="chk_reset_total_prod")
                
                if st.button("💥 PURGAR SIMULACIÓN Y ARRANCAR EN CEROS", use_container_width=True, type="secondary", disabled=not (check_reset_total and pin_produccion == config.PIN_MAESTRO)):
                    try:
                        # Purgar en Sheets de forma segura por coordenadas matriciales inversas
                        filas_limpieza = ws.get_all_values()
                        if len(filas_limpieza) > 1:
                            # Recorremos la hoja desde la última fila hacia arriba para no alterar los índices
                            for idx_real in range(len(filas_limpieza), 1, -1):
                                row = filas_limpieza[idx_real - 1]
                                reg_id = str(row[0]).strip()
                                if reg_id.startswith("SF6-VAL-"):
                                    ws.delete_rows(idx_real)
                        
                        data_manager.reiniciar_sistema()
                        
                        df_base = pd.DataFrame(STOCK_INICIAL)
                        df_base['Stock'] = 0  
                        df_base['Min'] = 0    
                        data_manager.guardar_inventario(df_base)
                        
                        st.session_state.db_inventario = df_base
                        st.session_state.vales_historial = []
                        st.session_state.maestro_auth = False 
                        
                        st.success("🎉 ¡Formateo Exitoso! El sistema ha sido desplegado a producción real en ceros.")
                        time.sleep(1.5)
                        st.rerun()
                    except Exception as e:
                        st.error(f"⚠️ Error en formateo: {e}")
